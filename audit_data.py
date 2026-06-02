import openpyxl
import pandas as pd

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"

print("=" * 80)
print("审慎审查原始Excel数据...")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path)

print(f"\n工作表: 清单汇总")
ws = wb['清单汇总']

print(f"\n第1行（表头）:")
for col in range(1, min(25, ws.max_column) + 1):
    val = ws.cell(row=1, column=col).value
    print(f"  列{col}: {val}")

print(f"\n第2-5行数据预览:")
for row in range(2, min(6, ws.max_row + 1)):
    print(f"\n--- 第{row}行 ---")
    for col in range(1, min(25, ws.max_column) + 1):
        val = ws.cell(row=row, column=col).value
        if val:
            print(f"  列{col}: {val}")

print(f"\n\n检查单价相关列:")
df = pd.read_excel(excel_path, sheet_name='清单汇总', header=0)
print(f"\n列名:")
for i, col in enumerate(df.columns):
    print(f"  {i+1}. {col}")

print(f"\n前10行数据预览（关键列）:")
key_cols = ['项目编码', '项目名称', '合价（元）_合同合价（元）_J', '合价（元）_结算合价（元）_K', '合价（元）_审核合价（元）_L']
for col in df.columns:
    if '单价' in col:
        key_cols.append(col)

print("\n选择关键列进行预览:")
print(df[key_cols].head(20))

print(f"\n\n检查审核合价的统计:")
print(f"审核合价列非空值数量: {df['合价（元）_审核合价（元）_L'].notna().sum()}")
print(f"审核合价列值 > 0 的数量: {(df['合价（元）_审核合价（元）_L'] > 0).sum()}")
print(f"审核合价列值汇总: {df['合价（元）_审核合价（元）_L'].sum():,.2f}")

print(f"\n\n检查单价列:")
price_cols = [col for col in df.columns if '单价' in col]
print(f"单价相关列: {price_cols}")
print(f"\n单价列的数据:")
print(df[price_cols].head(20))

print("=" * 80)
