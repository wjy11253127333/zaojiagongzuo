import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import shutil

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"
temp_path = r"c:\Users\aa\Desktop\造价清单文件标准化\temp_copy_fixed.xlsx"
output_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx"

print("=" * 80)
print("修复并更新成本分析文件...")
print("=" * 80)

# 复制原文件
shutil.copy2(excel_path, temp_path)
wb = openpyxl.load_workbook(temp_path)


# ==============================
# 步骤1: 直接从原始工作表读取完整数据
# ==============================
print("\n步骤1: 从原始工作表读取完整数据")

def read_complete_data(sheet_name):
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=2)
    print(f"  {sheet_name} 原始数据行数: {len(df)}")
    df_clean = df[df['项目编码'].notna() & (df['项目编码'] != '')].copy()
    print(f"  有效数据行数（有项目编码）: {len(df_clean)}")
    return df_clean

df_com = read_complete_data('表-08 分部分项工程和单价措施项目清单-综合楼')
df_edu = read_complete_data('表-08 分部分项工程和单价措施项目清单-教学楼')

# 添加业态标识
df_com['业态'] = '综合楼'
df_edu['业态'] = '教学楼'

# 合并数据
df_all = pd.concat([df_com, df_edu], ignore_index=True)
print(f"\n合并后总有效数据行数: {len(df_all)}")


# ==============================
# 步骤2: 计算单价和验证合价
# ==============================
print("\n步骤2: 计算单价并验证合价")

def safe_float(val):
    try:
        return float(val) if pd.notna(val) else 0.0
    except:
        return 0.0

def calculate_unit_price(qty, total):
    qty = safe_float(qty)
    total = safe_float(total)
    if qty > 0 and pd.notna(total) and total > 0:
        return total / qty
    return None

df_all['合同单价'] = df_all['单价（元）_投标单价（元）_P0']
df_all['送审单价'] = df_all['单价（元）_预算备案单价（元）_P1']
df_all['审核单价'] = df_all.apply(lambda x: calculate_unit_price(x['工程量_审核工程量_C'], x['合价（元）_审核合价（元）_L']), axis=1)

df_all['合同合价'] = df_all['合价（元）_合同合价（元）_J'].apply(safe_float)
df_all['结算合价'] = df_all['合价（元）_结算合价（元）_K'].apply(safe_float)
df_all['审核合价'] = df_all['合价（元）_审核合价（元）_L'].apply(safe_float)
df_all['合同工程量'] = df_all['工程量_合同工程量_A'].apply(safe_float)
df_all['送审工程量'] = df_all['工程量_送审工程量_B'].apply(safe_float)
df_all['审核工程量'] = df_all['工程量_审核工程量_C'].apply(safe_float)

print("\n数据预览（前10行）:")
print(df_all[['项目编码', '项目名称', '合同单价', '送审单价', '审核单价', '合同合价', '结算合价', '审核合价']].head(10))

print("\n数据统计:")
print(f"  合同单价有值: {df_all['合同单价'].notna().sum()}")
print(f"  送审单价有值: {df_all['送审单价'].notna().sum()}")
print(f"  审核单价有值: {df_all['审核单价'].notna().sum()}")
print(f"  审核合价有值: {(df_all['审核合价'] > 0).sum()}")
print(f"  审核合价总和: {df_all['审核合价'].sum():,.2f}")


# ==============================
# 步骤3: 删除旧的成本分析工作表（如果存在）
# ==============================
print("\n步骤3: 重建成本分析工作表")

for sheet_name in ['成本分析_汇总', '成本分析_综合', '成本分析_明细']:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"  删除旧工作表: {sheet_name}")


# ==============================
# 步骤4: 创建 Sheet 1 - 成本分析_汇总
# ==============================
ws1 = wb.create_sheet('成本分析_汇总', 0)

# 标题
title1 = ws1.cell(row=1, column=1, value='分部分项工程成本汇总分析')
title1.font = Font(bold=True, size=14)
ws1.merge_cells('A1:K1')

ws1.cell(row=2, column=1, value='分析日期：2026-05-27')
ws1.cell(row=2, column=7, value='单位：元')

# 表头
headers1 = ['序号', '业态', '项目数量', '合同合价', '结算合价', '审核合价', '偏差（结算-合同）', '偏差率（%）', '成本降低额', '成本降低率（%）']
header_fill1 = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font1 = Font(bold=True, size=11, color='FFFFFF')

for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font1
    cell.fill = header_fill1
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 汇总数据
summary_data = []
for category in ['综合楼', '教学楼']:
    df_cat = df_all[df_all['业态'] == category]
    summary_data.append({
        '业态': category,
        '项目数量': len(df_cat),
        '合同合价': df_cat['合同合价'].sum(),
        '结算合价': df_cat['结算合价'].sum(),
        '审核合价': df_cat['审核合价'].sum()
    })

# 填充数据
row_num1 = 5
total_contract = 0
total_settle = 0
total_audit = 0
total_count = 0

for i, data in enumerate(summary_data, 1):
    ws1.cell(row=row_num1, column=1, value=i)
    ws1.cell(row=row_num1, column=2, value=data['业态'])
    ws1.cell(row=row_num1, column=3, value=data['项目数量'])
    ws1.cell(row=row_num1, column=4, value=data['合同合价'])
    ws1.cell(row=row_num1, column=5, value=data['结算合价'])
    ws1.cell(row=row_num1, column=6, value=data['审核合价'])
    ws1.cell(row=row_num1, column=7, value=data['结算合价'] - data['合同合价'])
    dev_rate = (data['结算合价'] - data['合同合价']) / data['合同合价'] * 100 if data['合同合价'] > 0 else 0
    ws1.cell(row=row_num1, column=8, value=dev_rate / 100)
    red_amt = data['合同合价'] - data['审核合价'] if data['审核合价'] > 0 else None
    ws1.cell(row=row_num1, column=9, value=red_amt if red_amt is not None else 0)
    red_rate = red_amt / data['合同合价'] * 100 if data['审核合价'] > 0 and data['合同合价'] > 0 else 0
    ws1.cell(row=row_num1, column=10, value=red_rate / 100 if red_amt is not None else 0)

    total_contract += data['合同合价']
    total_settle += data['结算合价']
    total_audit += data['审核合价']
    total_count += data['项目数量']
    row_num1 += 1

# 合计行
ws1.cell(row=row_num1, column=2, value='合计')
ws1.cell(row=row_num1, column=3, value=total_count)
ws1.cell(row=row_num1, column=4, value=total_contract)
ws1.cell(row=row_num1, column=5, value=total_settle)
ws1.cell(row=row_num1, column=6, value=total_audit)
ws1.cell(row=row_num1, column=7, value=total_settle - total_contract)
total_dev_rate = (total_settle - total_contract) / total_contract * 100 if total_contract > 0 else 0
ws1.cell(row=row_num1, column=8, value=total_dev_rate / 100)
total_red_amt = total_contract - total_audit if total_audit > 0 else 0
ws1.cell(row=row_num1, column=9, value=total_red_amt)
total_red_rate = total_red_amt / total_contract * 100 if total_audit > 0 and total_contract > 0 else 0
ws1.cell(row=row_num1, column=10, value=total_red_rate / 100)

# 格式设置
for col in range(4, 11):
    for r in range(5, row_num1 + 1):
        cell = ws1.cell(row=r, column=col)
        if isinstance(cell.value, (int, float)):
            if col in [8, 10]:
                cell.number_format = '0.00%'
            else:
                cell.number_format = '#,##0.00'

for col in range(1, 11):
    ws1.cell(row=row_num1, column=col).font = Font(bold=True)
    ws1.cell(row=row_num1, column=col).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

ws1.column_dimensions['A'].width = 8
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 12
ws1.column_dimensions['I'].width = 18
ws1.column_dimensions['J'].width = 15

print("  ✅ Sheet1: 成本分析_汇总 已创建")


# ==============================
# 步骤5: 创建 Sheet 2 - 成本分析_综合
# ==============================
ws2 = wb.create_sheet('成本分析_综合', 1)

title2 = ws2.cell(row=1, column=1, value='综合成本汇总分析')
title2.font = Font(bold=True, size=14)
ws2.merge_cells('A1:E1')

ws2.cell(row=3, column=1, value='一、成本总体概况')
ws2.cell(row=3, column=1).font = Font(bold=True, size=12)

ws2.cell(row=4, column=1, value='合同总成本（元）')
ws2.cell(row=4, column=2, value=total_contract)
ws2.cell(row=4, column=2).number_format = '#,##0.00'

ws2.cell(row=5, column=1, value='送审总成本（元）')
ws2.cell(row=5, column=2, value=total_settle)
ws2.cell(row=5, column=2).number_format = '#,##0.00'

ws2.cell(row=6, column=1, value='审核总成本（元）')
ws2.cell(row=6, column=2, value=total_audit)
ws2.cell(row=6, column=2).number_format = '#,##0.00'

ws2.cell(row=7, column=1, value='送审与合同偏差（元）')
ws2.cell(row=7, column=2, value=total_settle - total_contract)
ws2.cell(row=7, column=2).number_format = '#,##0.00'

ws2.cell(row=8, column=1, value='送审与合同偏差率（%）')
ws2.cell(row=8, column=2, value=total_dev_rate / 100)
ws2.cell(row=8, column=2).number_format = '0.00%'

ws2.cell(row=10, column=1, value='二、成本降低效果分析（基于审核成本）')
ws2.cell(row=10, column=1).font = Font(bold=True, size=12)

ws2.cell(row=11, column=1, value='项目施工成本降低额（元）')
ws2.cell(row=11, column=2, value=total_red_amt)
ws2.cell(row=11, column=2).number_format = '#,##0.00'

ws2.cell(row=12, column=1, value='项目施工成本降低率（%）')
ws2.cell(row=12, column=2, value=total_red_rate / 100)
ws2.cell(row=12, column=2).number_format = '0.00%'

ws2.cell(row=14, column=1, value='三、按业态成本占比')
ws2.cell(row=14, column=1).font = Font(bold=True, size=12)

row2 = 15
for data in summary_data:
    cat_name = data['业态']
    proportion = (data['合同合价'] / total_contract * 100) if total_contract > 0 else 0
    ws2.cell(row=row2, column=1, value=f'{cat_name}合同占比（%）')
    ws2.cell(row=row2, column=2, value=proportion / 100)
    ws2.cell(row=row2, column=2).number_format = '0.00%'
    ws2.cell(row=row2, column=3, value=f'{cat_name}合同金额（元）')
    ws2.cell(row=row2, column=4, value=data['合同合价'])
    ws2.cell(row=row2, column=4).number_format = '#,##0.00'
    row2 += 1

ws2.cell(row=row2 + 1, column=1, value='四、成本管理方法论说明')
ws2.cell(row=row2 + 1, column=1).font = Font(bold=True, size=12)

ws2.cell(row=row2 + 2, column=1, value='1. 本分析基于《施工成本管理》方法论框架')
ws2.cell(row=row2 + 3, column=1, value='2. 采用"三算对比"方法：合同成本、送审成本、审核成本')
ws2.cell(row=row2 + 4, column=1, value='3. 成本降低率 = (合同成本 - 审核成本) / 合同成本 × 100%')
ws2.cell(row=row2 + 5, column=1, value='4. 偏差率 = (送审成本 - 合同成本) / 合同成本 × 100%')

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 20

print("  ✅ Sheet2: 成本分析_综合 已创建")


# ==============================
# 步骤6: 创建 Sheet 3 - 成本分析_明细
# ==============================
ws3 = wb.create_sheet('成本分析_明细', 2)

title3 = ws3.cell(row=1, column=1, value='分部分项工程成本明细数据（含单价）')
title3.font = Font(bold=True, size=14)
ws3.merge_cells('A1:N1')

headers3 = [
    '序号', '业态', '项目编码', '项目名称', '计量单位',
    '合同工程量', '送审工程量', '审核工程量',
    '合同单价', '送审单价', '审核单价',
    '合同合价', '结算合价', '审核合价'
]

header_fill3 = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font3 = Font(bold=True, size=11, color='FFFFFF')

for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = header_font3
    cell.fill = header_fill3
    cell.alignment = Alignment(horizontal='center', vertical='center')

row_num3 = 4
for idx, row_data in df_all.iterrows():
    ws3.cell(row=row_num3, column=1, value=idx + 1)
    ws3.cell(row=row_num3, column=2, value=row_data['业态'])
    ws3.cell(row=row_num3, column=3, value=row_data['项目编码'])
    ws3.cell(row=row_num3, column=4, value=row_data['项目名称'])
    ws3.cell(row=row_num3, column=5, value=row_data['计量单位'])
    ws3.cell(row=row_num3, column=6, value=safe_float(row_data['合同工程量']))
    ws3.cell(row=row_num3, column=7, value=safe_float(row_data['送审工程量']))
    ws3.cell(row=row_num3, column=8, value=safe_float(row_data['审核工程量']))
    ws3.cell(row=row_num3, column=9, value=safe_float(row_data['合同单价']))
    ws3.cell(row=row_num3, column=10, value=safe_float(row_data['送审单价']))
    ws3.cell(row=row_num3, column=11, value=safe_float(row_data['审核单价']))
    ws3.cell(row=row_num3, column=12, value=safe_float(row_data['合同合价']))
    ws3.cell(row=row_num3, column=13, value=safe_float(row_data['结算合价']))
    ws3.cell(row=row_num3, column=14, value=safe_float(row_data['审核合价']))

    for col in range(6, 15):
        cell = ws3.cell(row=row_num3, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

    row_num3 += 1
    if row_num3 > 600:
        break

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 30
ws3.column_dimensions['E'].width = 10
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 14
ws3.column_dimensions['H'].width = 14
ws3.column_dimensions['I'].width = 14
ws3.column_dimensions['J'].width = 14
ws3.column_dimensions['K'].width = 14
ws3.column_dimensions['L'].width = 16
ws3.column_dimensions['M'].width = 16
ws3.column_dimensions['N'].width = 16

print("  ✅ Sheet3: 成本分析_明细 已创建")


# ==============================
# 保存文件
# ==============================
print("\n步骤7: 保存文件")

wb.save(temp_path)
shutil.copy2(temp_path, output_path)

print("=" * 80)
print(f"✅ 文件已成功修复并保存:")
print(f"   文件路径: {output_path}")
print(f"   处理后数据统计:")
print(f"     总有效项目数: {len(df_all)}")
print(f"     合同总成本: {total_contract:,.2f}")
print(f"     送审总成本: {total_settle:,.2f}")
print(f"     审核总成本: {total_audit:,.2f}")
print(f"   新增工作表:")
print(f"     - Sheet1: 成本分析_汇总（分部分项工程成本汇总）")
print(f"     - Sheet2: 成本分析_综合（综合成本汇总分析）")
print(f"     - Sheet3: 成本分析_明细（含单价列的详细数据）")
print("=" * 80)
