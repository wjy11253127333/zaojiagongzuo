from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy

# 文件路径
output_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx'
excel_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx'

print("=" * 80)
print("补充缺失的汇总公式")
print("=" * 80)

# 加载文件
wb = load_workbook(output_path, data_only=False)
ws_summary = wb['清单汇总']

# 加载原始文件获取教学楼的汇总行
wb_orig = load_workbook(excel_path, data_only=False)
ws_edu = wb_orig['表-08 分部分项工程和单价措施项目清单-教学楼']

print(f"\n1. 检查教学楼汇总行...")
print(f"   教学楼总行数: {ws_edu.max_row}")

# 找到教学楼的汇总行（最后几行）
edu_last_row = ws_edu.max_row
print(f"   教学楼最后一行: {edu_last_row}")

# 检查最后几行的公式
print(f"\n2. 检查教学楼最后几行的公式...")
for row in range(edu_last_row - 3, edu_last_row + 1):
    for col in range(1, 8):
        cell = ws_edu.cell(row=row, column=col)
        if cell.value is not None:
            print(f"   行{row} 列{get_column_letter(col)}: {cell.value}")

# 确定教学楼数据在汇总表中的最后一行
summary_edu_start = 270  # 教学楼数据起始行
summary_edu_end = summary_edu_start + (edu_last_row - 5)  # 教学楼数据行数
print(f"\n3. 汇总表中教学楼数据范围: 第{summary_edu_start}行 - 第{summary_edu_end}行")

# 检查汇总表的最后几行
print(f"\n4. 检查汇总表最后几行的公式...")
for row in range(summary_edu_end - 3, summary_edu_end + 1):
    for col in range(1, 8):
        cell = ws_summary.cell(row=row, column=col)
        if cell.value is not None:
            print(f"   行{row} 列{get_column_letter(col)}: {cell.value}")

# 复制教学楼最后几行的公式（如果有的话）
print(f"\n5. 补充缺失的汇总公式...")
missing_count = 0

for row in range(edu_last_row - 2, edu_last_row + 1):
    for col in range(1, ws_edu.max_column + 1):
        source_cell = ws_edu.cell(row=row, column=col)
        target_row = summary_edu_start + (row - 5)
        target_cell = ws_summary.cell(row=target_row, column=col)

        # 检查是否是公式
        if source_cell.data_type == 'f':
            if target_cell.data_type != 'f':
                # 补充公式
                target_cell.value = source_cell.value
                missing_count += 1
                print(f"   补充公式: {get_column_letter(col)}{target_row} = {source_cell.value}")

print(f"\n6. 补充完成: {missing_count} 个公式")

# 重新统计公式数
formula_count = 0
for row in range(1, ws_summary.max_row + 1):
    for col in range(1, ws_summary.max_column + 1):
        if ws_summary.cell(row=row, column=col).data_type == 'f':
            formula_count += 1

print(f"\n7. 最终统计:")
print(f"   汇总表公式总数: {formula_count}")
print(f"   预期公式数: 5784 (3003 + 2781)")

# 保存文件
print(f"\n8. 保存文件...")
temp_output = r'c:\Users\aa\Desktop\造价清单文件标准化\temp_copy.xlsx'
wb.save(temp_output)

import shutil
shutil.copy(temp_output, output_path)

print(f"   文件已保存到: {output_path}")

print("\n" + "=" * 80)
print("补充完成！")
print("=" * 80)
