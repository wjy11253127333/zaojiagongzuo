from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 文件路径
excel_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx'

print("=" * 80)
print("详细分析教学楼工作表公式分布")
print("=" * 80)

# 加载原始文件
wb = load_workbook(excel_path, data_only=False)
ws_edu = wb['表-08 分部分项工程和单价措施项目清单-教学楼']

print(f"\n教学楼工作表统计:")
print(f"  总行数: {ws_edu.max_row}")
print(f"  总列数: {ws_edu.max_column}")

# 统计每个区域的公式数
formula_rows = []
for row in range(1, ws_edu.max_row + 1):
    row_formulas = 0
    for col in range(1, ws_edu.max_column + 1):
        if ws_edu.cell(row=row, column=col).data_type == 'f':
            row_formulas += 1
    if row_formulas > 0:
        formula_rows.append((row, row_formulas))

print(f"\n包含公式的行数: {len(formula_rows)}")

# 显示前20行和后20行的公式分布
print(f"\n前20行公式分布:")
for row, count in formula_rows[:20]:
    first_formula = None
    for col in range(1, min(6, ws_edu.max_column + 1)):
        cell = ws_edu.cell(row=row, column=col)
        if cell.data_type == 'f':
            first_formula = f"{get_column_letter(col)}{row}: {cell.value}"
            break
    print(f"  第{row:3d}行: {count:2d}个公式  示例: {first_formula}")

print(f"\n后20行公式分布:")
for row, count in formula_rows[-20:]:
    first_formula = None
    for col in range(1, min(6, ws_edu.max_column + 1)):
        cell = ws_edu.cell(row=row, column=col)
        if cell.data_type == 'f':
            first_formula = f"{get_column_letter(col)}{row}: {cell.value}"
            break
    print(f"  第{row:3d}行: {count:2d}个公式  示例: {first_formula}")

# 检查前5行是否有公式
print(f"\n检查前5行（表头区域）:")
for row in range(1, 6):
    row_formulas = 0
    for col in range(1, ws_edu.max_column + 1):
        if ws_edu.cell(row=row, column=col).data_type == 'f':
            row_formulas += 1
            if row_formulas <= 3:  # 只显示前3个
                print(f"  {get_column_letter(col)}{row}: {ws_edu.cell(row=row, column=col).value}")
    if row_formulas > 0:
        print(f"  第{row}行共{row_formulas}个公式")

print(f"\n检查后5行（汇总区域）:")
for row in range(ws_edu.max_row - 4, ws_edu.max_row + 1):
    row_formulas = 0
    for col in range(1, ws_edu.max_column + 1):
        if ws_edu.cell(row=row, column=col).data_type == 'f':
            row_formulas += 1
            if row_formulas <= 3:  # 只显示前3个
                print(f"  {get_column_letter(col)}{row}: {ws_edu.cell(row=row, column=col).value}")
    if row_formulas > 0:
        print(f"  第{row}行共{row_formulas}个公式")

# 总公式数
total_formulas = sum(count for _, count in formula_rows)
print(f"\n教学楼总公式数: {total_formulas}")
print(f"预期: 2781")
print(f"差异: {2781 - total_formulas}")
