import openpyxl
import pandas as pd

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    print(f"工作表数量: {len(wb.sheetnames)}")
    print(f"工作表名称: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"工作表: {sheet_name}")
        print(f"数据范围: {ws.max_row} 行 × {ws.max_column} 列")
        print(f"{'='*80}\n")

        # 读取前20行数据
        print("数据预览 (前20行):")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            # 过滤掉全为None的行
            if any(cell is not None for cell in row):
                # 限制每行的列数显示
                display_row = row[:20]  # 只显示前20列
                print(f"行 {i}: {[str(cell)[:30] if cell else '' for cell in display_row]}")

        # 检查合并单元格
        if ws.merged_cells:
            print(f"\n合并单元格数量: {len(ws.merged_cells.ranges)}")

    # 使用pandas读取更多信息
    print("\n\n" + "="*80)
    print("使用pandas读取数据:")
    print("="*80)

    # 读取所有工作表
    xls = pd.ExcelFile(excel_path)
    for sheet in xls.sheet_names:
        print(f"\n工作表: {sheet}")
        df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
        print(f"形状: {df.shape}")
        print(f"前5行:\n{df.head()}")

except Exception as e:
    print(f"读取Excel时出错: {e}")
    import traceback
    traceback.print_exc()
