import openpyxl

output_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx"

print("=" * 80)
print("验证修复后的成本分析文件")
print("=" * 80)

wb = openpyxl.load_workbook(output_path)

print(f"\n当前工作表: {wb.sheetnames}")

# 验证 Sheet 1
print("\n【验证 Sheet1: 成本分析_汇总】")
ws1 = wb['成本分析_汇总']
print(f"  数据范围: {ws1.max_row} 行 × {ws1.max_column} 列")
print(f"  表头: {[ws1.cell(row=4, column=col).value for col in range(1, ws1.max_column + 1)]}")

print("\n  汇总数据:")
for row in range(5, ws1.max_row + 1):
    print(f"  行{row}:")
    for col in range(1, ws1.max_column + 1):
        val = ws1.cell(row=row, column=col).value
        print(f"    列{col}: {val}")

# 验证 Sheet 3
print("\n【验证 Sheet3: 成本分析_明细】")
ws3 = wb['成本分析_明细']
print(f"  数据范围: {ws3.max_row} 行 × {ws3.max_column} 列")
print(f"  表头: {[ws3.cell(row=3, column=col).value for col in range(1, ws3.max_column + 1)]}")

print("\n  前10行数据预览:")
for row in range(4, min(14, ws3.max_row + 1)):
    print(f"  行{row}:")
    print(f"    序号: {ws3.cell(row=row, column=1).value}")
    print(f"    业态: {ws3.cell(row=row, column=2).value}")
    print(f"    项目编码: {ws3.cell(row=row, column=3).value}")
    print(f"    项目名称: {ws3.cell(row=row, column=4).value}")
    print(f"    合同单价: {ws3.cell(row=row, column=9).value}")
    print(f"    送审单价: {ws3.cell(row=row, column=10).value}")
    print(f"    审核单价: {ws3.cell(row=row, column=11).value}")
    print(f"    合同合价: {ws3.cell(row=row, column=12).value}")
    print(f"    结算合价: {ws3.cell(row=row, column=13).value}")
    print(f"    审核合价: {ws3.cell(row=row, column=14).value}")

print("\n=" * 80)
print("✅ 验证完成！")
print("=" * 80)
