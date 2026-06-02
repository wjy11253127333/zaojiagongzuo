"""
测试数据生成器
生成一个模拟的、非标准的Excel清单文件，用于测试处理程序
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os


def create_test_excel(output_path: str):
    """
    创建一个包含多级表头、合并单元格的非标准Excel清单文件
    包含多个工作表，模拟真实场景
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # ===== 工作表1：教学楼地上清单（含多级表头+合并单元格）=====
    ws1 = wb.create_sheet("教学楼地上清单")
    build_sheet_1(ws1)

    # ===== 工作表2：教学楼地下清单 =====
    ws2 = wb.create_sheet("教学楼地下清单")
    build_sheet_2(ws2)

    # ===== 工作表3：宿舍楼地上清单（表头格式不同）=====
    ws3 = wb.create_sheet("宿舍楼地上清单")
    build_sheet_3(ws3)

    wb.save(output_path)
    print(f"✅ 测试文件已生成：{output_path}")
    print(f"   包含 3 个工作表，含多级表头和合并单元格")
    return output_path


def build_sheet_1(ws):
    """教学楼地上清单 - 含合并单元格的多级表头"""
    # 第1行：合并的标题行
    ws.merge_cells("A1:H1")
    ws["A1"] = "教学楼地上 - 工程量清单"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 第2-3行：多级表头（合并单元格）
    # 第2行
    ws.merge_cells("A2:A3")
    ws["A2"] = "业态"  # 实际文件中可能没有，测试自动填充

    ws.merge_cells("B2:B3")
    ws["B2"] = "项目名称"

    ws.merge_cells("C2:C3")
    ws["C2"] = "项目特征描述"

    ws.merge_cells("D2:D3")
    ws["D2"] = "单位"

    ws.merge_cells("E2:E3")
    ws["E2"] = "工程量"

    ws.merge_cells("F2:G2")  # 综合单价横跨两列（故意制造不规范）
    ws["F2"] = "综合单价"
    ws["F3"] = "单价(元)"
    ws["G3"] = "人工费"

    ws.merge_cells("H2:H3")
    ws["H2"] = "综合合价"

    # 注意：故意不设置"备注"列，测试自动新增

    # 数据行（第4行开始）
    data = [
        ["砌块墙", "加气混凝土砌块 200mm厚", "m3", 120.5, 580.0, 120.0, 69890.0],
        ["砼柱", "C30商品砼 矩形柱 泵送", "m3", 85.2, 750.0, 90.0, 63900.0],
        ["砼梁", "C30砼 有梁板 抗渗P6", "m3", 210.0, 720.0, 85.0, 151200.0],
        ["钢筋", "HRB400直径12mm 钢筋制作安装", "t", 15.8, 5200.0, 200.0, 82160.0],
        ["模板", "柱模板 铝模", "m2", 450.0, 85.0, 10.0, 38250.0],
        ["砂浆抹灰", "墙面水泥砂浆抹灰 20mm厚", "m2", 3200.0, 38.0, 5.0, 121600.0],
        ["防水工程", "屋面SBS改性沥青防水卷材", "m2", 1800.0, 65.0, 8.0, 117000.0],
        ["脚手架", "综合钢脚手架 地上部分", "m2", 6500.0, 28.0, 3.0, 182000.0],
        ["楼地面", "细石混凝土楼地面 50mm厚", "m2", 4800.0, 45.0, 5.0, 216000.0],
        ["门窗", "铝合金窗 推拉窗", "m2", 850.0, 680.0, 50.0, 578000.0],
    ]

    for i, row in enumerate(data, start=4):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    # 设置列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40


def build_sheet_2(ws):
    """教学楼地下清单 - 不同表头格式"""
    # 表头（2行，无合并单元格，测试多级解析）
    ws["A1"] = "类型"
    ws["B1"] = "清单名称"
    ws["C1"] = "项目特征"
    ws["D1"] = "计量"
    ws["D2"] = "单位"
    ws["E1"] = "工程量"
    ws["F1"] = "综合单价(元)"
    ws["G1"] = "合价(元)"

    # 注意：这里是2行表头但无合并，测试解析逻辑

    data = [
        ["砼基础", "C30砼 独立基础", "m3", 200.0, 680.0, 136000.0],
        ["钢筋", "HRB400 基础钢筋", "t", 25.0, 5100.0, 127500.0],
        ["防水", "地下室底板防水 SBS", "m2", 3500.0, 72.0, 252000.0],
        ["土方", "挖土方 一、二类土", "m3", 5000.0, 25.0, 125000.0],
    ]

    for i, row in enumerate(data, start=3):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)


def build_sheet_3(ws):
    """宿舍楼地上清单 - 另一表头格式"""
    # 单行表头，但字段名不完全标准
    headers = ["建筑类型", "分项名称", "特征", "单位", "数量", "单价", "合价", "说明"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)

    data = [
        ["宿舍楼地上", "砌体墙", "加气砼砌块 120mm", "m3", 80.0, 560.0, 44800.0, ""],
        ["宿舍楼地上", "砼板", "C30有梁板 泵送 抗渗", "m3", 150.0, 730.0, 109500.0, ""],
        ["宿舍楼地上", "钢筋", "HRB400 直径8-25mm", "t", 12.0, 5150.0, 61800.0, ""],
        ["宿舍楼地上", "屋面防水", "SBS改性沥青防水", "m2", 1200.0, 62.0, 74400.0, ""],
        ["宿舍楼地上", "外墙涂料", "真石漆外墙涂料", "m2", 3800.0, 85.0, 323000.0, ""],
        ["宿舍楼地上", "栏杆", "楼梯不锈钢栏杆", "m", 450.0, 180.0, 81000.0, ""],
    ]

    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)


if __name__ == "__main__":
    output = os.path.join(os.path.dirname(__file__), "..", "测试清单_非标准格式.xlsx")
    output = os.path.abspath(output)
    create_test_excel(output)
