"""
多级表头解析模块 v0.1.2
功能：将复杂的多级合并表头拆分为标准一级表头（单行单单元格）
核心改进：
- 横向展开的父-子合并单元格用下划线连接（如：工程数量_教学楼1#~3#）
- 从子到父处理合并单元格（min_row从高到低），确保多层嵌套正确
- 非展开列（无子表头）保持原样
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.cell.cell import MergedCell
from utils import normalize_header_text
from config import REQUIRED_HEADERS


class HeaderParser:
    """
    多级表头解析器
    将合并单元格的多行表头展开为单行标准表头

    处理逻辑：
    1. 逐行逐列读取非MergedCell的值，用下划线纵向拼接
    2. 从子到父处理合并单元格，将父值下划线前缀到已有的子值前
    3. 非展开列（无子表头）保持原样
    """

    def __init__(self, wb=None):
        self.wb = wb

    def parse_sheet_headers(self, ws, header_rows: int, first_header_row: int = 1) -> list:
        """解析单个工作表的表头，处理合并单元格"""
        max_col = self._get_max_col(ws, header_rows, first_header_row)
        col_headers = [""] * max_col

        # 第一步：逐行逐列读取值（跳过MergedCell，取父级的top-left值）
        for row_idx in range(first_header_row, first_header_row + header_rows):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is not None:
                    val_str = normalize_header_text(str(val))
                    if col_headers[col_idx - 1]:
                        # 已有值，用下划线纵向拼接
                        col_headers[col_idx - 1] += "_" + val_str
                    else:
                        col_headers[col_idx - 1] = val_str

        # 第二步：从子到父处理合并单元格，用下划线横向拼接父-子关系
        col_headers = self._fill_from_merged(ws, col_headers, header_rows, max_col, first_header_row)
        return col_headers

    def _get_max_col(self, ws, header_rows: int, first_header_row: int = 1) -> int:
        max_col = 0
        for row_idx in range(first_header_row, first_header_row + header_rows):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    max_col = max(max_col, col_idx)
        return max(1, max_col)

    def _fill_from_merged(self, ws, col_headers: list, header_rows: int, max_col: int,
                          first_header_row: int = 1) -> list:
        """
        从合并单元格信息中补全表头

        核心逻辑：横向展开的父表头 → 下划线前缀到子表头
        例如：
          父表头（合并单元格）：工程数量
          子表头1：教学楼1#~3#
          子表头2：实验中心4#
          → 工程数量_教学楼1#~3# , 工程数量_实验中心4#

        处理顺序：从子到父（min_row 从高到低），确保多层嵌套正确
        例如：工程数量 → 地上 → 教学楼1#
          → 先处理"地上"：地上_教学楼1#
          → 再处理"工程数量"：工程数量_地上_教学楼1#

        first_header_row: 表头起始行，跳过该行以上的标题合并（如文件标题行）
        """
        # 从子到父排序（min_row越大越先处理）
        sorted_merges = sorted(ws.merged_cells.ranges,
                               key=lambda r: r.min_row, reverse=True)

        for merged_range in sorted_merges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col_r = merged_range.max_col

            # 跳过表头区域上方的合并单元格（如文件标题行）
            if max_row < first_header_row:
                continue
            # 跳过完全在表头区域下方的合并单元格
            if min_row >= first_header_row + header_rows:
                continue

            top_left = ws.cell(row=min_row, column=min_col).value
            if top_left is None:
                continue

            val_str = normalize_header_text(str(top_left))

            for c in range(min_col, max_col_r + 1):
                if c > len(col_headers):
                    continue
                current = col_headers[c - 1]
                if not current:
                    # 子列为空，直接填入父值（纵向合并跨行的列）
                    col_headers[c - 1] = val_str
                elif not current.startswith(val_str):
                    # 子列有值但缺父前缀，在前面补上（横向展开的列）
                    col_headers[c - 1] = val_str + "_" + current
                # else: 父值已包含在子列中（如逐行读取已拼接），跳过

        return col_headers

    def validate_headers(self, headers: list) -> dict:
        """校验表头是否包含必填字段"""
        norm_headers = [normalize_header_text(h) if h else "" for h in headers]
        required_norm = [normalize_header_text(h) for h in REQUIRED_HEADERS]

        mapped = {}
        missing = []
        for req in required_norm:
            found = None
            for i, h in enumerate(norm_headers):
                if h and (req == h or req in h or h in req):
                    found = i
                    break
            mapped[req] = found
            if found is None:
                missing.append(req)

        return {
            "is_valid": len(missing) == 0,
            "missing": missing,
            "mapped": mapped,
            "normalized_headers": norm_headers,
        }

    def auto_fill_building_type(self, headers: list, building_type: str) -> tuple:
        """
        确保"业态"列存在
        返回：(新表头列表, 是否新增了业态列)
        """
        result = list(headers)
        norm = [normalize_header_text(h) if h else "" for h in result]

        yt_idx = None
        for i, h in enumerate(norm):
            if h and ("业态" in h or "类型" in h):
                yt_idx = i
                break

        if yt_idx is None:
            result.insert(0, "业态")
            return result, True
        return result, False

    def ensure_remark_last(self, headers: list) -> list:
        """确保最后一列为"备注" """
        result = list(headers)
        norm = [normalize_header_text(h) if h else "" for h in result]

        remark_idx = None
        for i, h in enumerate(norm):
            if h and ("备注" in h or "remark" in h.lower()):
                remark_idx = i
                break

        if remark_idx is None:
            result.append("备注")
        elif remark_idx != len(result) - 1:
            remark_header = result.pop(remark_idx)
            result.append(remark_header)

        return result

    def standardize_headers(self, headers: list) -> list:
        """
        标准化表头：将各种写法统一为标准表头名
        使用最佳匹配（最长子串匹配），避免短关键词误匹配

        特殊处理：已用下划线拼接的多级表头（如"工程数量_教学楼1#"），
        匹配时使用完整字符串，看是否包含标准字段的关键词
        """
        mapping = {
            "业态": ["业态", "建筑类型", "楼栋类型", "类型"],
            "项目名称": ["项目名称", "清单名称", "分项名称", "名称"],
            "项目特征": ["项目特征", "特征", "特征描述", "工作描述", "规格"],
            "计量单位": ["计量单位", "单位", "单位名称"],
            "工程量": ["工程量", "数量", "清单量", "计取量"],
            "综合单价": ["综合单价", "单价", "清单单价"],
            "综合合价": ["综合合价", "合价", "清单合价", "总价"],
            "备注": ["备注", "说明", "备注说明"],
        }
        result = []
        for h in headers:
            if not h:
                result.append("")
                continue
            h_norm = normalize_header_text(h)
            # 最佳匹配：找到最长匹配的映射（避免短词误匹配）
            best_match = None
            best_len = 0
            for std, variants in mapping.items():
                for v in variants:
                    v_norm = normalize_header_text(v)
                    # 精确匹配优先级最高
                    if h_norm == v_norm:
                        best_match = std
                        best_len = len(v_norm)
                        break
                    # 子串匹配，记录最长匹配
                    elif v_norm in h_norm and len(v_norm) > best_len:
                        best_match = std
                        best_len = len(v_norm)
                    elif h_norm in v_norm and len(h_norm) > best_len:
                        best_match = std
                        best_len = len(h_norm)
                if best_match and best_len == len(h_norm):  # 精确匹配，直接break
                    break
            if best_match:
                result.append(best_match)
            else:
                result.append(h)
        return result
