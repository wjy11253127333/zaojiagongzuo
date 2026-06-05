"""
类型标记列添加脚本 — add_type_column.py

功能：根据"项目名称"和"项目特征描述"列的内容情况，在"业态"列后自动新增"分部分项标记"列。
  - C列(项目名称)有内容、D列(项目特征描述)为空 → 标记为"分部列"
  - C列(项目名称)有内容、D列(项目特征描述)有内容 → 标记为"分项列"
  - C列(项目名称)为空 → 不填写

使用方式：
    python add_type_column.py input.xlsx [output.xlsx]

作为管线独立步骤运行，前置步骤为 cleanup_empty_rows.py。
"""

import sys
import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── 格式常量 ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def find_header_row(ws) -> int:
    """
    自动检测表头行号。
    策略：从行1开始，找到第一个含"项目名称"或"项目特征描述"的行。

    参数:
        ws: 工作表对象
    返回:
        表头行号（1-based），未找到则返回3（默认值，与管线约定一致）
    """
    keywords = ["项目名称", "项目特征描述"]
    for row in range(1, min(12, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                for kw in keywords:
                    if kw in cell.value:
                        return row
    return 3


def find_column_by_keyword(ws, keyword: str, header_row: int) -> int:
    """
    根据关键字在表头行中查找列号（支持模糊子串匹配）。

    参数:
        ws: 工作表对象
        keyword: 关键字（如"项目名称"）
        header_row: 表头行号
    返回:
        列号（1-based），未找到则返回0
    """
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value and isinstance(cell.value, str):
            if keyword in cell.value:
                return col
    return 0


def add_type_column_to_worksheet(ws, header_row: int = None) -> dict:
    """
    为单个工作表在"业态"列后添加"分部分项标记"列。

    逻辑:
        - 项目名称有值 + 项目特征描述为空 → "分部列"
        - 项目名称有值 + 项目特征描述有值 → "分项列"
        - 项目名称无值 → 不填写

    参数:
        ws: 工作表对象
        header_row: 表头行号（None表示自动检测）
    返回:
        {"分部列": n, "分项列": n, "跳过": n, "标记列": col_letter}
    """
    if header_row is None:
        header_row = find_header_row(ws)

    # 查找目标列
    name_col = find_column_by_keyword(ws, "项目名称", header_row)
    feature_col = find_column_by_keyword(ws, "项目特征描述", header_row)

    if name_col == 0:
        print(f"  警告: {ws.title} 未找到'项目名称'列，跳过")
        return None

    print(f"  {ws.title}:")
    name_letter = openpyxl.utils.get_column_letter(name_col)
    print(f"    项目名称: {name_letter}(列{name_col})", end="")
    if feature_col > 0:
        feature_letter = openpyxl.utils.get_column_letter(feature_col)
        print(f", 项目特征描述: {feature_letter}(列{feature_col})")
    else:
        print(", 项目特征描述: (未找到)")

    # 查找"业态"列，在其后插入新列
    yt_col = find_column_by_keyword(ws, "业态", header_row)
    if yt_col > 0:
        new_col = yt_col + 1
        ws.insert_cols(new_col)  # 物理插入空列，右侧列自动右移（公式引用同步调整）
        print(f"    业态: {openpyxl.utils.get_column_letter(yt_col)}(列{yt_col})"
              f" → 在 {openpyxl.utils.get_column_letter(new_col)} 插入'分部分项标记'")
    else:
        # 兼容：未找到"业态"列时追加到末尾
        new_col = ws.max_column + 1
        print(f"    未找到'业态'列，追加到末尾 {openpyxl.utils.get_column_letter(new_col)}")
    new_col_letter = openpyxl.utils.get_column_letter(new_col)

    # 写入表头
    header_cell = ws.cell(row=header_row, column=new_col)
    header_cell.value = "分部分项标记"
    header_cell.font = HEADER_FONT
    header_cell.fill = HEADER_FILL
    header_cell.alignment = CENTER_ALIGN
    header_cell.border = THIN_BORDER

    # 逐行处理数据
    data_start = header_row + 1
    stats = {"分部列": 0, "分项列": 0, "跳过": 0}

    for row in range(data_start, ws.max_row + 1):
        # 跳过全空行
        is_empty = True
        for col in range(1, ws.max_column):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and not isinstance(cell, MergedCell):
                is_empty = False
                break
        if is_empty:
            continue

        # 获取项目名称和项目特征描述的值
        name_val = ws.cell(row=row, column=name_col).value
        feature_val = ws.cell(row=row, column=feature_col).value if feature_col > 0 else None

        # 判断逻辑
        if name_val is not None and str(name_val).strip() != "":
            if feature_val is not None and str(feature_val).strip() != "":
                ws.cell(row=row, column=new_col).value = "分项列"
                ws.cell(row=row, column=new_col).font = BODY_FONT
                ws.cell(row=row, column=new_col).alignment = CENTER_ALIGN
                ws.cell(row=row, column=new_col).border = THIN_BORDER
                stats["分项列"] += 1
            else:
                ws.cell(row=row, column=new_col).value = "分部列"
                ws.cell(row=row, column=new_col).font = BODY_FONT
                ws.cell(row=row, column=new_col).alignment = CENTER_ALIGN
                ws.cell(row=row, column=new_col).border = THIN_BORDER
                stats["分部列"] += 1
        else:
            stats["跳过"] += 1

    print(f"    新增列: {new_col_letter}(分部分项标记), "
          f"分部列={stats['分部列']}, "
          f"分项列={stats['分项列']}, "
          f"跳过={stats['跳过']}")

    return stats


def add_type_column(input_path: str, output_path: str | None = None) -> str:
    """
    为 Excel 文件中所有工作表在"业态"列后添加"分部分项标记"列。

    参数:
        input_path: 输入文件路径
        output_path: 输出文件路径（默认: 输入文件名_type.xlsx）
    返回:
        输出文件路径
    """
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_type{ext}"

    print(f"加载: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    total = {"分部列": 0, "分项列": 0, "跳过": 0}
    for ws in wb.worksheets:
        result = add_type_column_to_worksheet(ws)
        if result:
            total["分部列"] += result["分部列"]
            total["分项列"] += result["分项列"]
            total["跳过"] += result["跳过"]

    if total["分部列"] + total["分项列"] == 0:
        print("  未找到符合条件的行，无需添加。")
    else:
        print(f"  合计: 分部列={total['分部列']}, "
              f"分项列={total['分项列']}, "
              f"跳过={total['跳过']}")

    wb.save(output_path)
    print(f"保存: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("用法: python add_type_column.py input.xlsx [output.xlsx]")
        print("示例: python add_type_column.py 综合楼_clean.xlsx 综合楼_类型.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_path):
        print(f"错误: 文件不存在 — {input_path}")
        sys.exit(1)

    add_type_column(input_path, output_path)


if __name__ == "__main__":
    main()
