"""
Excel读取与结构识别模块
功能：读取上传文件，识别工作表数量及结构（单表或多表组合）
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import pandas as pd
from openpyxl.cell.cell import MergedCell
from utils import normalize_header_text, extract_building_type


class ExcelReader:
    """Excel读取与结构识别"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.sheet_names = []
        self.sheet_structures = {}  # {sheet_name: {"nrows", "ncols", "has_merge", "header_rows"}}

    def load(self):
        """加载Excel文件"""
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
        self.sheet_names = self.wb.sheetnames
        return self

    def recognize_structure(self) -> dict:
        """
        识别所有工作表的结构
        返回：{
            "sheet_count": int,
            "is_multi_sheet": bool,
            "sheets": {sheet_name: structure_info},
            "recommendation": str,  # "single" or "multi"
        }
        """
        if self.wb is None:
            self.load()

        structures = {}
        for name in self.sheet_names:
            ws = self.wb[name]
            info = self._analyze_sheet(ws)
            structures[name] = info

        is_multi = len(self.sheet_names) > 1
        return {
            "sheet_count": len(self.sheet_names),
            "is_multi_sheet": is_multi,
            "sheets": structures,
            "recommendation": "multi" if is_multi else "single",
        }

    def _analyze_sheet(self, ws) -> dict:
        """分析单个工作表的结构"""
        merged_ranges = list(ws.merged_cells.ranges)
        has_merge = len(merged_ranges) > 0

        # 找到有数据的最大行/列
        max_row = 0
        max_col = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

        if max_row == 0:
            max_col = 0

        # 检测表头起始行和表头行数
        first_header_row = self._detect_first_header_row(ws)
        header_rows = self._detect_header_rows(ws, first_header_row)

        # 提取业态（从工作表名）
        building_type = extract_building_type(ws.title)

        return {
            "title": ws.title,
            "max_row": max_row,
            "max_col": max_col,
            "has_merge": has_merge,
            "merged_ranges": [str(r) for r in merged_ranges],
            "first_header_row": first_header_row,
            "header_rows": header_rows,
            "building_type": building_type,
            "is_empty": max_row == 0,
        }

    def _detect_first_header_row(self, ws) -> int:
        """
        检测第一行真正的表头行（跳过文件标题行）
        规则：
          1. 跳过标题行：≤2个非空单元格 且 任一单元格>15字
          2. 从上往下找第一个至少有3个独立非空单元格的行
        """
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            distinct_count = 0
            cell_texts = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    distinct_count += 1
                    cell_texts.append(str(cell.value).strip())

            # 标题行判定：存在 >35字单元格 → 跳过（真实列名最长~26字）
            if any(len(t) > 35 for t in cell_texts):
                continue

            if distinct_count >= 3:
                return row_idx
        return 1

    def _detect_header_rows(self, ws, first_header_row: int = 1) -> int:
        """
        检测表头占几行（通过合并单元格模式判断）
        从 first_header_row 开始查找连续有合并单元格的行，
        遇到第一个没有合并的行就停止
        """
        merge_count_per_row = {}
        for merged_range in ws.merged_cells.ranges:
            min_row = merged_range.min_row
            if min_row >= first_header_row:
                for r in range(min_row, merged_range.max_row + 1):
                    merge_count_per_row[r] = merge_count_per_row.get(r, 0) + 1

        header_rows = 0
        for r in range(first_header_row, first_header_row + 20):
            if merge_count_per_row.get(r, 0) > 0:
                header_rows += 1
            else:
                break  # 遇到没有合并的行，表头区域结束
        return max(1, header_rows)

    def read_sheet_raw(self, sheet_name: str, header_rows: int = None,
                       first_header_row: int = None) -> dict:
        """
        读取原始工作表数据（不做处理）
        """
        ws = self.wb[sheet_name]
        info = self._analyze_sheet(ws)
        if header_rows is None:
            header_rows = info.get("header_rows", 1)
        if first_header_row is None:
            first_header_row = info.get("first_header_row", 1)

        raw_headers = []
        for row_idx in range(first_header_row, first_header_row + header_rows):
            row_data = []
            for col_idx in range(1, info["max_col"] + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            raw_headers.append(row_data)

        data_start = first_header_row + header_rows
        data = []
        for row_idx in range(data_start, info["max_row"] + 1):
            row_data = []
            for col_idx in range(1, info["max_col"] + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            data.append(row_data)

        return {
            "raw_headers": raw_headers,
            "first_header_row": first_header_row,
            "header_rows": header_rows,
            "data": data,
            "building_type": info["building_type"],
            "max_col": info["max_col"],
        }

    def get_all_sheets_raw(self) -> list:
        """读取所有工作表的原始数据"""
        results = []
        for name in self.sheet_names:
            raw = self.read_sheet_raw(name)
            raw["sheet_name"] = name
            results.append(raw)
        return results
