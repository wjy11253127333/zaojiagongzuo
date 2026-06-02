"""
表头拆解工具 — header_flattener.py

功能：自动检测多级表头区域，拆解成单行表头（下划线连接），
      写入原位并删除多余表头行，保持原表格式不变。

处理流程：
  1. 复制原文件
  2. 逐个工作表自动检测表头起始行和表头行数
  3. 解析多级表头 → 下划线拼接为单行
  4. 检测无表头有数据的列 → 自动命名 "增项_N"（全局递增）
  5. 在"备注"列后插入 "业态" 列（数据行填入工作表名称）
  6. 全表合并单元格处理（表头区取消/数据区纵向填充/数据区横向不填充）
  7. 清空表头区域内容，写入拆解后的单行表头
  8. 填充业态列数据
  9. 扫描公式修正引用（行号偏移 + 被删区域警告）
  10. 物理删除多余表头空行（从下往上删）
  11. 统一格式美化（微软雅黑10pt, 深蓝表头, 全表细线边框）
  12. 保存新文件

用法：
  python header_flattener.py <输入Excel路径> [输出Excel路径]
"""

import os
import re
import sys
import shutil

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.cell.cell import MergedCell
from excel_reader import ExcelReader
from header_parser import HeaderParser


def _fix_formulas_after_row_delete(ws, deleted_rows: list, max_col: int) -> dict:
    """
    扫描工作表所有含公式的单元格，修正因删除行导致的公式引用偏移。

    参数:
        ws: openpyxl Worksheet
        deleted_rows: 将要删除的行号列表（从下往上排序）
        max_col: 工作表最大列数
    返回:
        {"fixed": 修正数量, "warned": 警告数量, "details": [...]}
    """
    if not deleted_rows:
        return {"fixed": 0, "warned": 0, "details": []}

    # deleted_rows 已从下往上排序（物理删除顺序）
    # 计算每个被删行之前累计已删行数
    # 例如删除 [5, 4]: 最终所有 R>5 的行号-2
    deleted_set = set(deleted_rows)
    max_deleted = max(deleted_rows)
    min_deleted = min(deleted_rows)
    shift = len(deleted_rows)  # 统一偏移量

    fixed_count = 0
    warned_count = 0
    details = []

    # 公式单元格引用正则: 捕获可选的 [工作表名!] + [列字母] + [行号]
    # 支持: A1, $A$1, A$1, $A1, Sheet2!A1, 'Sheet Name'!A1, Sheet2!$A$1
    cell_ref_pattern = re.compile(
        r"(?:(?P<sheet>('[^']+'|[A-Za-z0-9_\-.\u4e00-\u9fff\s]+)!))?"
        r"(?P<col>\$?[A-Z]{1,3})"
        r"(?P<row>\$?\d+)"
    )

    def fix_formula(match):
        nonlocal fixed_count, warned_count
        sheet_part = match.group("sheet") or ""
        col_part = match.group("col")
        row_str = match.group("row")

        # 解析行号
        is_absolute = row_str.startswith("$")
        row_num = int(row_str.lstrip("$"))

        if row_num < min_deleted:
            # 在被删区域之上，不变
            return match.group(0)

        if min_deleted <= row_num <= max_deleted:
            # 在被删区域之内 → 替换为警告文本
            warned_count += 1
            return '"表头区域已删除，用户确认数据是否正确"'

        # 在被删区域之下 → 行号 - shift
        new_row = row_num - shift
        new_row_str = f"${new_row}" if is_absolute else str(new_row)
        result = f"{sheet_part}{col_part}{new_row_str}"
        fixed_count += 1
        return result

    max_row = ws.max_row
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                original = cell.value
                new_formula = cell_ref_pattern.sub(fix_formula, original)
                if new_formula != original:
                    cell.value = new_formula
                    details.append(f"  {cell.coordinate}: {original} → {new_formula}")

    return {"fixed": fixed_count, "warned": warned_count, "details": details}


def _shift_formula_columns(formula: str, col_offset: int) -> str:
    """
    将公式中所有列引用按 col_offset 平移（正值右移，负值左移）。

    参数:
        formula: 公式字符串（以 = 开头）
        col_offset: 列偏移量（例：-2 表示所有列引用左移2列）
    返回:
        平移后的公式字符串
    """
    import re

    # 匹配 $?[A-Z]+[0-9]+ 的单元格引用
    pattern = re.compile(r'(\$?)([A-Z]+)(\d+)')

    def replacer(m):
        dollar = m.group(1)
        col_str = m.group(2)
        row_str = m.group(3)
        try:
            col_idx = openpyxl.utils.column_index_from_string(col_str)
        except ValueError:
            return m.group(0)  # 无法解析，保留原样
        new_col_idx = col_idx + col_offset
        if new_col_idx < 1:
            return m.group(0)  # 溢出 A 列，保留原样
        new_col = openpyxl.utils.get_column_letter(new_col_idx)
        return f"{dollar}{new_col}{row_str}"

    return pattern.sub(replacer, formula)


def _fix_formulas_after_col_insert(ws, insert_col: int, max_row: int, max_col: int) -> dict:
    """
    扫描工作表所有含公式的单元格，修正因插入列导致的列引用偏移。

    参数:
        ws: openpyxl Worksheet
        insert_col: 插入列的 1-based 列号
        max_row: 工作表最大行数
        max_col: 工作表最大列数（插入前）
    返回:
        {"fixed": 修正数量}
    """
    # 匹配单元格引用：可选工作表前缀 + 可选$ + 列字母 + 行号
    cell_ref_pattern = re.compile(
        r"(?:(?P<sheet>('[^']+'|[A-Za-z0-9_\-.\u4e00-\u9fff\s]+)!))?"
        r"(?P<col>\$?[A-Z]{1,3})"
        r"(?P<row>\$?\d+)"
    )

    fixed_count = 0

    def fix_formula(match):
        nonlocal fixed_count
        sheet_part = match.group("sheet") or ""
        col_str = match.group("col")
        row_str = match.group("row")

        # 解析列字母
        is_abs_col = col_str.startswith("$")
        col_letter = col_str.lstrip("$")
        col_idx = openpyxl.utils.column_index_from_string(col_letter)

        if col_idx >= insert_col:
            # 该列在插入列之后 → 右移1列
            new_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            new_col = f"${new_letter}" if is_abs_col else new_letter
            fixed_count += 1
            return f"{sheet_part}{new_col}{row_str}"
        return match.group(0)

    # 扫描公式范围：插入列之后的所有列 + 全行范围
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                original = cell.value
                new_formula = cell_ref_pattern.sub(fix_formula, original)
                if new_formula != original:
                    cell.value = new_formula

    return {"fixed": fixed_count}


def _apply_formatting(ws, first_header_row: int, max_col: int):
    """
    统一工作表格式。

    参数:
        ws: openpyxl Worksheet
        first_header_row: 标准化后的表头行号
        max_col: 工作表最大列数
    """
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )

    # 样式定义
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_align = Alignment(horizontal="right", vertical="center")
    title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_row = ws.max_row

    # 判断列类型：表头含数字关键词 → 数值列
    numeric_keywords = ["工程量", "合价", "单价", "数量", "金额", "造价", "增减"]
    numeric_cols = set()
    for col_idx in range(1, max_col + 1):
        h = ws.cell(row=first_header_row, column=col_idx).value
        if h and any(kw in str(h) for kw in numeric_keywords):
            numeric_cols.add(col_idx)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # 边框
            cell.border = thin_border

            if row_idx == first_header_row:
                # 表头行
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            elif row_idx < first_header_row:
                # 标题行（Row 1-2）：保持原样，只加边框和居中
                cell.alignment = title_align
                if cell.font and cell.font.name:
                    # 已有字体，不改
                    pass
                else:
                    cell.font = data_font
            else:
                # 数据行
                cell.font = data_font
                if col_idx in numeric_cols:
                    cell.alignment = number_align
                else:
                    cell.alignment = text_align

            # 行高
            if row_idx == first_header_row:
                ws.row_dimensions[row_idx].height = None  # 自适应
            elif row_idx >= first_header_row + 1:
                ws.row_dimensions[row_idx].height = 20

    # 剥离溢出列的旧格式（超出 parsed 范围的列，如 Z/AA）
    # 这些列已被清空值为 None，但可能残留原始文件的边框/填充格式
    extra_max_col = ws.max_column
    if extra_max_col > max_col:
        empty_fill = PatternFill(fill_type=None)
        empty_border = Border()
        for row_idx in range(1, max_row + 1):
            for col_idx in range(max_col + 1, extra_max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = empty_border
                cell.fill = empty_fill


def _get_columns_with_images(ws) -> set:
    """
    获取工作表中所有含图片的列号集合（1-based）。

    Excel 图片不存储在 cell.value 中，需要单独检测。
    通过 openpyxl 的 _images 属性获取图片锚点信息。
    """
    cols = set()
    # 兼容 openpyxl >= 3.1 的 images 和旧版的 _images
    img_list = getattr(ws, 'images', None) or getattr(ws, '_images', [])
    for img in img_list:
        anchor = getattr(img, 'anchor', None)
        if anchor is None:
            continue
        # 字符串格式: "A1", "B2" 等
        if isinstance(anchor, str):
            col_letter = ''.join(c for c in anchor if c.isalpha())
            if col_letter:
                cols.add(openpyxl.utils.column_index_from_string(col_letter))
        else:
            # CellAnchor / OneCellAnchor / TwoCellAnchor 对象
            # _from.col 是 0-based
            from_obj = getattr(anchor, '_from', None)
            if from_obj is not None:
                col_idx = getattr(from_obj, 'col', None)
                if col_idx is not None:
                    cols.add(col_idx + 1)
                else:
                    # 回退: 尝试 col_off
                    col_off = getattr(from_obj, 'colOff', None)
                    if col_off is not None:
                        # colOff 是 EMU 单位，除以列宽(~1,000,000 EMU/列)近似计算
                        cols.add(int(col_off / 1000000) + 1)
    return cols


def _detect_orphaned_columns(ws, parsed_cols: int, data_start_row: int, max_row: int) -> list:
    """
    扫描数据行中超出 parsed 范围的列，返回有数据的列号列表（1-based）。

    场景：表头解析时某些列的表头为空（如 AA 列的残留数据未被 _get_max_col 检测到），
          但这些列在数据行中有实际内容（包括图片），需要自动命名后纳入标准化清单。
    """
    orphans = []
    scan_limit = max_row  # 扫描全部数据行，避免漏掉表格末尾的数据
    cols_with_images = _get_columns_with_images(ws)  # 图片也算数据
    for col_idx in range(parsed_cols + 1, ws.max_column + 1):
        has_data = False
        # 先检查图片
        if col_idx in cols_with_images:
            has_data = True
        else:
            # 再检查数据行值
            for row_idx in range(data_start_row, scan_limit + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    has_data = True
                    break
        if has_data:
            orphans.append(col_idx)
    return orphans



def flatten_headers(input_path: str, output_path: str = None) -> str:
    """
    对输入文件的所有工作表进行表头拆解，输出新文件

    参数:
        input_path: 输入Excel文件路径
        output_path: 输出路径（默认：原文件名_拆解.xlsx）
    返回:
        输出文件路径
    """
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = base + "_拆解.xlsx"

    # Step 0: 复制原文件
    print(f"[1/5] 复制原文件...")
    shutil.copy2(input_path, output_path)
    print(f"       {os.path.basename(input_path)} → {os.path.basename(output_path)}")

    # Step 1: 加载并分析
    print(f"[2/5] 加载并自动检测表头区域...")
    reader = ExcelReader(input_path)
    reader.load()

    # 用新的 workbook 进行操作
    wb = openpyxl.load_workbook(output_path)
    parser = HeaderParser()

    results = []
    orphan_counter = 0  # 全局计数器：无表头有数据列 → "增项_N"，跨工作表递增
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info = reader._analyze_sheet(reader.wb[sheet_name])

        if info["is_empty"]:
            print(f"       [{sheet_name}] → 空表，跳过")
            results.append({"sheet": sheet_name, "status": "skipped (empty)"})
            continue

        first_header_row = info.get("first_header_row", 1)
        header_rows = info.get("header_rows", 1)

        if header_rows <= 1:
            print(f"       [{sheet_name}] → 单行表头，无需拆解")
            results.append({"sheet": sheet_name, "status": "single-row, skipped"})
            continue

        print(f"       [{sheet_name}] 表头行: {first_header_row}-{first_header_row + header_rows - 1} ({header_rows}行)")

        # Step 2: 解析表头
        parsed = parser.parse_sheet_headers(
            reader.wb[sheet_name], header_rows, first_header_row
        )

        # 清理尾部空列
        while parsed and not parsed[-1]:
            parsed.pop()

        # Step 2a: 检测无表头但有数据的列 → 自动命名为 "增项_N"
        # 场景：如 AA 列表头为空但数据行有值，_get_max_col 会漏掉
        data_start_row = first_header_row + header_rows
        orphans = _detect_orphaned_columns(
            ws, len(parsed), data_start_row, info["max_row"]
        )
        # 孤儿列紧凑化：连续追加增项_N 到 parsed 末尾，不保留间隙空列
        # 后续 Step 5b 会将原位数据迁移到紧凑列位置
        orphan_mapping = {}  # {new_parsed_idx(0-based): original_col_idx}
        if orphans:
            for col_idx in sorted(orphans):
                orphan_counter += 1
                orphan_mapping[len(parsed)] = col_idx
                parsed.append(f"增项_{orphan_counter}")
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                print(f"          检测到增项列: {col_letter}(列{col_idx}) → 增项_{orphan_counter} (紧凑列{len(parsed)})")

        # Step 2b: 在"备注"列后插入"业态"列
        # 搜索 parsed 中包含"备注"的表头，在其后插入；找不到则追加到末尾
        yt_col_idx = None
        yt_inserted = False
        for i, h in enumerate(parsed):
            if h and "备注" in str(h):
                parsed.insert(i + 1, "业态")
                yt_col_idx = i + 2  # 1-based 列号
                yt_inserted = True
                print(f"          业态列: 插入在'备注'(列{i+1})之后 → 列{yt_col_idx}")
                break
        if yt_col_idx is None:
            parsed.append("业态")
            yt_col_idx = len(parsed)
            print(f"          业态列: 未找到'备注', 追加到末尾列{yt_col_idx}")

        # Step 2b 后续: 业态列插入 parsed 后，orphan_mapping 的键位同步右移
        # 增项列在 parsed 中的位置 ≥ 业态插入点，需要 +1
        if orphan_mapping and yt_inserted:
            insert_idx = yt_col_idx - 1  # 0-based parsed 插入索引
            adj = {}
            for k, v in orphan_mapping.items():
                adj[k + 1 if k >= insert_idx else k] = v
            orphan_mapping = adj

        # Step 3: 全表合并单元格处理
        #   表头区：直接取消合并（Step 4会清空内容）
        #   数据区纵向合并（同列多行）：取消合并 → 顶部值向下填充
        #   数据区横向/矩形合并：直接取消合并（不填充）
        unmerge_count = 0
        fill_down_count = 0

        # 需要复制一份合并范围列表，因为遍历过程中会修改
        merge_list = list(ws.merged_cells.ranges)

        for mr in merge_list:
            mr_str = str(mr)
            # 判断是否在表头区域（与表头行有重叠）
            in_header = (mr.min_row <= first_header_row + header_rows - 1 and
                         mr.max_row >= first_header_row)

            if in_header:
                # 表头区：直接取消
                ws.unmerge_cells(mr_str)
                unmerge_count += 1
            elif mr.min_row >= data_start_row:
                # 数据区
                if mr.min_col == mr.max_col and mr.min_row < mr.max_row:
                    # 纵向合并（同列多行）：取top-left值 → 取消 → 向下填充
                    val = ws.cell(row=mr.min_row, column=mr.min_col).value
                    ws.unmerge_cells(mr_str)
                    for r in range(mr.min_row + 1, mr.max_row + 1):
                        ws.cell(row=r, column=mr.min_col).value = val
                    fill_down_count += 1
                    unmerge_count += 1
                else:
                    # 横向/矩形合并：直接取消，不填充
                    ws.unmerge_cells(mr_str)
                    unmerge_count += 1

        # Step 3.5: 物理插入业态列（仅当业态插入在"备注"之后时）
        # parsed 中已插入"业态"，但 worksheet 中还没有新列 → 数据会错位
        # 需要 ws.insert_cols() 物理新建列 + 修正公式中的列引用偏移
        if yt_inserted:
            pre_insert_max_col = ws.max_column
            pre_insert_max_row = ws.max_row
            ws.insert_cols(yt_col_idx)
            formula_col_result = _fix_formulas_after_col_insert(
                ws, yt_col_idx, pre_insert_max_row, pre_insert_max_col
            )
            yt_col_letter = openpyxl.utils.get_column_letter(yt_col_idx)
            print(f"           列插入: 在列{yt_col_letter}插入'业态', 公式修正{formula_col_result['fixed']}处")

            # 业态列插入后，孤儿列原位索引 >= yt_col_idx 的右移1位
            if orphan_mapping:
                adj = {}
                for new_i, orig_c in orphan_mapping.items():
                    adj[new_i] = orig_c + 1 if orig_c >= yt_col_idx else orig_c
                orphan_mapping = adj

        # Step 4: 清空表头区域内容
        # 以实际拆解列数为准写入区域（max_col）；多余列（如 Z/AA 的残留值）
        # 也需要清空为 None，避免旧数据残留在可见位置
        max_col = len(parsed)
        clear_max_col = min(max(ws.max_column, max_col), max_col + 100)
        for row_idx in range(first_header_row, first_header_row + header_rows):
            for col_idx in range(1, clear_max_col + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None
                except (ValueError, AttributeError):
                    pass

        # Step 5: 写入拆解后的单行表头到表头区域第一行
        target_row = first_header_row
        for col_idx, header_text in enumerate(parsed, 1):
            cell = ws.cell(row=target_row, column=col_idx)
            cell.value = header_text if header_text else None

        # Step 5a: 填充"业态"列数据（所有数据行 = 工作表名称）
        for row_idx in range(data_start_row, ws.max_row + 1):
            ws.cell(row=row_idx, column=yt_col_idx).value = sheet_name

        # Step 5b: 迁移孤儿列数据到紧凑位置，清除原位及溢出残留
        # 孤儿列原位（如 AA/AD）可能分散，紧凑化后需要把数据搬到新列
        if orphan_mapping:
            for new_idx_0, orig_col in orphan_mapping.items():
                new_col = new_idx_0 + 1  # 1-based
                col_offset = new_col - orig_col  # 负值=左移, 0=原位
                if col_offset != 0:
                    # 数据需要从原位搬到紧凑列
                    for row_idx in range(data_start_row, ws.max_row + 1):
                        src_val = ws.cell(row=row_idx, column=orig_col).value
                        if src_val is not None:
                            ws.cell(row=row_idx, column=new_col).value = src_val
                        # 清除原位值
                        ws.cell(row=row_idx, column=orig_col).value = None
                # col_offset==0: 数据已因业态列插入自动移到目标位, 无需搬移
            # 统一清除所有超出紧凑范围的残留（表头+数据），避免列间错位覆盖
            for col_idx in range(len(parsed) + 1, ws.max_column + 1):
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None
            print(f"           孤儿列迁移: {len(orphan_mapping)} 列数据从原位→紧凑位, 公式已平移, 溢出列数据已清除")

        # Step 6: 扫描所有公式，修正因删除行导致的引用偏移
        # 收集将要删除的空行（除 target_row 外的所有表头行）
        rows_to_delete = []
        for r in range(first_header_row, first_header_row + header_rows):
            if r != target_row:
                rows_to_delete.append(r)
        rows_to_delete.sort(reverse=True)  # 从下往上

        formula_result = _fix_formulas_after_row_delete(ws, rows_to_delete, max_col)
        print(f"           公式修正: {formula_result['fixed']} 处引用调整, {formula_result['warned']} 处被删区域警告")

        # Step 7: 物理删除多余表头空行（从下往上删）
        for row_idx in rows_to_delete:
            ws.delete_rows(row_idx)
        print(f"           物理删除 {len(rows_to_delete)} 个空行")

        # Step 8: 统一格式美化
        _apply_formatting(ws, first_header_row, max_col)
        print(f"           格式统一: 微软雅黑10pt, 深蓝表头, 全表细线边框")

        results.append({
            "sheet": sheet_name,
            "status": "flattened",
            "from_rows": f"{first_header_row}-{first_header_row + header_rows - 1}",
            "to_row": first_header_row,
            "columns": len(parsed),
            "unmerged": unmerge_count,
            "fill_down": fill_down_count,
            "formula_fixed": formula_result["fixed"],
            "formula_warned": formula_result["warned"],
            "rows_deleted": len(rows_to_delete),
        })

        preview = [h[:25] if h else "(空)" for h in parsed[:6]]
        print(f"           拆解 {len(parsed)} 列, 取消 {unmerge_count} 个合并, 纵向填充 {fill_down_count} 个")
        print(f"           首列: {preview}")

    # Step 7: 保存
    print(f"[3/5] 保存文件...")
    wb.save(output_path)

    # 打印汇总
    print(f"[4/5] 处理完成！")
    print(f"{'='*60}")
    for r in results:
        print(f"  {r['sheet']}: {r['status']}")
        if r['status'] == 'flattened':
            fd = r.get('fill_down', 0)
            ff = r.get('formula_fixed', 0)
            fw = r.get('formula_warned', 0)
            rd = r.get('rows_deleted', 0)
            print(f"    原表头 {r['from_rows']} → 新表头第 {r['to_row']} 行, {r['columns']} 列")
            print(f"    合并: 取消 {r['unmerged']} 个, 纵向填充 {fd} 列 | 公式: 修正 {ff} 处, 警告 {fw} 处 | 删除 {rd} 空行")
    print(f"{'='*60}")
    print(f"[5/5] 输出: {output_path}")

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python header_flattener.py <输入Excel> [输出Excel]")
    else:
        input_path = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
        flatten_headers(input_path, output_path)
