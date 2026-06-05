"""
----------------------------------------------
清单表合并脚本 v1.0
----------------------------------------------
功能：
  1. 扫描工作簿，识别清单表（表头命中 >= 4 个关键字）
  2. 为每个清单表添加"业态"列（值=工作表名简称）
  3. pd.concat 合并（同名纵向拼接，异名横向新增列）
  4. 序号统一重新编号 1-N
  5. Unnamed 列自动重命名为"扩展列_N"
  6. 生成"合并说明"Sheet（列审计：同名列/异名列/空列）
  7. 格式标准化（深蓝表头、微软雅黑10pt、全表细线边框）

用法：
    python sheet_merger.py "input.xlsx"

输出：
    在原文件新增「清单汇总」Sheet +「合并说明」Sheet
    若原文件被占用，另存为 _含清单汇总.xlsx
"""

import os
import re
import sys
import shutil
from copy import copy as _copy_obj
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import normalize_header_text
from audit_logger import AuditLogger   # 新增：导入审计日志器

# ===== 配置 =====
LIST_KEYWORDS    = ["项目特征", "项目名称", "项目编码", "计量单位", "工程量"]
MIN_HIT          = 4
OUTPUT_SHEET     = "清单合并"    # 原："清单汇总"
REPORT_SHEET     = "合并说明"    # 原："操作记录"（列审计报告：同名列/异名列/空列分析）

# 样式常量
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_FONT  = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT  = Font(name="微软雅黑", size=10)
LEFT_ALIGN = Alignment(horizontal="left",    vertical="center", wrap_text=True)
CTR_ALIGN  = Alignment(horizontal="center",  vertical="center")
NUM_ALIGN  = Alignment(horizontal="right",   vertical="center")


# ============================================================
# 文件解析 & Flat 检测（Phase 1 新增）
# ============================================================

def _resolve_type_file(file_path: str, prefer_flat: bool = True) -> str:
    """
    自动解析输入文件路径，优先使用 header_flattener 预处理版本。

    查找顺序：
        1. {original}_type_v2.xlsx   (header_flattener v2 输出)
        2. {original}_type.xlsx      (header_flattener v1 输出)
        3. {original} 自身            (回退到原始文件)

    Args:
        file_path:   输入的 Excel 文件路径
        prefer_flat: 是否优先查找 flat 版本（默认 True）

    Returns:
        解析后的实际文件路径（已验证存在）
    """
    base, ext = os.path.splitext(file_path)

    if prefer_flat:
        candidates = [
            f"{base}_type_v2{ext}",
            f"{base}_type{ext}",
            file_path,
        ]
        for cand in candidates:
            if os.path.isfile(cand):
                if cand != file_path:
                    print(f"  [INFO] 自动使用预处理版本：{os.path.basename(cand)}")
                return cand

    # 不 prefer flat 或所有候选都不存在 → 返回原始路径
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")
    return file_path


def _is_flat_header_file(ws) -> bool:
    """
    判断当前 Worksheet 是否为已 flatten 的单级表头文件。

    检测特征（header_flattener 处理后的典型结构）：
      - Row 1-2 为标题/副标题行（仅有 1~2 个非空单元格）
      - 表头行含下划线连接关键词，如 "工程量_合同工程量_A"

    Args:
        ws: openpyxl Worksheet 对象

    Returns:
        bool: 是否判定为已处理的 flat 文件
    """
    hdr_row = get_header_row(ws, max_scan=5)
    if hdr_row is None:
        return False

    # 特征1：Row 1 和 Row 2 只有少量非空单元格（标题行特征）
    for check_row in [1, 2]:
        non_empty = sum(
            1 for c in range(1, min(ws.max_column + 1, 5))
            if ws.cell(row=check_row, column=c).value
        )
        if non_empty > 2:
            return False

    # 特征2：表头行有 >=3 个列名包含 "_X_" 下划线模式（flatten 标志）
    underline_pattern = re.compile(r'_\w+_')
    header_texts = [
        str(ws.cell(row=hdr_row, column=c).value or "")
        for c in range(1, min(ws.max_column + 1, 20))
    ]
    flat_count = sum(1 for h in header_texts if underline_pattern.search(h))

    return flat_count >= 3


# ============================================================
# 工作表识别
# ============================================================

def is_list_sheet(ws, max_scan=10) -> bool:
    """判断工作表是否为清单表（表头含 >= MIN_HIT 个关键字）"""
    hdr_row = get_header_row(ws, max_scan)
    if hdr_row is None:
        return False
    headers = [ws.cell(row=hdr_row, column=c).value or ""
               for c in range(1, ws.max_column + 1)]
    hdr_text = "".join(normalize_header_text(str(h)) for h in headers if h)
    return sum(1 for kw in LIST_KEYWORDS if kw in hdr_text) >= MIN_HIT


def get_header_row(ws, max_scan=10) -> int | None:
    """自动检测表头行（跳过前置标题行）"""
    # 先跳过内容超长的行（通常是标题行，如 >35字）
    start_row = 1
    for row in range(1, min(ws.max_row + 1, max_scan + 1)):
        cells = [str(ws.cell(row=row, column=c).value or "").strip()
                 for c in range(1, min(ws.max_column + 1, 10))]
        if any(len(c) > 35 for c in cells):
            continue
        start_row = row
        break

    best_row, best_hits = start_row, 0
    for row in range(start_row, min(ws.max_row + 1, start_row + max_scan)):
        cells = [str(ws.cell(row=row, column=c).value or "")
                 for c in range(1, ws.max_column + 1)]
        hdr_text = "".join(normalize_header_text(c) for c in cells if c)
        hits = sum(1 for kw in LIST_KEYWORDS if kw in hdr_text)
        if hits > best_hits:
            best_hits, best_row = hits, row

    return best_row if best_hits >= MIN_HIT else None


def get_sheet_headers(ws):
    """读取工作表的表头行，返回 (header_row, headers_list)"""
    hdr_row = get_header_row(ws)
    if hdr_row is None:
        return None, []
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=hdr_row, column=c).value
        headers.append(str(val).strip() if val else "")
    # 去除尾部空列
    while headers and not headers[-1]:
        headers.pop()
    return hdr_row, headers


# ============================================================
# Unnamed 列重命名
# ============================================================

def _rename_unnamed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """将 Unnamed 列动态重命名为 扩展列_N"""
    new_cols = []
    unnamed_idx = 0
    for col in df.columns:
        col_str = str(col).strip()
        if col_str.startswith("Unnamed:") or col_str == "" or col_str == "nan":
            unnamed_idx += 1
            new_cols.append(f"扩展列_{unnamed_idx}")
        else:
            new_cols.append(col_str)
    df.columns = new_cols
    return df


# ============================================================
# 表头重复过滤
# ============================================================

def _is_header_duplicate(df: pd.DataFrame, header_col: str, threshold: float = 0.7) -> bool:
    """判断DataFrame中某列是否有 >= threshold 的行与该列名重复"""
    if header_col not in df.columns:
        return False
    col_values = df[header_col].dropna().astype(str)
    if len(col_values) == 0:
        return False
    match_count = (col_values == header_col).sum()
    return match_count / len(col_values) >= threshold


# ============================================================
# 列审计 + 合并说明
# ============================================================

def _generate_merge_report(all_dfs: list, merged: pd.DataFrame,
                           sheet_names: list) -> pd.DataFrame:
    """逐列分析来源：同名列（>=2表）vs 异名列（仅1表）"""
    report_rows = []
    for col in merged.columns:
        source_sheets = []
        has_data_sheets = []
        for i, df in enumerate(all_dfs):
            if col in df.columns:
                source_sheets.append(sheet_names[i])
                # 检查是否有数据（非空值）
                non_null = df[col].dropna()
                if len(non_null) > 0:
                    has_data_sheets.append(sheet_names[i])

        source_count = len(source_sheets)
        col_type = "同名列" if source_count >= 2 else "异名列"
        coverage = ", ".join(has_data_sheets) if has_data_sheets else "无"

        # 检查是否为空列
        merged_non_null = merged[col].dropna() if col in merged.columns else pd.Series()
        is_empty = len(merged_non_null) == 0
        action = "物理删除" if is_empty else "保留"

        report_rows.append({
            "列名": col,
            "来源类型": col_type,
            "来源表": ", ".join(source_sheets) if source_sheets else "无",
            "有数据表": coverage,
            "行覆盖率": f"{len(merged_non_null)}/{len(merged)}" if not is_empty else "0",
            "处理": action,
        })

    return pd.DataFrame(report_rows)


# ============================================================
# 格式标准化
# ============================================================

def _apply_list_merge_formatting(ws):
    """为「清单汇总」Sheet 应用格式标准"""
    max_row = ws.max_row
    max_col = ws.max_column

    # 判断列类型
    numeric_keywords = ["工程量", "合价", "单价", "数量", "金额", "造价", "增减"]
    numeric_cols = set()
    center_cols = {1}  # 序号列居中
    for col_idx in range(1, max_col + 1):
        h = ws.cell(row=1, column=col_idx).value
        if h and any(kw in str(h) for kw in numeric_keywords):
            numeric_cols.add(col_idx)
        if h and "计量单位" in str(h):
            center_cols.add(col_idx)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            if row_idx == 1:
                # 表头行
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = HDR_ALIGN
            else:
                # 数据行
                cell.font = DATA_FONT
                if col_idx in center_cols:
                    cell.alignment = CTR_ALIGN
                elif col_idx in numeric_cols:
                    cell.alignment = NUM_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN

            # 行高
            if row_idx == 1:
                ws.row_dimensions[row_idx].height = None  # 自适应
            elif row_idx >= 2:
                ws.row_dimensions[row_idx].height = 20

    # 列宽自适应（上限50）
    for col_idx in range(1, max_col + 1):
        max_width = 8
        for row_idx in range(1, min(max_row + 1, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_width = max(max_width, min(len(str(val)) * 1.2, 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width

    # 冻结首行
    ws.freeze_panes = "A2"


def _apply_report_formatting(ws, report_df: pd.DataFrame):
    """为「合并说明」Sheet 应用格式"""
    # 表头
    for col_idx in range(1, len(report_df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = THIN_BORDER
        cell.value = report_df.columns[col_idx - 1]

    # 数据
    for row_idx in range(len(report_df)):
        for col_idx in range(len(report_df.columns)):
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            val = report_df.iloc[row_idx, col_idx]
            if pd.notna(val):
                cell.value = val
            cell.alignment = CTR_ALIGN if col_idx in (1, 5) else LEFT_ALIGN

    # 列宽
    col_widths = [18, 12, 30, 25, 12, 12]
    for col_idx, w in enumerate(col_widths, 1):
        if col_idx <= len(report_df.columns):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"


# ============================================================
# Sheet 复制工具（Step 6 多文件合并用）
# ============================================================

def _copy_sheet(src_wb, src_sheet_name: str, dst_wb, dst_sheet_name: str):
    """
    将源工作簿中的 Sheet 复制到目标工作簿（手动逐格复制）。

    Args:
        src_wb:          源 openpyxl Workbook
        src_sheet_name:  源 Sheet 名称
        dst_wb:          目标 openpyxl Workbook
        dst_sheet_name:  目标 Sheet 名称
    """
    ws_src = src_wb[src_sheet_name]
    ws_dst = dst_wb.create_sheet(dst_sheet_name)

    # 复制单元格值 + 样式
    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row,
                                 min_col=1, max_col=ws_src.max_column):
        for cell in row:
            dst_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    dst_cell.font = _copy_obj(cell.font)
                    dst_cell.fill = _copy_obj(cell.fill)
                    dst_cell.border = _copy_obj(cell.border)
                    dst_cell.alignment = _copy_obj(cell.alignment)
                    dst_cell.number_format = cell.number_format
                except Exception:
                    pass

    # 复制列宽
    for col_letter, col_dim in ws_src.column_dimensions.items():
        if col_dim.width:
            ws_dst.column_dimensions[col_letter].width = col_dim.width

    # 复制行高
    for row_num, row_dim in ws_src.row_dimensions.items():
        if row_dim.height:
            ws_dst.row_dimensions[row_num].height = row_dim.height

    # 复制合并单元格
    for merged_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_range))

    # 复制冻结窗格
    if ws_src.freeze_panes:
        ws_dst.freeze_panes = ws_src.freeze_panes


# ============================================================
# 主处理函数（多文件版本）
# ============================================================

def merge_sheets(
    input_files: list,
    output_path: str | None = None,
    audit_logger: "AuditLogger | None" = None,
    prefer_flat: bool = True,       # Phase 1 新增：优先使用 _type_v2 预处理文件
    original_file: str | None = None,  # v1.3：原始未处理文件的路径（用于保留真正的原始清单表）
):
    """
    主处理函数（多文件版）：
        扫描所有输入文件 → 识别清单表 → 加业态列 → 合并 → 格式标准化 → 写操作记录

    Args:
        input_files:    Excel 文件路径列表（支持1个或多个文件）
        output_path:    输出文件路径（为 None 时自动生成）
        audit_logger:   AuditLogger 实例（为 None 时内部新建）
        prefer_flat:    是否优先使用 _type_v2 预处理文件
        original_file:  真正的原始文件路径（用于在输出中保留含多级表头的原始清单表）

    Returns:
        dict: {"list_sheets", "output_file", "rows", "cols", "dropped_cols"}
    """
    # ---- 初始化审计日志器 ----
    if audit_logger is None:
        audit_logger = AuditLogger()

    # ---- 规范化 input_files ----
    if isinstance(input_files, str):
        input_files = [input_files]

    print(f"\n{'='*60}")
    print(f"清单表合并 v1.1 — 多文件模式")
    print(f"  输入文件数：{len(input_files)}")
    print(f"{'='*60}")

    # ---- Step 0a：自动解析 _type_v2 预处理文件（Phase 1 新增）----
    resolved_files = []
    for f in input_files:
        try:
            resolved = _resolve_type_file(f, prefer_flat=prefer_flat)
        except FileNotFoundError as e:
            print(f"  [WARN] {e}")
            audit_logger.log_error("文件解析", str(e))
            continue
        resolved_files.append(resolved)

    input_files = resolved_files

    if any("_type_v2" in f for f in input_files):
        print("  [MODE] 检测到预处理文件（_type_v2），使用单级表头合并模式")

    # ---- Step 0：审计 — 文件扫描 ----
    audit_logger.log_file_scan(input_files, source="多文件模式")

    # ---- Step 1：遍历所有文件，收集清单表 DataFrame ----
    all_dfs          = []
    list_sheet_names = []       # 用于列审计报告
    all_headers      = {}       # {file_sheet: [headers]}
    processed_files  = []       # 成功处理的文件列表

    for file_idx, file_path in enumerate(input_files, 1):
        short_name = os.path.basename(file_path)
        print(f"\n[{file_idx}/{len(input_files)}] 处理文件：{short_name}")

        if not os.path.isfile(file_path):
            print(f"  [WARN] 文件不存在，跳过：{file_path}")
            audit_logger.log_error("文件检查", f"文件不存在：{file_path}")
            continue

        # 用 openpyxl 识别清单表
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except PermissionError:
            print(f"  [WARN] 文件被占用（PermissionError），跳过：{short_name}")
            audit_logger.log_error("文件打开", f"文件被占用：{short_name}", file=short_name)
            continue
        except Exception as e:
            print(f"  [WARN] 文件打开失败：{e}")
            audit_logger.log_error("文件打开", str(e), file=short_name)
            continue

        sheet_names = wb.sheetnames
        file_list_sheets = []

        for sn in sheet_names:
            if sn in (OUTPUT_SHEET, REPORT_SHEET):
                continue
            ws = wb[sn]
            if not is_list_sheet(ws):
                continue
            file_list_sheets.append(sn)

        print(f"  识别到 {len(file_list_sheets)} 个清单表：{', '.join(file_list_sheets)}")

        # 逐个读取清单表
        for sn in file_list_sheets:
            ws = wb[sn]
            hdr_row, headers = get_sheet_headers(ws)
            if hdr_row is None:
                continue

            # ==== Phase 1 新增：检测 flat 文件 ====
            is_flat = _is_flat_header_file(ws)

            # 用 pandas 读数据
            try:
                if is_flat:
                    # Flat 文件：openpyxl 手读单级表头 + 手读数据（避 pd.read_excel bug）
                    headers_list = [
                        str(ws.cell(row=hdr_row, column=c).value or "").strip()
                        for c in range(1, ws.max_column + 1)
                    ]
                    # 去除尾部空列
                    while headers_list and not headers_list[-1]:
                        headers_list.pop()

                    # 用手动逐格读取数据（pd.read_excel 对某些单元格返回 nan）
                    data = []
                    for row_idx in range(hdr_row + 1, ws.max_row + 1):
                        row_data = [
                            ws.cell(row=row_idx, column=c).value
                            for c in range(1, len(headers_list) + 1)
                        ]
                        data.append(row_data)
                    df = pd.DataFrame(data, columns=headers_list)
                    del data  # 释放内存
                else:
                    # 原始文件：使用原有逻辑
                    df = pd.read_excel(
                        file_path, sheet_name=sn,
                        header=hdr_row - 1,   # pandas header 是 0-based
                    )
            except Exception as e:
                print(f"  [WARN] 读取失败：{sn} — {e}")
                audit_logger.log_error("读取工作表", str(e), file=short_name, sheet=sn)
                continue

            # 只保留有效列名（非空列）
            valid_cols = [c for c in df.columns if str(c).strip() not in ("", "nan")]
            df = df[valid_cols].copy()

            # 重命名 Unnamed 列
            df = _rename_unnamed_columns(df)

            # 过滤重复表头行
            for col in list(df.columns):
                if _is_header_duplicate(df, col):
                    df = df[df[col] != col]

            # 增项列来源标注：增项_N → 增项_N_Sheet名
            rename_map = {}
            for col in df.columns:
                if str(col).startswith("增项_"):
                    rename_map[col] = f"{col}_{sn}"
            if rename_map:
                df.rename(columns=rename_map, inplace=True)
                print(f"          增项列来源标注：{list(rename_map.values())}")

            # 加业态列（值 = 工作表全名）
            # Phase 1：如果 flat 文件已含业态列，不重复插入
            if "业态" not in df.columns:
                df.insert(0, "业态", sn)
            # else: _type_v2 文件已有业态列（header_flattener 填入），保持原值

            # 记录
            all_dfs.append(df)
            list_sheet_names.append(f"{short_name}_{sn}")
            all_headers[f"{short_name}_{sn}"] = df.columns.tolist()

            row_count = len(df)
            col_count = len(df.columns)
            print(f"  [{sn}] {row_count}行 x {col_count}列")

            # 审计日志
            audit_logger.log_sheet_detect(
                short_name, sn, df.columns.tolist(), row_count,
                hit_keywords=["序号", "项目编码", "项目名称", "项目特征", "计量单位"],
            )

        wb.close()
        if file_list_sheets:
            processed_files.append(file_path)

    # ---- 检查：是否找到清单表 ----
    if len(all_dfs) == 0:
        msg = "未找到任何清单表（需表头命中 >= {MIN_HIT} 个关键字）"
        print(f"\n[ERROR] {msg}")
        audit_logger.log_error("清单表识别", msg)
        return {"list_sheets": [], "output_file": None, "rows": 0, "cols": 0, "dropped_cols": 0}

    print(f"\n共识别 {len(all_dfs)} 个清单表，开始合并...")

    # ---- Step 2：审计 — 表头对比 ----
    audit_logger.log_header_compare(all_headers)

    # ---- Step 3：合并（基准列序锁定 + pd.concat outer join）----
    # Phase 1：锁定第一个 df 的列序作为基准
    base_columns = all_dfs[0].columns.tolist()

    merged = pd.concat(all_dfs, join="outer", ignore_index=True, sort=False)

    # 删除无效行（避免表头后出现空行）
    # 策略：
    #   - 删除「项目编码」为空 且 「项目名称」为空的行（真正无意义空行）
    #   - 保留分部标题行（如"土石方工程"，项目名称有值但项目编码为空）
    before_drop = len(merged)

    if "项目编码" in merged.columns and "项目名称" in merged.columns:
        # 项目编码 非空（有值） → 保留
        cond_code_ok = merged["项目编码"].notna() & (
            merged["项目编码"].astype(str).str.strip().ne("")
        )
        # 项目名称 非空（分部标题行） → 保留
        cond_name_ok = merged["项目名称"].notna() & (
            merged["项目名称"].astype(str).str.strip().ne("")
        )
        # 任一条件满足即保留
        mask = cond_code_ok | cond_name_ok
        removed = before_drop - mask.sum()
        if removed > 0:
            print(f"  [INFO] 已删除 {int(removed)} 行（项目编码和项目名称均为空）")
        merged = merged[mask].reset_index(drop=True)
    else:
        # 降级：只按关键列判断
        key_col = None
        for candidate in ["项目编码", "序号"]:
            if candidate in merged.columns:
                key_col = candidate
                break
        if key_col is None:
            for c in merged.columns:
                if str(c).strip() not in ("", "序号", "业态", "分部分项标记"):
                    key_col = c
                    break

        if key_col and key_col in merged.columns:
            mask = merged[key_col].notna() & (merged[key_col].astype(str).str.strip().ne(""))
            removed = before_drop - mask.sum()
            if removed > 0:
                print(f"  [INFO] 已删除 {int(removed)} 行「{key_col}」为空的数据")
            merged = merged[mask].reset_index(drop=True)
        else:
            merged.dropna(how="all", inplace=True)
            merged.reset_index(drop=True, inplace=True)

    # 显式重排列序：同名列按基准表顺序，差异列追加到末尾
    extra_columns = [c for c in merged.columns if c not in base_columns]
    merged = merged[base_columns + extra_columns]

    # 重新编号：先删源表自带的序号列，再统一插入 1-based 编号
    if "序号" in merged.columns:
        merged.drop(columns=["序号"], inplace=True)
    merged.insert(0, "序号", range(1, len(merged) + 1))

    print(f"\n合并结果：{merged.shape[0]}行 x {merged.shape[1]}列")

    # ---- Step 3b：审计 — 合并结果 ----
    # 计算新增的差异列
    all_col_sets = [set(df.columns) for df in all_dfs]
    common_cols  = set.intersection(*all_col_sets) if all_col_sets else set()
    unique_cols  = set.union(*all_col_sets) - common_cols if all_col_sets else set()
    added_cols   = sorted(unique_cols)

    audit_logger.log_merge_result(merged, added_cols=list(added_cols))

    # ---- Step 4：列审计（来源分析 + 空列删除）----
    report_df = _generate_merge_report(all_dfs, merged, list_sheet_names)

    # 删除空列（但保留增项_*列 — 用户有意创建的扩展列，即使当前无数据也应保留）
    cols_to_drop = report_df[report_df["处理"] == "物理删除"]["列名"].tolist()
    cols_to_drop = [c for c in cols_to_drop if c in merged.columns
                    and c not in ("序号", "业态")
                    and not str(c).startswith("增项_")]  # v1.3：增项列不删
    if cols_to_drop:
        merged.drop(columns=cols_to_drop, inplace=True)
        print(f"物理删除 {len(cols_to_drop)} 个空列：{[c[:20] for c in cols_to_drop[:5]]}...")

    print(f"最终：{merged.shape[0]}行 x {merged.shape[1]}列")

    # ---- Step 5：确定输出路径 ----
    if output_path is None:
        # 自动生成：取第一个文件名 + "_清单合并.xlsx"
        first_name = os.path.splitext(os.path.basename(input_files[0]))[0]
        output_path = os.path.join(os.path.dirname(input_files[0]), f"{first_name}_清单合并.xlsx")

    # 若输出文件已存在，自动重命名
    base, ext = os.path.splitext(output_path)
    counter = 2
    while os.path.exists(output_path):
        output_path = f"{base}_v{counter}{ext}"
        counter += 1

    # ---- Step 6：写入 Excel（保留原始 Sheet + 新增合并结果）----
    try:
        import tempfile

        # 确定基础文件：优先使用原始文件（保留真正含多级表头的原始清单表）
        # 若未传 original_file，则回退到第一个成功处理的文件
        if original_file and os.path.isfile(original_file):
            base_file = original_file
            print(f"  [INFO] 基底文件使用原始文件：{os.path.basename(base_file)}")
        else:
            base_file = processed_files[0] if processed_files else input_files[0]

        # 确保输出目录存在
        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        # 复制基础文件到输出路径（保留所有原始 Sheet）
        if os.path.abspath(base_file) != os.path.abspath(output_path):
            shutil.copy2(base_file, output_path)

        # 用 openpyxl 加载输出文件，准备追加其他文件的 Sheet 和新 Sheet
        wb = openpyxl.load_workbook(output_path)

        # ---- 多文件支持：从第二个文件开始复制 Sheet 到输出文件 ----
        for other_file in input_files[1:]:
            if not os.path.isfile(other_file):
                continue
            try:
                wb_other = openpyxl.load_workbook(other_file, data_only=True)
            except Exception as e:
                print(f"  [WARN] 无法打开 {os.path.basename(other_file)}: {e}")
                continue

            for sn in wb_other.sheetnames:
                if sn in (OUTPUT_SHEET, REPORT_SHEET):
                    continue

                ws_new = wb_other[sn]
                new_name = sn

                # 冲突检测：表头对比
                if sn in wb.sheetnames:
                    ws_existing = wb[sn]

                    # 获取两个 Sheet 的表头行
                    hdr_row_existing = get_header_row(ws_existing)
                    hdr_row_new = get_header_row(ws_new)

                    if hdr_row_existing is not None and hdr_row_new is not None:
                        headers_existing = [
                            str(ws_existing.cell(row=hdr_row_existing, column=c).value or "").strip()
                            for c in range(1, min(ws_existing.max_column + 1, 50))
                        ]
                        headers_new = [
                            str(ws_new.cell(row=hdr_row_new, column=c).value or "").strip()
                            for c in range(1, min(ws_new.max_column + 1, 50))
                        ]

                        # 去除尾部空列
                        while headers_existing and not headers_existing[-1]:
                            headers_existing.pop()
                        while headers_new and not headers_new[-1]:
                            headers_new.pop()

                        if headers_existing == headers_new:
                            print(f"  [WARN] 疑似重复上传：Sheet「{sn}」"
                                  f"在 {os.path.basename(base_file)} 与 {os.path.basename(other_file)} "
                                  f"中表头完全相同，已跳过")
                            audit_logger.log_error(
                                "Sheet冲突",
                                f"表头相同，跳过重复 Sheet「{sn}」（来源：{os.path.basename(other_file)}）",
                                remark="建议检查是否重复上传同一文件",
                            )
                            continue
                        else:
                            # 表头不同，重命名（D3-A 规则）
                            file_short = os.path.basename(other_file).split("_type")[0].split(".")[0]
                            new_name = f"{sn}({file_short})"
                            counter = 1
                            while new_name in wb.sheetnames:
                                new_name = f"{sn}({file_short})_{counter}"
                                counter += 1
                            print(f"  [INFO] Sheet「{sn}」重命名为「{new_name}」（表头不同）")

                # 复制 Sheet（手动复制单元格）
                _copy_sheet(wb_other, sn, wb, new_name)

            wb_other.close()

        # 若已存在同名目标 Sheet，先删除（覆盖模式）
        for sheet_name in (OUTPUT_SHEET, REPORT_SHEET):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        # 保存并关闭（保留原始 Sheet 的状态），然后用 pandas 追加新 Sheet
        wb.save(output_path)
        wb.close()

        # 用 pandas 写 merged + report_df（追加模式）
        with pd.ExcelWriter(
            output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            merged.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
            report_df.to_excel(writer, sheet_name=REPORT_SHEET, index=False)

        # 格式标准化
        wb = openpyxl.load_workbook(output_path)

        ws_output = wb[OUTPUT_SHEET]
        _apply_list_merge_formatting(ws_output)

        ws_report = wb[REPORT_SHEET]
        _apply_report_formatting(ws_report, report_df)

        wb.save(output_path)
        wb.close()

        original_sheet_count = len(openpyxl.load_workbook(output_path).sheetnames) - 2
        print(f"\n[OK] 已保存到：{os.path.basename(output_path)}")
        print(f"       原始清单表已保留（{original_sheet_count} 个 Sheet）")
        audit_logger.log_output(output_path, [OUTPUT_SHEET, REPORT_SHEET])
        audit_logger.log_format_apply(OUTPUT_SHEET)
        audit_logger.log_format_apply(REPORT_SHEET)

    except Exception as e:
        print(f"\n[ERROR] 写入文件失败：{e}")
        audit_logger.log_error("写文件", str(e), remark=output_path)
        raise

    # ---- Step 7：汇总 ----
    print(f"\n{'='*60}")
    print(f"处理完成！")
    print(f"  清单表数：{len(all_dfs)}")
    print(f"  输入文件：{len(processed_files)} 个")
    print(f"  输出：{os.path.basename(output_path)}")
    print(f"  Sheets：{OUTPUT_SHEET} ({merged.shape[0]}行x{merged.shape[1]}列)")
    print(f"         {REPORT_SHEET} ({len(report_df)} 行)")
    print(f"{'='*60}\n")

    return {
        "list_sheets": list_sheet_names,
        "output_file": output_path,
        "rows": merged.shape[0],
        "cols": merged.shape[1],
        "dropped_cols": len(cols_to_drop),
    }


# ============================================================
# CLI 入口（多文件支持）
# ============================================================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python sheet_merger.py <输入Excel1> [输入Excel2] ...")
    else:
        input_files = sys.argv[1:]
        for f in input_files:
            if not os.path.exists(f):
                print(f"[ERROR] 文件不存在：{f}")
                sys.exit(1)
        merge_sheets(input_files)
