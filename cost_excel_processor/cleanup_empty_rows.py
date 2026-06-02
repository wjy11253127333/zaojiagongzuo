"""
空行清理脚本 — cleanup_empty_rows.py

功能：删除所有工作表的尾部连续空行。
空行定义：核心数据列（A列到业态列前一列）全部为 None/空。

使用方式：
    python cleanup_empty_rows.py input.xlsx [output.xlsx]

作为管线最后一步独立运行，前置步骤为 header_flattener.py。
"""

import sys
import os
import openpyxl


def find_core_end_col(ws, header_row: int = 3) -> int:
    """
    从表头行定位"业态"列，返回核心数据列的结束位置（业态列前一列）。

    参数:
        ws: 工作表对象
        header_row: 表头所在行号（默认3）
    返回:
        核心数据列结束列号（1-based）。若找不到业态列，回退为 ws.max_column
    """
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and isinstance(val, str) and "业态" in val:
            return col - 1
    # 回退：未找到业态列，整表都算核心列
    return ws.max_column


def is_core_empty(ws, row: int, core_end_col: int) -> bool:
    """
    检查指定行在核心数据列范围内是否全部为空。

    参数:
        ws: 工作表对象
        row: 行号
        core_end_col: 核心数据列结束列号
    返回:
        True 如果核心列全部为空
    """
    for col in range(1, core_end_col + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            return False
    return True


def find_trailing_empty_rows(ws, data_start_row: int, core_end_col: int) -> list:
    """
    从工作表底部向上扫描，收集尾部连续空行。

    参数:
        ws: 工作表对象
        data_start_row: 数据起始行号（表头下一行）
        core_end_col: 核心数据列结束列号
    返回:
        尾部空行号列表（按原始行号降序排列，即从下往上）
    """
    trailing = []
    for row in range(ws.max_row, data_start_row - 1, -1):
        if is_core_empty(ws, row, core_end_col):
            trailing.append(row)
        else:
            break  # 遇到第一个非空行，停止
    return trailing


def cleanup_worksheet(ws, header_row: int = 3) -> int:
    """
    清理单个工作表的尾部空行。

    参数:
        ws: 工作表对象
        header_row: 表头所在行号
    返回:
        删除的空行数
    """
    core_end = find_core_end_col(ws, header_row)
    if core_end < 1:
        return 0

    data_start = header_row + 1  # 数据从表头下一行开始
    trailing = find_trailing_empty_rows(ws, data_start, core_end)

    if not trailing:
        return 0

    # 从下往上逐行删除，避免行号变化导致偏移
    for row in trailing:
        ws.delete_rows(row, 1)

    return len(trailing)


def cleanup_file(input_path: str, output_path: str | None = None) -> str:
    """
    清理 Excel 文件中所有工作表的尾部空行。

    参数:
        input_path: 输入文件路径
        output_path: 输出文件路径（默认: 输入文件名_clean.xlsx）
    返回:
        输出文件路径
    """
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_clean{ext}"

    print(f"加载: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    total_deleted = 0
    for ws in wb.worksheets:
        deleted = cleanup_worksheet(ws)
        if deleted > 0:
            total_deleted += deleted
            print(f"  {ws.title}: 删除 {deleted} 行尾部空行 "
                  f"(剩 {ws.max_row} 行)")

    if total_deleted == 0:
        print("  未发现尾部空行，无需清理。")
    else:
        print(f"  合计删除 {total_deleted} 行。")

    wb.save(output_path)
    print(f"保存: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("用法: python cleanup_empty_rows.py input.xlsx [output.xlsx]")
        print("示例: python cleanup_empty_rows.py 综合楼_紧凑.xlsx 综合楼_最终.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(input_path):
        print(f"错误: 文件不存在 — {input_path}")
        sys.exit(1)

    cleanup_file(input_path, output_path)


if __name__ == "__main__":
    main()
