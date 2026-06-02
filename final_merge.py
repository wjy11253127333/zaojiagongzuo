from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy
import re

# 文件路径
excel_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx'
output_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx'

print("=" * 80)
print("最终修复 - 正确处理教学楼数据")
print("=" * 80)

# 重新加载原始文件
wb = load_workbook(excel_path, data_only=False)
ws_com = wb['表-08 分部分项工程和单价措施项目清单-综合楼']
ws_edu = wb['表-08 分部分项工程和单价措施项目清单-教学楼']

# 创建新的清单汇总表
print("\n1. 创建清单汇总表...")
if '清单汇总' in wb.sheetnames:
    del wb['清单汇总']

wb.create_sheet('清单汇总', 0)
ws_summary = wb['清单汇总']

# 复制综合楼（从第1行到第266行）
print("\n2. 复制综合楼工作表...")
for row in range(1, ws_com.max_row + 1):
    for col in range(1, ws_com.max_column + 1):
        source_cell = ws_com.cell(row=row, column=col)
        target_cell = ws_summary.cell(row=row, column=col)

        # 复制值或公式
        if source_cell.data_type == 'f':
            target_cell.value = source_cell.value
        else:
            target_cell.value = source_cell.value

        # 复制格式
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy.copy(source_cell.alignment)

# 复制列宽
for col_letter in ws_com.column_dimensions:
    ws_summary.column_dimensions[col_letter].width = ws_com.column_dimensions[col_letter].width

# 复制行高
for row in ws_com.row_dimensions:
    if row in ws_com.row_dimensions:
        ws_summary.row_dimensions[row].height = ws_com.row_dimensions[row].height

# 复制合并单元格
for merged_range in ws_com.merged_cells.ranges:
    ws_summary.merge_cells(str(merged_range))

print(f"   综合楼已复制: {ws_com.max_row} 行")

# 确定教学楼数据起始行
edu_start_row = 268  # 从第268行开始（跳过综合楼的266行+1行分隔+1行）

print(f"\n3. 处理教学楼数据...")

# 调整公式中的行号引用
def adjust_formula_row(formula, offset):
    """调整公式中的行号引用"""
    if not formula or not formula.startswith('='):
        return formula

    # 匹配单元格引用（A1, B2, AA10等）
    pattern = r'([A-Z]+)(\d+)'

    def replace_func(match):
        col_letter = match.group(1)
        row_num = int(match.group(2))
        new_row_num = row_num + offset
        return f"{col_letter}{new_row_num}"

    # 替换公式中的所有行号
    adjusted_formula = re.sub(pattern, replace_func, formula)
    return adjusted_formula

# 只复制教学楼的数据部分（第4行开始）
# 跳过前3行的标题和表头
edu_data_start = 4
edu_offset = edu_start_row - edu_data_start  # 偏移量 = 268 - 4 = 264

print(f"   教学楼数据起始行（原表）: 第{edu_data_start}行")
print(f"   教学楼数据起始行（汇总表）: 第{edu_start_row}行")
print(f"   行号偏移量: {edu_offset}")

for row in range(edu_data_start, ws_edu.max_row + 1):
    for col in range(1, ws_edu.max_column + 1):
        source_cell = ws_edu.cell(row=row, column=col)
        target_row = row + edu_offset  # 行号 + 264
        target_cell = ws_summary.cell(row=target_row, column=col)

        # 复制值或公式
        if source_cell.data_type == 'f':
            # 调整公式中的行号引用
            adjusted_formula = adjust_formula_row(source_cell.value, edu_offset)
            target_cell.value = adjusted_formula
        else:
            target_cell.value = source_cell.value

        # 复制格式
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy.copy(source_cell.alignment)

    # 复制行高
    if row in ws_edu.row_dimensions:
        target_row = row + edu_offset
        ws_summary.row_dimensions[target_row].height = ws_edu.row_dimensions[row].height

# 复制教学楼的合并单元格（调整行号）
for merged_range in ws_edu.merged_cells.ranges:
    min_col, min_row, max_col, max_row = merged_range.bounds
    if min_row >= edu_data_start:  # 只复制数据部分的合并单元格
        new_min_row = min_row + edu_offset
        new_max_row = max_row + edu_offset
        new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
        try:
            ws_summary.merge_cells(new_range)
        except:
            pass

# 添加教学楼分隔标题（在第267行）
print(f"\n4. 添加教学楼分隔标题...")
ws_summary.cell(row=267, column=1).value = "教学楼数据"
ws_summary.cell(row=267, column=1).font = copy.copy(ws_summary.cell(row=3, column=1).font)

# 验证结果
print("\n5. 验证结果...")
summary_formula_count = 0
for row in range(1, ws_summary.max_row + 1):
    for col in range(1, ws_summary.max_column + 1):
        if ws_summary.cell(row=row, column=col).data_type == 'f':
            summary_formula_count += 1

print(f"   汇总表总行数: {ws_summary.max_row}")
print(f"   汇总表总列数: {ws_summary.max_column}")
print(f"   汇总表公式数: {summary_formula_count}")
print(f"   预期公式数: 5784 (3003 + 2781)")

# 检查关键行
print(f"\n6. 检查关键行...")

# 综合楼汇总行
print(f"\n   综合楼汇总行 (第4行):")
cell_v4 = ws_summary.cell(row=4, column=22)  # V列
cell_w4 = ws_summary.cell(row=4, column=23)  # W列
print(f"     V4: {cell_v4.value if cell_v4.value else 'None'}")
print(f"     W4: {cell_w4.value if cell_w4.value else 'None'}")

# 教学楼汇总行
print(f"\n   教学楼汇总行 (第268行):")
cell_t268 = ws_summary.cell(row=268, column=20)  # T列
cell_u268 = ws_summary.cell(row=268, column=21)  # U列
cell_x268 = ws_summary.cell(row=268, column=24)  # X列
print(f"     T268: {cell_t268.value if cell_t268.value else 'None'}")
print(f"     U268: {cell_u268.value if cell_u268.value else 'None'}")
print(f"     X268: {cell_x268.value if cell_x268.value else 'None'}")

# 教学楼第一行数据
print(f"\n   教学楼第一行数据 (第269行):")
cell_i269 = ws_summary.cell(row=269, column=9)  # I列
cell_j269 = ws_summary.cell(row=269, column=10)  # J列
cell_k269 = ws_summary.cell(row=269, column=11)  # K列
print(f"     I269: {cell_i269.value if cell_i269.value else 'None'}")
print(f"     J269: {cell_j269.value if cell_j269.value else 'None'}")
print(f"     K269: {cell_k269.value if cell_k269.value else 'None'}")

# 保存文件
print("\n7. 保存文件...")
temp_output = r'c:\Users\aa\Desktop\造价清单文件标准化\temp_copy.xlsx'
wb.save(temp_output)

import shutil
shutil.copy(temp_output, output_path)

print(f"   文件已保存到: {output_path}")

print("\n" + "=" * 80)
print("修复完成！")
print("=" * 80)
print(f"\n✓ 清单汇总表结构:")
print(f"  - 综合楼数据: 第1行 - 第266行")
print(f"  - 分隔标题: 第267行")
print(f"  - 教学楼汇总行: 第268行")
print(f"  - 教学楼数据: 第269行 - 第508行")
print(f"✓ 公式总数: {summary_formula_count}")
print(f"✓ 公式引用已调整")
print(f"✓ 输出文件: {output_path}")
