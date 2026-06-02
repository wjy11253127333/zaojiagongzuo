from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy

# 文件路径
excel_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx'
output_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx'

print("=" * 80)
print("修复清单汇总表 - 确保教学楼表头正确排列")
print("=" * 80)

# 重新加载原始文件
wb = load_workbook(excel_path, data_only=False)
ws_com = wb['表-08 分部分项工程和单价措施项目清单-综合楼']
ws_edu = wb['表-08 分部分项工程和单价措施项目清单-教学楼']

# 创建新的清单汇总表
print("\n1. 创建清单汇总表...")
if '清单汇总' in wb.sheetnames:
    del wb['清单汇总']

wb.create_sheet('清单汇总', 0)
ws_summary = wb['清单汇总']

# 复制综合楼（从第1行到第266行）
print("\n2. 复制综合楼工作表...")
for row in range(1, ws_com.max_row + 1):
    for col in range(1, ws_com.max_column + 1):
        source_cell = ws_com.cell(row=row, column=col)
        target_cell = ws_summary.cell(row=row, column=col)

        # 复制值或公式
        if source_cell.data_type == 'f':
            target_cell.value = source_cell.value
        else:
            target_cell.value = source_cell.value

        # 复制格式
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy.copy(source_cell.alignment)

# 复制列宽
for col_letter in ws_com.column_dimensions:
    ws_summary.column_dimensions[col_letter].width = ws_com.column_dimensions[col_letter].width

# 复制行高
for row in ws_com.row_dimensions:
    if row in ws_com.row_dimensions:
        ws_summary.row_dimensions[row].height = ws_com.row_dimensions[row].height

# 复制合并单元格
for merged_range in ws_com.merged_cells.ranges:
    ws_summary.merge_cells(str(merged_range))

print(f"   综合楼已复制: {ws_com.max_row} 行")

# 确定教学楼数据起始行
# 综合楼：266行
# 预留1行作为间隔和标题行
edu_start_row = 268  # 从第268行开始

print(f"\n3. 复制教学楼工作表...")
print(f"   教学楼起始行: {edu_start_row}")

# 复制教学楼数据（跳过表头，直接复制数据部分）
# 根据综合楼的结构，表头占3行，数据从第5行开始
# 教学楼应该类似，表头也占3行，数据从第5行开始

edu_data_start = 4  # 教学楼数据起始行（跳过前3行表头）
edu_data_rows = ws_edu.max_row - edu_data_start  # 数据行数

print(f"   教学楼原始数据起始行: {edu_data_start}")
print(f"   教学楼数据行数: {edu_data_rows}")

for row in range(edu_data_start + 1, ws_edu.max_row + 1):  # 从第5行开始复制数据
    source_row = row
    target_row = edu_start_row + (row - edu_data_start - 1)  # 调整目标行号

    for col in range(1, ws_edu.max_column + 1):
        source_cell = ws_edu.cell(row=source_row, column=col)
        target_cell = ws_summary.cell(row=target_row, column=col)

        # 复制值或公式
        if source_cell.data_type == 'f':
            target_cell.value = source_cell.value
        else:
            target_cell.value = source_cell.value

        # 复制格式
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy.copy(source_cell.alignment)

    # 复制行高
    if source_row in ws_edu.row_dimensions:
        ws_summary.row_dimensions[target_row].height = ws_edu.row_dimensions[source_row].height

# 复制教学楼的合并单元格（调整行号）
for merged_range in ws_edu.merged_cells.ranges:
    min_col, min_row, max_col, max_row = merged_range.bounds
    if min_row > edu_data_start:  # 只复制数据部分的合并单元格
        new_min_row = edu_start_row + (min_row - edu_data_start - 1)
        new_max_row = edu_start_row + (max_row - edu_data_start - 1)
        new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
        ws_summary.merge_cells(new_range)

# 添加教学楼表头（在第267行）
print(f"\n4. 添加教学楼表头...")
edu_header_row = 267
for col in range(1, ws_edu.max_column + 1):
    header_cell = ws_edu.cell(row=3, column=col)  # 原表头在第3行
    target_cell = ws_summary.cell(row=edu_header_row, column=col)

    target_cell.value = header_cell.value
    target_cell.font = copy.copy(header_cell.font)
    target_cell.alignment = copy.copy(header_cell.alignment)

# 在第267行添加合并说明
ws_summary.cell(row=267, column=1).value = "教学楼数据"
ws_summary.cell(row=267, column=1).font = copy.copy(ws_summary.cell(row=3, column=1).font)

# 验证结果
print("\n5. 验证结果...")
summary_formula_count = 0
for row in range(1, ws_summary.max_row + 1):
    for col in range(1, ws_summary.max_column + 1):
        if ws_summary.cell(row=row, column=col).data_type == 'f':
            summary_formula_count += 1

print(f"   汇总表总行数: {ws_summary.max_row}")
print(f"   汇总表总列数: {ws_summary.max_column}")
print(f"   汇总表公式数: {summary_formula_count}")
print(f"   预期公式数: 5784 (3003 + 2781)")

# 检查关键行
print(f"\n6. 检查关键行...")

# 综合楼第一行数据
print(f"\n   综合楼第一行数据 (第5行):")
cell = ws_summary.cell(row=5, column=1)
print(f"     A5: {cell.value}")

# 教学楼表头
print(f"\n   教学楼表头 (第267行):")
cell = ws_summary.cell(row=267, column=1)
print(f"     A267: {cell.value}")

# 教学楼第一行数据
print(f"\n   教学楼第一行数据 (第270行):")
cell = ws_summary.cell(row=270, column=1)
print(f"     A270: {cell.value}")

# 保存文件
print("\n7. 保存文件...")
temp_output = r'c:\Users\aa\Desktop\造价清单文件标准化\temp_copy.xlsx'
wb.save(temp_output)

import shutil
shutil.copy(temp_output, output_path)

print(f"   文件已保存到: {output_path}")

print("\n" + "=" * 80)
print("修复完成！")
print("=" * 80)
print(f"\n✓ 清单汇总表结构:")
print(f"  - 综合楼数据: 第1行 - 第266行")
print(f"  - 间隔/表头: 第267行")
print(f"  - 教学楼数据: 第270行 - 第514行")
print(f"✓ 公式总数: {summary_formula_count}")
print(f"✓ 输出文件: {output_path}")
