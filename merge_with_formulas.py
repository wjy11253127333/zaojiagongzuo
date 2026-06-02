from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy

# 文件路径
excel_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx'
output_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx'

print("=" * 80)
print("开始处理：提取公式并创建清单汇总表")
print("=" * 80)

# 加载工作簿，保留公式
print("\n1. 加载Excel文件（保留公式）...")
wb = load_workbook(excel_path, data_only=False)

# 定义工作表名称
sheet_com = '表-08 分部分项工程和单价措施项目清单-综合楼'
sheet_edu = '表-08 分部分项工程和单价措施项目清单-教学楼'
sheet_summary = '清单汇总'

# 提取公式信息
def extract_formulas_from_sheet(ws, sheet_name):
    """提取工作表中的所有公式"""
    formulas = {}
    print(f"\n   分析工作表: {sheet_name}")
    print(f"   工作表大小: {ws.max_row} 行 x {ws.max_column} 列")

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f':  # 如果是公式
                cell_addr = f"{get_column_letter(col)}{row}"
                formulas[cell_addr] = {
                    'row': row,
                    'col': col,
                    'formula': cell.value,
                    'data_type': cell.data_type
                }

    print(f"   发现 {len(formulas)} 个公式单元格")
    return formulas

# 提取两个工作表的公式
print("\n2. 提取公式信息...")
ws_com = wb[sheet_com]
ws_edu = wb[sheet_edu]

formulas_com = extract_formulas_from_sheet(ws_com, sheet_com)
formulas_edu = extract_formulas_from_sheet(ws_edu, sheet_edu)

print(f"\n   综合楼工作表公式数: {len(formulas_com)}")
print(f"   教学楼工作表公式数: {len(formulas_edu)}")

# 创建新的清单汇总表
print("\n3. 创建清单汇总表...")

# 检查是否已存在汇总表，如果存在则删除
if sheet_summary in wb.sheetnames:
    print(f"   删除旧的 '{sheet_summary}' 工作表...")
    del wb[sheet_summary]

# 在最后创建一个新的空白工作表
wb.create_sheet(sheet_summary, 0)  # 在最前面插入
ws_summary = wb[sheet_summary]

print(f"   已创建新的 '{sheet_summary}' 工作表")

# 定义合并策略
# 综合楼和教学楼的列对应关系需要确定
# 根据分析，两个工作表的列结构有差异，需要统一

print("\n4. 设计清单汇总表结构...")

# 综合楼列结构（35列）：
# A-Z, AA-AI (1-35)
# 序号(A), 项目编码(B), 项目名称(C), 项目特征描述(D), 计量单位(E),
# 工程量相关(F-L): 合同工程量, 送审工程量, 审核工程量, 施工单位复核, 增减工程量等
# 单价相关(N-S): 投标单价, 预算备案单价, 偏差率, P1*L*(1-15%), P1*(1+15%), 增减部分单价
# 合价相关(T-Y): 合同合价, 结算合价, 审核合价, 增减部分金额, 核增减金额, 备注(Z)

# 教学楼列结构（27列）：
# A-Y (1-27)
# 类似结构但列数不同

# 合并策略：
# 1. 复制综合楼的完整表头和数据到汇总表
# 2. 在综合楼数据后追加教学楼的数据
# 3. 调整行号引用

# 复制综合楼工作表的数据和公式
print("\n5. 复制综合楼工作表数据...")

# 复制单元格值和公式
def copy_sheet_data(source_ws, target_ws, start_row=1, offset_row=0, copy_formulas=True):
    """复制工作表数据，支持公式引用调整"""
    copied_cells = 0
    formula_cells = 0

    for row in range(1, source_ws.max_row + 1):
        for col in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=row, column=col)
            target_row = row + offset_row
            target_cell = target_ws.cell(row=target_row, column=col)

            # 复制值或公式
            if source_cell.data_type == 'f' and copy_formulas:
                # 复制公式
                target_cell.value = source_cell.value
                formula_cells += 1
            else:
                # 复制值
                target_cell.value = source_cell.value

            # 复制格式
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

            copied_cells += 1

    return copied_cells, formula_cells

# 复制综合楼（offset_row=0）
copied, formulas = copy_sheet_data(ws_com, ws_summary, offset_row=0)
print(f"   综合楼数据已复制: {copied} 个单元格, 其中 {formulas} 个公式")

# 确定教学楼数据的起始行
# 综合楼的最后一行 + 1行间隔
com_last_row = ws_com.max_row
edu_start_row = com_last_row + 2  # 空一行分隔

print(f"\n6. 复制教学楼工作表数据...")
print(f"   综合楼最后一行: {com_last_row}")
print(f"   教学楼数据起始行: {edu_start_row}")

# 注意：教学楼的列数较少（27列），需要特殊处理
# 我们将教学楼的数据复制到综合楼的列结构中

# 复制教学楼的数据
edu_copied, edu_formulas = 0, 0

for row in range(1, ws_edu.max_row + 1):
    for col in range(1, ws_edu.max_column + 1):
        source_cell = ws_edu.cell(row=row, column=col)
        target_row = edu_start_row + row - 1
        target_cell = ws_summary.cell(row=target_row, column=col)

        # 复制值或公式
        if source_cell.data_type == 'f':
            target_cell.value = source_cell.value
            edu_formulas += 1
        else:
            target_cell.value = source_cell.value

        # 复制格式
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy.copy(source_cell.protection)
            target_cell.alignment = copy.copy(source_cell.alignment)

        edu_copied += 1

print(f"   教学楼数据已复制: {edu_copied} 个单元格, 其中 {edu_formulas} 个公式")

# 复制列宽
print("\n7. 复制列宽...")
for col in range(1, max(ws_com.max_column, ws_edu.max_column) + 1):
    col_letter = get_column_letter(col)
    if col_letter in ws_com.column_dimensions:
        ws_summary.column_dimensions[col_letter].width = ws_com.column_dimensions[col_letter].width
    elif col_letter in ws_edu.column_dimensions:
        ws_summary.column_dimensions[col_letter].width = ws_edu.column_dimensions[col_letter].width

print("   列宽复制完成")

# 复制行高
print("8. 复制行高...")
for row in range(1, com_last_row):
    if row in ws_com.row_dimensions:
        ws_summary.row_dimensions[row].height = ws_com.row_dimensions[row].height

for row in range(1, ws_edu.max_row + 1):
    target_row = edu_start_row + row - 1
    if row in ws_edu.row_dimensions:
        ws_summary.row_dimensions[target_row].height = ws_edu.row_dimensions[row].height

print("   行高复制完成")

# 复制合并单元格
print("9. 复制合并单元格...")
for merged_range in ws_com.merged_cells.ranges:
    ws_summary.merge_cells(str(merged_range))

for merged_range in ws_edu.merged_cells.ranges:
    # 需要调整行号
    min_col, min_row, max_col, max_row = merged_range.bounds
    new_min_row = min_row + edu_start_row - 1
    new_max_row = max_row + edu_start_row - 1
    new_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
    ws_summary.merge_cells(new_range)

print("   合并单元格复制完成")

# 验证公式数量
print("\n10. 验证汇总表公式...")
summary_formula_count = 0
for row in range(1, ws_summary.max_row + 1):
    for col in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=row, column=col)
        if cell.data_type == 'f':
            summary_formula_count += 1

print(f"   汇总表公式总数: {summary_formula_count}")
print(f"   预期公式数: {len(formulas_com) + len(formulas_edu)}")

if summary_formula_count >= len(formulas_com) + len(formulas_edu):
    print("   ✓ 公式数量验证通过")
else:
    print(f"   ⚠ 公式数量可能不完整，差异: {len(formulas_com) + len(formulas_edu) - summary_formula_count}")

# 保存文件
print("\n11. 保存文件...")
temp_output = r'c:\Users\aa\Desktop\造价清单文件标准化\temp_copy.xlsx'
wb.save(temp_output)

import shutil
shutil.copy(temp_output, output_path)
print(f"   文件已保存到: {output_path}")

print("\n" + "=" * 80)
print("处理完成！")
print("=" * 80)
print(f"\n总结：")
print(f"  - 综合楼原始公式数: {len(formulas_com)}")
print(f"  - 教学楼原始公式数: {len(formulas_edu)}")
print(f"  - 汇总表公式总数: {summary_formula_count}")
print(f"  - 输出文件: {output_path}")
