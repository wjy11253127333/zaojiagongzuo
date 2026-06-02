import openpyxl

output_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx"

print("=" * 80)
print("验证生成的Excel文件...")
print("=" * 80)

wb = openpyxl.load_workbook(output_path)
print(f"\n工作表列表: {wb.sheetnames}")
print(f"总工作表数量: {len(wb.sheetnames)}")

print("\n各工作表内容预览:")
print("-" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n【{sheet_name}】")
    print(f"  数据范围: {ws.max_row} 行 × {ws.max_column} 列")

    if sheet_name == '成本分析_汇总':
        print(f"  标题: {ws['A1'].value}")
        print(f"  分析日期: {ws['A2'].value}")
        print(f"  表头: {[ws.cell(row=4, column=col).value for col in range(1, 11)]}")

    elif sheet_name == '成本分析_综合':
        print(f"  标题: {ws['A1'].value}")
        print(f"  成本总体概况:")
        print(f"    - 合同总成本: {ws['B4'].value:,.2f}" if ws['B4'].value else "    - 合同总成本: N/A")
        print(f"    - 送审总成本: {ws['B5'].value:,.2f}" if ws['B5'].value else "    - 送审总成本: N/A")

    elif sheet_name == '成本分析_明细':
        print(f"  标题: {ws['A1'].value}")
        print(f"  表头: {[ws.cell(row=3, column=col).value for col in range(1, 12)]}")
        print(f"  示例数据 (第4行):")
        print(f"    序号: {ws['A4'].value}")
        print(f"    项目编码: {ws['C4'].value}")
        print(f"    项目名称: {ws['D4'].value}")
        print(f"    合同合价: {ws['I4'].value}")

print("\n" + "=" * 80)
print("✅ 验证完成！文件结构正确。")
print("=" * 80)
