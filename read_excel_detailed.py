import openpyxl
import pandas as pd
import json

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"
output_path = r"c:\Users\aa\Desktop\造价清单文件标准化\excel_structure.txt"

try:
    wb = openpyxl.load_workbook(excel_path)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"工作表数量: {len(wb.sheetnames)}\n")
        f.write(f"工作表名称: {wb.sheetnames}\n\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"\n{'='*80}\n")
            f.write(f"工作表: {sheet_name}\n")
            f.write(f"数据范围: {ws.max_row} 行 × {ws.max_column} 列\n")
            f.write(f"{'='*80}\n\n")
            
            # 读取前30行数据
            f.write("数据预览 (前30行):\n")
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
                if any(cell is not None for cell in row):
                    display_row = row[:15]  # 只显示前15列
                    f.write(f"行 {i}: {[str(cell)[:40] if cell else '' for cell in display_row]}\n")
            
            # 检查合并单元格
            if ws.merged_cells:
                f.write(f"\n合并单元格数量: {len(ws.merged_cells.ranges)}\n")
        
        # 使用pandas读取更多信息
        f.write("\n\n" + "="*80 + "\n")
        f.write("使用pandas读取数据:\n")
        f.write("="*80 + "\n")
        
        # 读取所有工作表
        xls = pd.ExcelFile(excel_path)
        for sheet in xls.sheet_names:
            f.write(f"\n工作表: {sheet}\n")
            df = pd.read_excel(excel_path, sheet_name=sheet, header=None)
            f.write(f"形状: {df.shape}\n")
            f.write(f"列名/表头:\n{df.head(10).to_string()}\n")
    
    print(f"Excel结构信息已保存到: {output_path}")
    
except Exception as e:
    print(f"读取Excel时出错: {e}")
    import traceback
    traceback.print_exc()
