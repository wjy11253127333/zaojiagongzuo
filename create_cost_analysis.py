import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from collections import defaultdict

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"
output_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx"

print("=" * 80)
print("开始处理Excel文件...")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path)
print(f"原始工作表: {wb.sheetnames}")

df_summary = pd.read_excel(excel_path, sheet_name='清单汇总', header=0)

df_clean = df_summary.copy()
df_clean = df_clean[df_clean['项目编码'].notna() & (df_clean['项目编码'] != '')].copy()

def safe_float(value):
    try:
        return float(value) if pd.notna(value) else 0.0
    except:
        return 0.0

df_clean['合同合价'] = df_clean['合价（元）_合同合价（元）_J'].apply(safe_float)
df_clean['结算合价'] = df_clean['合价（元）_结算合价（元）_K'].apply(safe_float)
df_clean['审核合价'] = df_clean['合价（元）_审核合价（元）_L'].apply(safe_float)
df_clean['合同工程量'] = df_clean['工程量_合同工程量_A'].apply(safe_float)
df_clean['送审工程量'] = df_clean['工程量_送审工程量_B'].apply(safe_float)
df_clean['审核工程量'] = df_clean['工程量_审核工程量_C'].apply(safe_float)

print(f"清洗后有效数据行数: {len(df_clean)}")

summary_by_category = df_clean.groupby('业态').agg({
    '项目编码': 'count',
    '合同合价': 'sum',
    '结算合价': 'sum',
    '审核合价': 'sum',
    '合同工程量': 'sum'
}).rename(columns={'项目编码': '项目数量'})

summary_by_category['偏差（结算-合同）'] = summary_by_category['结算合价'] - summary_by_category['合同合价']
summary_by_category['偏差率（%）'] = (summary_by_category['偏差（结算-合同）'] / summary_by_category['合同合价'] * 100).round(2)
summary_by_category['成本降低额'] = summary_by_category['合同合价'] - summary_by_category['审核合价']
summary_by_category['成本降低率（%）'] = (summary_by_category['成本降低额'] / summary_by_category['合同合价'] * 100).round(2)

print("\n按业态统计结果:")
print(summary_by_category)

def create_cost_summary_sheet(wb):
    print("\n创建 Sheet1: 分部分项工程成本汇总...")
    if '成本分析_汇总' in wb.sheetnames:
        del wb['成本分析_汇总']
    ws = wb.create_sheet('成本分析_汇总', 0)

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')

    title = ws.cell(row=1, column=1, value='分部分项工程成本汇总分析')
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:H1')

    ws.cell(row=2, column=1, value='分析日期：2026-05-27')
    ws.cell(row=2, column=5, value='单位：元')

    headers = ['序号', '业态', '项目数量', '合同合价', '结算合价', '审核合价', '偏差（结算-合同）', '偏差率（%）', '成本降低额', '成本降低率（%）']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 5
    total_contract = 0
    total_settlement = 0
    total_audit = 0
    total_count = 0

    for idx, (category, data) in enumerate(summary_by_category.iterrows(), 1):
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=category.replace('表-08 分部分项工程和单价措施项目清单-', ''))
        ws.cell(row=row_num, column=3, value=int(data['项目数量']))
        ws.cell(row=row_num, column=4, value=data['合同合价'])
        ws.cell(row=row_num, column=5, value=data['结算合价'])
        ws.cell(row=row_num, column=6, value=data['审核合价'] if data['审核合价'] > 0 else '-')
        ws.cell(row=row_num, column=7, value=data['偏差（结算-合同）'])
        ws.cell(row=row_num, column=8, value=f"{data['偏差率（%）']:.2f}%")
        ws.cell(row=row_num, column=9, value=data['成本降低额'] if data['审核合价'] > 0 else '-')
        ws.cell(row=row_num, column=10, value=f"{data['成本降低率（%）']:.2f}%" if data['审核合价'] > 0 else '-')

        total_contract += data['合同合价']
        total_settlement += data['结算合价']
        total_audit += data['审核合价'] if data['审核合价'] > 0 else 0
        total_count += int(data['项目数量'])
        row_num += 1

    ws.cell(row=row_num, column=1, value='')
    ws.cell(row=row_num, column=2, value='合计')
    ws.cell(row=row_num, column=3, value=total_count)
    ws.cell(row=row_num, column=4, value=total_contract)
    ws.cell(row=row_num, column=5, value=total_settlement)
    ws.cell(row=row_num, column=6, value=total_audit if total_audit > 0 else '-')
    ws.cell(row=row_num, column=7, value=total_settlement - total_contract)

    total_deviation_rate = ((total_settlement - total_contract) / total_contract * 100) if total_contract > 0 else 0
    ws.cell(row=row_num, column=8, value=f"{total_deviation_rate:.2f}%")

    total_reduction = total_contract - total_audit
    ws.cell(row=row_num, column=9, value=total_reduction if total_audit > 0 else '-')
    total_reduction_rate = (total_reduction / total_contract * 100) if total_contract > 0 and total_audit > 0 else 0
    ws.cell(row=row_num, column=10, value=f"{total_reduction_rate:.2f}%" if total_audit > 0 else '-')

    for col in range(1, 11):
        ws.cell(row=row_num, column=col).font = Font(bold=True)
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for col in range(4, 10):
        for row in range(5, row_num + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15

    return ws

def create_comprehensive_analysis_sheet(wb):
    print("创建 Sheet2: 综合成本汇总分析...")
    if '成本分析_综合' in wb.sheetnames:
        del wb['成本分析_综合']
    ws = wb.create_sheet('成本分析_综合', 1)

    title = ws.cell(row=1, column=1, value='综合成本汇总分析')
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')

    ws.cell(row=3, column=1, value='一、成本总体概况')
    ws.cell(row=3, column=1).font = Font(bold=True, size=12)

    total_contract = summary_by_category['合同合价'].sum()
    total_settlement = summary_by_category['结算合价'].sum()
    total_audit = summary_by_category['审核合价'].sum()

    ws.cell(row=4, column=1, value='合同总成本（元）')
    ws.cell(row=4, column=2, value=total_contract)
    ws.cell(row=4, column=2).number_format = '#,##0.00'

    ws.cell(row=5, column=1, value='送审总成本（元）')
    ws.cell(row=5, column=2, value=total_settlement)
    ws.cell(row=5, column=2).number_format = '#,##0.00'

    ws.cell(row=6, column=1, value='审核总成本（元）')
    ws.cell(row=6, column=2, value=total_audit if total_audit > 0 else '数据不完整')
    if total_audit > 0:
        ws.cell(row=6, column=2).number_format = '#,##0.00'

    ws.cell(row=7, column=1, value='送审与合同偏差（元）')
    ws.cell(row=7, column=2, value=total_settlement - total_contract)
    ws.cell(row=7, column=2).number_format = '#,##0.00'

    ws.cell(row=8, column=1, value='送审与合同偏差率（%）')
    deviation_rate = ((total_settlement - total_contract) / total_contract * 100) if total_contract > 0 else 0
    ws.cell(row=8, column=2, value=deviation_rate)
    ws.cell(row=8, column=2).number_format = '0.00%'

    ws.cell(row=10, column=1, value='二、成本降低效果分析（基于审核成本）')
    ws.cell(row=10, column=1).font = Font(bold=True, size=12)

    if total_audit > 0:
        reduction_amount = total_contract - total_audit
        reduction_rate = reduction_amount / total_contract * 100

        ws.cell(row=11, column=1, value='项目施工成本降低额（元）')
        ws.cell(row=11, column=2, value=reduction_amount)
        ws.cell(row=11, column=2).number_format = '#,##0.00'

        ws.cell(row=12, column=1, value='项目施工成本降低率（%）')
        ws.cell(row=12, column=2, value=reduction_rate)
        ws.cell(row=12, column=2).number_format = '0.00%'
    else:
        ws.cell(row=11, column=1, value='成本降低效果分析')
        ws.cell(row=11, column=2, value='审核数据不完整，无法计算')

    ws.cell(row=14, column=1, value='三、按业态成本占比')
    ws.cell(row=14, column=1).font = Font(bold=True, size=12)

    row = 15
    for category, data in summary_by_category.iterrows():
        category_name = category.replace('表-08 分部分项工程和单价措施项目清单-', '')
        proportion = (data['合同合价'] / total_contract * 100) if total_contract > 0 else 0

        ws.cell(row=row, column=1, value=f'{category_name}合同占比（%）')
        ws.cell(row=row, column=2, value=proportion)
        ws.cell(row=row, column=2).number_format = '0.00%'

        ws.cell(row=row, column=3, value=f'{category_name}合同金额（元）')
        ws.cell(row=row, column=4, value=data['合同合价'])
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        row += 1

    ws.cell(row=row + 1, column=1, value='四、成本管理方法论说明')
    ws.cell(row=row + 1, column=1).font = Font(bold=True, size=12)

    ws.cell(row=row + 2, column=1, value='1. 本分析基于《施工成本管理》方法论框架')
    ws.cell(row=row + 3, column=1, value='2. 采用"三算对比"方法：合同成本、送审成本、审核成本')
    ws.cell(row=row + 4, column=1, value='3. 成本降低率 = (合同成本 - 审核成本) / 合同成本 × 100%')
    ws.cell(row=row + 5, column=1, value='4. 偏差率 = (送审成本 - 合同成本) / 合同成本 × 100%')

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    return ws

def create_detailed_data_sheet(wb, df):
    print("创建 Sheet3: 详细数据列表...")
    if '成本分析_明细' in wb.sheetnames:
        del wb['成本分析_明细']
    ws = wb.create_sheet('成本分析_明细', 2)

    title = ws.cell(row=1, column=1, value='分部分项工程成本明细数据')
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:K1')

    headers = ['序号', '业态', '项目编码', '项目名称', '计量单位', '合同工程量', '送审工程量', '审核工程量', '合同合价', '结算合价', '审核合价']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 4
    for idx, row_data in df.iterrows():
        ws.cell(row=row_num, column=1, value=idx + 1)
        ws.cell(row=row_num, column=2, value=row_data['业态'].replace('表-08 分部分项工程和单价措施项目清单-', ''))
        ws.cell(row=row_num, column=3, value=row_data['项目编码'])
        ws.cell(row=row_num, column=4, value=row_data['项目名称'])
        ws.cell(row=row_num, column=5, value=row_data['计量单位'])
        ws.cell(row=row_num, column=6, value=row_data['合同工程量'])
        ws.cell(row=row_num, column=7, value=row_data['送审工程量'])
        ws.cell(row=row_num, column=8, value=row_data['审核工程量'])
        ws.cell(row=row_num, column=9, value=row_data['合同合价'])
        ws.cell(row=row_num, column=10, value=row_data['结算合价'])
        ws.cell(row=row_num, column=11, value=row_data['审核合价'] if row_data['审核合价'] > 0 else '-')

        for col in range(6, 12):
            cell = ws.cell(row=row_num, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

        row_num += 1
        if row_num > 500:
            print(f"  已处理 {row_num - 4} 行，为控制文件大小限制为500行...")
            break

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18

    return ws

create_cost_summary_sheet(wb)
create_comprehensive_analysis_sheet(wb)
create_detailed_data_sheet(wb, df_clean)

print("\n保存Excel文件...")
wb.save(output_path)
print(f"✅ 文件已保存至: {output_path}")
print("=" * 80)
print("\n✅ 成本分析工作表创建完成！")
print(f"   - Sheet1: 成本分析_汇总（分部分项工程成本汇总）")
print(f"   - Sheet2: 成本分析_综合（综合成本汇总分析）")
print(f"   - Sheet3: 成本分析_明细（详细数据列表）")
print("=" * 80)
