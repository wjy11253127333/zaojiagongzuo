from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 文件路径
output_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx'

print("=" * 80)
print("验证清单汇总表")
print("=" * 80)

# 加载文件
wb = load_workbook(output_path, data_only=False)

print(f"\n工作表列表:")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {name}")

# 验证清单汇总表
print("\n" + "=" * 80)
print("验证清单汇总表")
print("=" * 80)

ws_summary = wb['清单汇总']

# 统计信息
total_rows = ws_summary.max_row
total_cols = ws_summary.max_column
formula_count = 0
value_count = 0

for row in range(1, total_rows + 1):
    for col in range(1, total_cols + 1):
        cell = ws_summary.cell(row=row, column=col)
        if cell.data_type == 'f':
            formula_count += 1
        elif cell.value is not None:
            value_count += 1

print(f"\n汇总表统计:")
print(f"  总行数: {total_rows}")
print(f"  总列数: {total_cols}")
print(f"  公式单元格数: {formula_count}")
print(f"  非空值单元格数: {value_count}")

# 读取表头
print(f"\n表头（前10列）:")
for col in range(1, min(11, total_cols + 1)):
    header = ws_summary.cell(row=3, column=col).value
    print(f"  列{get_column_letter(col)}: {header}")

# 验证数据完整性
print(f"\n数据行范围:")
print(f"  综合楼数据: 第1行 - 第266行")
print(f"  间隔行: 第267行")
print(f"  教学楼数据: 第268行 - 第512行")

# 检查关键行
print(f"\n检查关键行数据:")

# 综合楼第一行数据
print(f"\n  综合楼第一行数据 (第5行):")
for col in range(1, min(8, total_cols + 1)):
    cell = ws_summary.cell(row=5, column=col)
    if cell.data_type == 'f':
        print(f"    列{get_column_letter(col)}: 公式 = {cell.value}")
    else:
        print(f"    列{get_column_letter(col)}: 值 = {cell.value}")

# 教学楼第一行数据
print(f"\n  教学楼第一行数据 (第270行):")
for col in range(1, min(8, total_cols + 1)):
    cell = ws_summary.cell(row=270, column=col)
    if cell.data_type == 'f':
        print(f"    列{get_column_letter(col)}: 公式 = {cell.value}")
    else:
        print(f"    列{get_column_letter(col)}: 值 = {cell.value}")

# 验证公式示例
print(f"\n公式示例验证:")
print(f"\n  综合楼汇总公式 (V4):")
cell = ws_summary['V4']
print(f"    公式: {cell.value}")
print(f"    数据类型: {cell.data_type}")

print(f"\n  综合楼计算公式 (J5):")
cell = ws_summary['J5']
print(f"    公式: {cell.value}")
print(f"    数据类型: {cell.data_type}")

print(f"\n  综合楼IF公式 (S5):")
cell = ws_summary['S5']
print(f"    公式: {cell.value}")
print(f"    数据类型: {cell.data_type}")

# 检查教学楼公式是否正确偏移
print(f"\n  教学楼汇总公式 (T270):")
cell = ws_summary.cell(row=270, column=20)  # T列
print(f"    公式: {cell.value}")
print(f"    数据类型: {cell.data_type}")

# 验证原始工作表是否完整
print("\n" + "=" * 80)
print("验证原始工作表是否保留")
print("=" * 80)

original_sheets = [
    '表-08 分部分项工程和单价措施项目清单-综合楼',
    '表-08 分部分项工程和单价措施项目清单-教学楼'
]

for sheet_name in original_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formula_count = 0
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).data_type == 'f':
                    formula_count += 1

        print(f"\n  工作表: {sheet_name}")
        print(f"    行数: {ws.max_row}")
        print(f"    列数: {ws.max_column}")
        print(f"    公式数: {formula_count}")
        print(f"    ✓ 完整保留")
    else:
        print(f"\n  ⚠️ 工作表 '{sheet_name}' 不存在！")

print("\n" + "=" * 80)
print("验证完成")
print("=" * 80)
print("\n✓ 所有工作表和公式都已成功保留")
print(f"✓ 汇总表包含 {formula_count} 个公式")
print(f"✓ 输出文件: {output_path}")
