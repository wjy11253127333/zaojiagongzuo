from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import os

# 文件路径
excel_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx'

# 加载工作簿，保留公式
print("正在加载Excel文件（保留公式）...")
wb = load_workbook(excel_path, data_only=False)

# 查看所有工作表
print("\n=== 工作表列表 ===")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"{idx}. {sheet_name}")

# 分析两个目标工作表
target_sheets = [
    '表-08 分部分项工程和单价措施项目清单-综合楼',
    '表-08 分部分项工程和单价措施项目清单-教学楼'
]

formula_summary = {}

for sheet_name in target_sheets:
    print(f"\n{'='*60}")
    print(f"分析工作表：{sheet_name}")
    print('='*60)

    if sheet_name not in wb.sheetnames:
        print(f"⚠️ 工作表 '{sheet_name}' 不存在！")
        continue

    ws = wb[sheet_name]

    # 读取表头（前3行）
    print(f"\n【表头结构】")
    headers = []
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=3, column=col).value
        if header_value:
            headers.append((col, header_value))
            print(f"  列{get_column_letter(col)}: {header_value}")

    # 统计公式单元格
    formula_count = 0
    total_cells = 0
    data_row_count = 0

    formulas_detail = []

    print(f"\n【公式统计】")
    print(f"工作表最大行数: {ws.max_row}")
    print(f"工作表最大列数: {ws.max_column}")

    # 遍历所有单元格查找公式
    for row in range(1, ws.max_row + 1):
        row_formulas = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            total_cells += 1

            # 检查是否是公式单元格
            if cell.data_type == 'f':  # formula
                formula_count += 1
                formula_text = cell.value
                row_formulas.append({
                    'cell': f"{get_column_letter(col)}{row}",
                    'formula': formula_text
                })

        if row_formulas:
            data_row_count += 1
            formulas_detail.extend(row_formulas)

    print(f"总单元格数: {total_cells}")
    print(f"包含公式的单元格数: {formula_count}")
    print(f"包含公式的行数: {data_row_count}")

    # 显示部分公式示例
    print(f"\n【公式示例（前20个）】")
    for i, f in enumerate(formulas_detail[:20], 1):
        print(f"  {i}. {f['cell']}: {f['formula']}")

    if len(formulas_detail) > 20:
        print(f"  ... 还有 {len(formulas_detail) - 20} 个公式")

    # 统计公式类型
    formula_types = {}
    for f in formulas_detail:
        formula = f['formula']
        # 提取函数名
        if formula.startswith('='):
            formula = formula[1:]

        func_name = '其他'
        if 'SUM' in formula.upper():
            func_name = 'SUM'
        elif 'IF' in formula.upper():
            func_name = 'IF'
        elif 'ROUND' in formula.upper():
            func_name = 'ROUND'
        elif '*' in formula and '/' in formula:
            func_name = '混合运算'
        elif '*' in formula:
            func_name = '乘法'
        elif '/' in formula:
            func_name = '除法'
        elif '+' in formula or '-' in formula:
            func_name = '加减运算'

        formula_types[func_name] = formula_types.get(func_name, 0) + 1

    print(f"\n【公式类型统计】")
    for ftype, count in sorted(formula_types.items(), key=lambda x: x[1], reverse=True):
        print(f"  {ftype}: {count}个")

    # 保存到summary
    formula_summary[sheet_name] = {
        'formula_count': formula_count,
        'formulas': formulas_detail,
        'headers': headers,
        'max_row': ws.max_row,
        'max_column': ws.max_column
    }

# 保存公式详情到文件
output_file = r'c:\Users\aa\Desktop\造价清单文件标准化\formula_analysis.txt'
with open(output_file, 'w', encoding='utf-8') as f:
    f.write("Excel公式分析报告\n")
    f.write("=" * 80 + "\n\n")

    for sheet_name, info in formula_summary.items():
        f.write(f"\n工作表：{sheet_name}\n")
        f.write(f"-" * 80 + "\n")
        f.write(f"公式总数: {info['formula_count']}\n")
        f.write(f"工作表大小: {info['max_row']} 行 x {info['max_column']} 列\n\n")

        f.write("所有公式列表：\n")
        for i, formula in enumerate(info['formulas'], 1):
            f.write(f"{i:4d}. {formula['cell']:8s} : {formula['formula']}\n")

print(f"\n\n公式分析报告已保存到: {output_file}")
print("分析完成！")
