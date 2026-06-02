from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 文件路径
output_path = r'c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx'

print("=" * 80)
print("详细检查清单汇总表的公式位置")
print("=" * 80)

# 加载文件
wb = load_workbook(output_path, data_only=False)
ws_summary = wb['清单汇总']

print(f"\n检查第267-270行的所有单元格...")

# 检查第267-270行
for row in range(267, 271):
    print(f"\n第{row}行:")
    has_content = False
    for col in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=row, column=col)
        if cell.value is not None:
            has_content = True
            cell_type = "公式" if cell.data_type == 'f' else "值"
            print(f"  {get_column_letter(col)}{row}: [{cell_type}] {cell.value}")

    if not has_content:
        print(f"  (空行)")

print(f"\n检查第4行的公式...")
for col in range(1, ws_summary.max_column + 1):
    cell = ws_summary.cell(row=4, column=col)
    if cell.data_type == 'f':
        print(f"  {get_column_letter(col)}4: 公式 = {cell.value}")
