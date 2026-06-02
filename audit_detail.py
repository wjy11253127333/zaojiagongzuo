import openpyxl
import pandas as pd

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"

print("=" * 80)
print("详细检查综合楼工作表的数据关系...")
print("=" * 80)

df = pd.read_excel(excel_path, sheet_name='表-08 分部分项工程和单价措施项目清单-综合楼', header=2)

print(f"\n列名:")
for i, col in enumerate(df.columns):
    print(f"  {i+1}. {col}")

print(f"\n前20行数据（关键列）:")
key_cols = ['项目编码', '项目名称', '计量单位', '工程量_合同工程量_A', '单价（元）_投标单价（元）_P0', '合价（元）_合同合价（元）_J']
preview = df[key_cols].head(20)
print(preview)

print(f"\n\n验证单价 × 工程量 = 合价:")
for i in range(3, min(20, len(df))):
    qty = df.iloc[i]['工程量_合同工程量_A']
    price = df.iloc[i]['单价（元）_投标单价（元）_P0']
    total = df.iloc[i]['合价（元）_合同合价（元）_J']

    if pd.notna(qty) and pd.notna(price) and pd.notna(total):
        calc = qty * price
        diff = abs(calc - total)
        print(f"  行{i+3} (项目编码: {df.iloc[i]['项目编码']}):")
        print(f"    工程量: {qty:,.2f} × 单价: {price:,.2f} = 计算: {calc:,.2f}, 实际: {total:,.2f}")
        if diff > 1:
            print(f"    ⚠️  差异: {diff:,.2f}")
        else:
            print(f"    ✅ 匹配")

print("\n\n检查审核合价的有效性:")
audit_cols = ['项目编码', '项目名称', '合价（元）_合同合价（元）_J', '合价（元）_结算合价（元）_K', '合价（元）_审核合价（元）_L']
audit_data = df[df['合价（元）_审核合价（元）_L'].notna() & (df['合价（元）_审核合价（元）_L'] > 0)][audit_cols]
print(audit_data)

print("\n\n检查原始综合楼工作表是否有更多单价相关列:")
ws = openpyxl.load_workbook(excel_path)['表-08 分部分项工程和单价措施项目清单-综合楼']
print(f"总列数: {ws.max_column}")
print("前4行（表头区域）:")
for row in range(1, 5):
    print(f"行{row}:")
    for col in range(1, min(30, ws.max_column) + 1):
        val = ws.cell(row=row, column=col).value
        if val:
            print(f"  列{col}: {val}")

print("=" * 80)
