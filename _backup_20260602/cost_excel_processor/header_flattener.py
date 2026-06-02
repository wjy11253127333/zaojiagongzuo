"""
表头拆解工具 — header_flattener.py

功能：自动检测多级表头区域，拆解成单行表头（下划线连接），
      写入原位并删除多余表头行，保持原表格式不变。

处理流程：
  1. 复制原文件
  2. 逐个工作表自动检测表头起始行和表头行数
  3. 解析多级表头 → 下划线拼接为单行
  4. 全表合并单元格处理（表头区取消/数据区纵向填充/数据区横向不填充）
  5. 清空表头区域内容，写入拆解后的单行表头
  6. 扫描公式修正引用（行号偏移 + 被删区域警告）
  7. 物理删除多余表头空行（从下往上删）
  8. 统一格式美化（微软雅黑10pt, 深蓝表头, 全表细线边框）
  9. 保存新文件

用法：
  python header_flattener.py <输入Excel路径> [输出Excel路径]
"""

import os
import re
import sys
import shutil

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.cell.cell import MergedCell
from excel_reader import ExcelReader
from header_parser import HeaderParser


def _fix_formulas_after_row_delete(ws, deleted_rows: list, max_col: int) -> dict:
    """
    扫描工作表所有含公式的单元格，修正因删除行导致的公式引用偏移。

    参数:
        ws: openpyxl Worksheet
        deleted_rows: 将要删除的行号列表（从下往上排序）
        max_col: 工作表最大列数
    返回:
        {"fixed": 修正数量, "warned": 警告数量, "details": [...]}
    """
    if not deleted_rows:
        return {"fixed": 0, "warned": 0, "details": []}

    # deleted_rows 已从下往上排序（物理删除顺序）
    # 计算每个被删行之前累计已删行数
    # 例如删除 [5, 4]: 最终所有 R>5 的行号-2
    deleted_set = set(deleted_rows)
    max_deleted = max(deleted_rows)
    min_deleted = min(deleted_rows)
    shift = len(deleted_rows)  # 统一偏移量

    fixed_count = 0
    warned_count = 0
    details = []

    # 公式单元格引用正则: 捕获可选的 [工作表名!] + [列字母] + [行号]
    # 支持: A1, $A$1, A$1, $A1, Sheet2!A1, 'Sheet Name'!A1, Sheet2!$A$1
    cell_ref_pattern = re.compile(
        r"(?:(?P<sheet>('[^']+'|[A-Za-z0-9_\-.\u4e00-\u9fff\s]+)!))?"
        r"(?P<col>\$?[A-Z]{1,3})"
        r"(?P<row>\$?\d+)"
    )

    def fix_formula(match):
        nonlocal fixed_count, warned_count
        sheet_part = match.group("sheet") or ""
        col_part = match.group("col")
        row_str = match.group("row")

        # 解析行号
        is_absolute = row_str.startswith("$")
        row_num = int(row_str.lstrip("$"))

        if row_num < min_deleted:
            # 在被删区域之上，不变
            return match.group(0)

        if min_deleted <= row_num <= max_deleted:
            # 在被删区域之内 → 替换为警告文本
            warned_count += 1
            return '"表头区域已删除，用户确认数据是否正确"'

        # 在被删区域之下 → 行号 - shift
        new_row = row_num - shift
        new_row_str = f"${new_row}" if is_absolute else str(new_row)
        result = f"{sheet_part}{col_part}{new_row_str}"
        fixed_count += 1
        return result

    max_row = ws.max_row
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                original = cell.value
                new_formula = cell_ref_pattern.sub(fix_formula, original)
                if new_formula != original:
                    cell.value = new_formula
                    details.append(f"  {cell.coordinate}: {original} → {new_formula}")

    return {"fixed": fixed_count, "warned": warned_count, "details": details}


def _apply_formatting(ws, first_header_row: int, max_col: int):
    """
    统一工作表格式。

    参数:
        ws: openpyxl Worksheet
        first_header_row: 标准化后的表头行号
        max_col: 工作表最大列数
    """
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )

    # 样式定义
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_align = Alignment(horizontal="right", vertical="center")
    title_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_row = ws.max_row

    # 判断列类型：表头含数字关键词 → 数值列
    numeric_keywords = ["工程量", "合价", "单价", "数量", "金额", "造价", "增减"]
    numeric_cols = set()
    for col_idx in range(1, max_col + 1):
        h = ws.cell(row=first_header_row, column=col_idx).value
        if h and any(kw in str(h) for kw in numeric_keywords):
            numeric_cols.add(col_idx)

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # 边框
            cell.border = thin_border

            if row_idx == first_header_row:
                # 表头行
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            elif row_idx < first_header_row:
                # 标题行（Row 1-2）：保持原样，只加边框和居中
                cell.alignment = title_align
                if cell.font and cell.font.name:
                    # 已有字体，不改
                    pass
                else:
                    cell.font = data_font
            else:
                # 数据行
                cell.font = data_font
                if col_idx in numeric_cols:
                    cell.alignment = number_align
                else:
                    cell.alignment = text_align

            # 行高
            if row_idx == first_header_row:
                ws.row_dimensions[row_idx].height = None  # 自适应
            elif row_idx >= first_header_row + 1:
                ws.row_dimensions[row_idx].height = 20

    # 剥离溢出列的旧格式（超出 parsed 范围的列，如 Z/AA）
    # 这些列已被清空值为 None，但可能残留原始文件的边框/填充格式
    extra_max_col = ws.max_column
    if extra_max_col > max_col:
        empty_fill = PatternFill(fill_type=None)
        empty_border = Border()
        for row_idx in range(1, max_row + 1):
            for col_idx in range(max_col + 1, extra_max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = empty_border
                cell.fill = empty_fill
def flatten_headers(input_path: str, output_path: str = None) -> str:
    """
    对输入文件的所有工作表进行表头拆解，输出新文件

    参数:
        input_path: 输入Excel文件路径
        output_path: 输出路径（默认：原文件名_拆解.xlsx）
    返回:
        输出文件路径
    """
    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = base + "_拆解.xlsx"

    # Step 0: 复制原文件
    print(f"[1/5] 复制原文件...")
    shutil.copy2(input_path, output_path)
    print(f"       {os.path.basename(input_path)} → {os.path.basename(output_path)}")

    # Step 1: 加载并分析
    print(f"[2/5] 加载并自动检测表头区域...")
    reader = ExcelReader(input_path)
    reader.load()

    # 用新的 workbook 进行操作
    wb = openpyxl.load_workbook(output_path)
    parser = HeaderParser()

    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info = reader._analyze_sheet(reader.wb[sheet_name])

        if info["is_empty"]:
            print(f"       [{sheet_name}] → 空表，跳过")
            results.append({"sheet": sheet_name, "status": "skipped (empty)"})
            continue

        first_header_row = info.get("first_header_row", 1)
        header_rows = info.get("header_rows", 1)

        if header_rows <= 1:
            print(f"       [{sheet_name}] → 单行表头，无需拆解")
            results.append({"sheet": sheet_name, "status": "single-row, skipped"})
            continue

        print(f"       [{sheet_name}] 表头行: {first_header_row}-{first_header_row + header_rows - 1} ({header_rows}行)")

        # Step 2: 解析表头
        parsed = parser.parse_sheet_headers(
            reader.wb[sheet_name], header_rows, first_header_row
        )

        # 清理尾部空列
        while parsed and not parsed[-1]:
            parsed.pop()

        # Step 3: 全表合并单元格处理
        #   表头区：直接取消合并（Step 4会清空内容）
        #   数据区纵向合并（同列多行）：取消合并 → 顶部值向下填充
        #   数据区横向/矩形合并：直接取消合并（不填充）
        data_start_row = first_header_row + header_rows
        unmerge_count = 0
        fill_down_count = 0

        # 需要复制一份合并范围列表，因为遍历过程中会修改
        merge_list = list(ws.merged_cells.ranges)

        for mr in merge_list:
            mr_str = str(mr)
            # 判断是否在表头区域（与表头行有重叠）
            in_header = (mr.min_row <= first_header_row + header_rows - 1 and
                         mr.max_row >= first_header_row)

            if in_header:
                # 表头区：直接取消
                ws.unmerge_cells(mr_str)
                unmerge_count += 1
            elif mr.min_row >= data_start_row:
                # 数据区
                if mr.min_col == mr.max_col and mr.min_row < mr.max_row:
                    # 纵向合并（同列多行）：取top-left值 → 取消 → 向下填充
                    val = ws.cell(row=mr.min_row, column=mr.min_col).value
                    ws.unmerge_cells(mr_str)
                    for r in range(mr.min_row + 1, mr.max_row + 1):
                        ws.cell(row=r, column=mr.min_col).value = val
                    fill_down_count += 1
                    unmerge_count += 1
                else:
                    # 横向/矩形合并：直接取消，不填充
                    ws.unmerge_cells(mr_str)
                    unmerge_count += 1

        # Step 4: 清空表头区域内容
        # 以实际拆解列数为准写入区域（max_col）；多余列（如 Z/AA 的残留值）
        # 也需要清空为 None，避免旧数据残留在可见位置
        max_col = len(parsed)
        clear_max_col = min(max(ws.max_column, max_col), max_col + 100)
        for row_idx in range(first_header_row, first_header_row + header_rows):
            for col_idx in range(1, clear_max_col + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = None
                except (ValueError, AttributeError):
                    pass

        # Step 5: 写入拆解后的单行表头到表头区域第一行
        target_row = first_header_row
        for col_idx, header_text in enumerate(parsed, 1):
            cell = ws.cell(row=target_row, column=col_idx)
            cell.value = header_text if header_text else None

        # Step 6: 扫描所有公式，修正因删除行导致的引用偏移
        # 收集将要删除的空行（除 target_row 外的所有表头行）
        rows_to_delete = []
        for r in range(first_header_row, first_header_row + header_rows):
            if r != target_row:
                rows_to_delete.append(r)
        rows_to_delete.sort(reverse=True)  # 从下往上

        formula_result = _fix_formulas_after_row_delete(ws, rows_to_delete, max_col)
        print(f"           公式修正: {formula_result['fixed']} 处引用调整, {formula_result['warned']} 处被删区域警告")

        # Step 7: 物理删除多余表头空行（从下往上删）
        for row_idx in rows_to_delete:
            ws.delete_rows(row_idx)
        print(f"           物理删除 {len(rows_to_delete)} 个空行")

        # Step 8: 统一格式美化
        _apply_formatting(ws, first_header_row, max_col)
        print(f"           格式统一: 微软雅黑10pt, 深蓝表头, 全表细线边框")

        results.append({
            "sheet": sheet_name,
            "status": "flattened",
            "from_rows": f"{first_header_row}-{first_header_row + header_rows - 1}",
            "to_row": first_header_row,
            "columns": len(parsed),
            "unmerged": unmerge_count,
            "fill_down": fill_down_count,
            "formula_fixed": formula_result["fixed"],
            "formula_warned": formula_result["warned"],
            "rows_deleted": len(rows_to_delete),
        })

        preview = [h[:25] if h else "(空)" for h in parsed[:6]]
        print(f"           拆解 {len(parsed)} 列, 取消 {unmerge_count} 个合并, 纵向填充 {fill_down_count} 个")
        print(f"           首列: {preview}")

    # Step 7: 保存
    print(f"[3/5] 保存文件...")
    wb.save(output_path)

    # 打印汇总
    print(f"[4/5] 处理完成！")
    print(f"{'='*60}")
    for r in results:
        print(f"  {r['sheet']}: {r['status']}")
        if r['status'] == 'flattened':
            fd = r.get('fill_down', 0)
            ff = r.get('formula_fixed', 0)
            fw = r.get('formula_warned', 0)
            rd = r.get('rows_deleted', 0)
            print(f"    原表头 {r['from_rows']} → 新表头第 {r['to_row']} 行, {r['columns']} 列")
            print(f"    合并: 取消 {r['unmerged']} 个, 纵向填充 {fd} 列 | 公式: 修正 {ff} 处, 警告 {fw} 处 | 删除 {rd} 空行")
    print(f"{'='*60}")
    print(f"[5/5] 输出: {output_path}")

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python header_flattener.py <输入Excel> [输出Excel]")
    else:
        input_path = sys.argv[1]
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
        flatten_headers(input_path, output_path)
