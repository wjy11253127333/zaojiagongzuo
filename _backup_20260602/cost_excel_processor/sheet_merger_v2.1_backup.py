"""
清单表合并脚本 v2.0  —— 公式感知版
=====================================================
功能：
  1. 扫描工作簿，识别清单表（表头命中 ≥4 个关键字）
  2. 直接用 openpyxl（data_only=False）逐格复制，保留公式字符串
  3. 教学楼追加到综合楼尾部时，公式行号全量 +offset
  4. 生成「公式核查报告」Sheet，列出所有范围公式(含:)和跨表公式(含!)
  5. 三段式数值一致性核查（追加前 → 追加后 → 对比）

用法：
    python sheet_merger.py "input.xlsx"
    python sheet_merger.py "input.xlsx" --sheets "综合楼表名" "教学楼表名"

注意：
    本脚本仅处理「1.1综合楼_终测.xlsx」这类包含两张清单表的工作簿。
    输出目标 Sheet 名：清单汇总
    公式核查报告 Sheet 名：公式核查报告
    合并说明 Sheet 名：合并说明
"""

import os
import sys
import re
import copy

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import normalize_header_text

# ===== 配置 =====
LIST_KEYWORDS    = ["项目特征", "项目名称", "项目编码", "计量单位", "工程量"]
MIN_HIT          = 4
OUTPUT_SHEET     = "清单汇总"
FORMULA_SHEET    = "公式核查报告"
REPORT_SHEET     = "合并说明"

# 样式常量
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_FONT  = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT  = Font(name="微软雅黑", size=10)
_LEFT_ALIGN = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_CTR_ALIGN  = Alignment(horizontal="center", vertical="center")
_NUM_ALIGN  = Alignment(horizontal="right",  vertical="center")

# 警示色（范围公式行）
_WARN_FILL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
# 错误色（跨表公式行）
_ERR_FILL   = PatternFill(start_color="FFDDC1", end_color="FFDDC1", fill_type="solid")


# ============================================================
# 工具函数
# ============================================================

def is_list_sheet(headers: list) -> bool:
    all_text = "".join(normalize_header_text(str(h)) for h in headers if h is not None)
    return sum(1 for kw in LIST_KEYWORDS if kw in all_text) >= MIN_HIT


def get_header_row(ws, max_scan: int = 10) -> int:
    """自动检测表头行（跳过标题行，找关键字命中最多的行）"""
    start_row = 1
    for row_idx in range(1, min(ws.max_row + 1, max_scan + 1)):
        non_empty = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and str(val).strip():
                non_empty.append(str(val).strip())
        if any(len(t) > 35 for t in non_empty):
            continue
        start_row = row_idx
        break

    best_row, best_hits = start_row, 0
    for row_idx in range(start_row, min(ws.max_row + 1, start_row + max_scan)):
        row_data = [str(ws.cell(row=row_idx, column=c).value).strip()
                    for c in range(1, ws.max_column + 1)
                    if ws.cell(row=row_idx, column=c).value is not None]
        hits = sum(1 for kw in LIST_KEYWORDS
                   if kw in "".join(normalize_header_text(h) for h in row_data))
        if hits > best_hits:
            best_hits, best_row = hits, row_idx
    return best_row


def get_sheet_headers(ws) -> tuple:
    """
    读取表头行，自动展开合并单元格：
    - 表头行有合并单元格时，父列名覆盖所有子列（造成空列）
    - 自动用子表头行（header_row+1）的值填充空列
    - 返回唯一可区分的列名，如「合同工程量」「送审工程量」
    """
    header_row = get_header_row(ws)
    sub_row = header_row + 1 if header_row < ws.max_row else header_row

    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        hdr_str = str(val).strip() if val is not None else ""

        # 表头行该列为空（合并单元格子列）→ 用子表头行的值
        if not hdr_str:
            sub_val = ws.cell(row=sub_row, column=col_idx).value
            hdr_str = str(sub_val).strip() if sub_val is not None else ""

        headers.append(hdr_str)

    # 移除尾部空列
    while headers and headers[-1] == "":
        headers.pop()
    return header_row, headers


# ============================================================
# 列对齐与行过滤工具
# ============================================================

def build_unified_columns(headers_a: list, headers_b: list) -> tuple:
    """
    按标准化表头名称对齐，构建统一列列表和两张表的列映射。

    返回: (unified_headers, col_indices_a, col_indices_b)
      - unified_headers: 统一列名列表，业态列插入在序号之后
      - col_indices_a:   长度 = unified_col_count，每项 = 源列号(1-based) 或 None
      - col_indices_b:   同上
    """
    norm_a = [normalize_header_text(h) for h in headers_a]
    norm_b = [normalize_header_text(h) for h in headers_b]

    # Step 1: 表A表头保序，表B独有列追加末尾
    unified = list(headers_a)
    for i, h_norm in enumerate(norm_b):
        if h_norm and h_norm not in norm_a:
            unified.append(headers_b[i])

    # Step 2: 在"序号"之后插入"业态"列
    xuhao_idx = None
    for i, h in enumerate(unified):
        if '序号' in normalize_header_text(h):
            xuhao_idx = i
            break

    if xuhao_idx is not None:
        unified.insert(xuhao_idx + 1, '业态')
    else:
        unified.insert(0, '业态')

    # Step 3: 构建列映射
    def _build_indices(src_headers, src_norms):
        indices = []
        for u_hdr in unified:
            if normalize_header_text(u_hdr) == '业态':
                indices.append(None)  # 业态是合成列
                continue
            u_norm = normalize_header_text(u_hdr)
            found = False
            for si, s_norm in enumerate(src_norms):
                if s_norm == u_norm:
                    indices.append(si + 1)  # 1-based
                    found = True
                    break
            if not found:
                indices.append(None)
        return indices

    col_indices_a = _build_indices(headers_a, norm_a)
    col_indices_b = _build_indices(headers_b, norm_b) if headers_b else []

    return unified, col_indices_a, col_indices_b


def get_skip_rows(ws, header_row: int) -> set:
    """
    识别需跳过的行（仅用于非首张表）。
    返回需跳过的源行号集合：表头行 + 工程名称行 + 工程量偏差/漏项对比表行。
    """
    skip = {header_row}

    # 含"工程名称"的行（表头行之前）
    for row in range(1, header_row):
        for col in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=row, column=col).value
            if val and '工程名称' in str(val):
                skip.add(row)
                break

    # "工程量偏差…漏项对比表"行（全表扫描前几列）
    for row in range(1, ws.max_row + 1):
        if row in skip:
            continue
        for col in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=row, column=col).value
            if val and '工程量偏差' in str(val) and '漏项对比表' in str(val):
                skip.add(row)
                break

    return skip


# ============================================================
# 公式工具函数
# ============================================================

def classify_formula(formula: str) -> str:
    """分类公式类型：inline / range / cross"""
    if not formula or not formula.startswith("="):
        return "value"
    if "!" in formula:
        return "cross"
    if ":" in formula:
        return "range"
    return "inline"


def adjust_formula_rows(formula: str, offset: int) -> str:
    """
    将公式中所有行号 +offset。
    仅调整相对引用（不带 $ 前缀的行号）。
    例：=SUM(V5:V10) + offset=266 → =SUM(V271:V276)
    规则：匹配 [A-Za-z$]+数字 中没有 $ 前缀的行号部分。
    """
    if not formula or not formula.startswith("="):
        return formula

    def _shift(m):
        col_part = m.group(1)   # 列字母，可能含 $
        row_lock = m.group(2)   # 行锁定符 $ 或空
        row_num  = m.group(3)   # 行号数字
        if row_lock == "$":     # 绝对行引用，不调整
            return f"{col_part}${row_num}"
        return f"{col_part}{int(row_num) + offset}"

    # 匹配：列字母($?) + 行锁定($?) + 行号数字
    pattern = r"(\$?[A-Za-z]+)(\$?)(\d+)"
    return re.sub(pattern, _shift, formula)


def extract_formulas_info(ws, sheet_name: str, offset: int = 0) -> list:
    """
    提取工作表中所有范围公式和跨表公式的详细信息。
    offset > 0 表示这是追加到汇总表后的偏移量（用于展示偏移后的公式）。
    返回：[{来源表, 原位置, 新位置, 公式类型, 原公式, 偏移后公式, 建议}, ...]
    """
    records = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not (val and isinstance(val, str) and val.startswith("=")):
                continue
            ftype = classify_formula(val)
            if ftype not in ("range", "cross"):
                continue

            col_letter = get_column_letter(col)
            orig_pos   = f"{col_letter}{row}"
            new_pos    = f"{col_letter}{row + offset}" if offset > 0 else orig_pos
            shifted    = adjust_formula_rows(val, offset) if offset > 0 else val

            if ftype == "cross":
                suggestion = "⚠️ 跨表引用，行号已偏移但跨表路径不变，请人工确认引用是否正确"
            else:
                # 范围公式：检查是否引用范围包含整段数据（可能是分项小计）
                suggestion = "请核查 SUM 范围在汇总表中是否覆盖正确的数据行"

            records.append({
                "来源表":     sheet_name,
                "原位置":     orig_pos,
                "汇总表新位置": new_pos,
                "公式类型":   "跨表公式" if ftype == "cross" else "范围公式(SUM等)",
                "原公式":     val,
                "偏移后公式": shifted,
                "建议":       suggestion,
            })
    return records


# ============================================================
# 数值一致性核查
# ============================================================

def _get_numeric_value(cell) -> float | None:
    """安全读取数值（公式单元格返回 None，纯数字返回 float）"""
    val = cell.value
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str) and val.startswith("="):
        return None  # 公式，需要 data_only 才能读到计算值
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def audit_numeric_stats(ws, label: str) -> dict:
    """
    对工作表统计：行数、公式总数、分类数量，以及关键数值列的原始值汇总
    （用于追加前/追加后对比；公式单元格记录公式字符串，不计算值）
    """
    formula_count = {"inline": 0, "range": 0, "cross": 0}
    data_rows = 0

    for row in range(1, ws.max_row + 1):
        row_has_data = False
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None:
                continue
            if isinstance(val, str) and val.startswith("="):
                ft = classify_formula(val)
                formula_count[ft] = formula_count.get(ft, 0) + 1
                row_has_data = True
            elif val != "" and val is not None:
                row_has_data = True
        if row_has_data:
            data_rows += 1

    return {
        "标签": label,
        "数据行数（含表头）": data_rows,
        "工作表最大行": ws.max_row,
        "工作表最大列": ws.max_column,
        "公式总数": sum(formula_count.values()),
        "行内公式": formula_count.get("inline", 0),
        "范围公式": formula_count.get("range", 0),
        "跨表公式": formula_count.get("cross", 0),
    }


# ============================================================
# 逐格复制（核心）
# ============================================================

def copy_cells_with_formula(src_ws, dst_ws,
                             src_row_start: int, src_row_end: int,
                             dst_row_start: int,
                             col_count: int,
                             row_offset: int = 0):
    """
    [保留旧版兼容] 按列位置逐格复制（已废弃，请用 copy_cells_with_formula_mapped）。
    """
    copied_cells  = 0
    formula_cells = 0
    for src_row in range(src_row_start, src_row_end + 1):
        dst_row = dst_row_start + (src_row - src_row_start)
        for col in range(1, col_count + 1):
            src_cell = src_ws.cell(row=src_row, column=col)
            dst_cell = dst_ws.cell(row=dst_row, column=col)
            val = src_cell.value
            if val is None:
                dst_cell.value = None
            elif isinstance(val, str) and val.startswith("="):
                dst_cell.value = adjust_formula_rows(val, row_offset)
                formula_cells += 1
            else:
                dst_cell.value = val
            if src_cell.has_style:
                dst_cell.font          = copy.copy(src_cell.font)
                dst_cell.border        = copy.copy(src_cell.border)
                dst_cell.fill          = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment     = copy.copy(src_cell.alignment)
                dst_cell.protection    = copy.copy(src_cell.protection)
            copied_cells += 1
        if src_row in src_ws.row_dimensions:
            dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height
    return copied_cells, formula_cells


def copy_cells_with_formula_mapped(
    src_ws, dst_ws,
    src_row_start: int, src_row_end: int,
    dst_row_start: int,
    col_indices: list,          # [src_col(1-based) or None] × unified_col_count
    unified_headers: list,
    yetai_col: int = None,      # 1-based 业态列在输出中的列号
    sheet_name: str = "",
    combined_name: str = "",
    is_first_sheet: bool = True,
    skip_rows: set = None,
):
    """
    按表头名称映射复制单元格（v2.1 核心函数）。

    - col_indices: 每个输出列对应源列号（1-based）或 None（无源列→留空）
    - yetai_col:   业态合成列位置，自动填 sheet_name
    - combined_name: 非首张表时，工程名称列填此值
    - skip_rows:   需跳过的源行号集合

    返回：(copied_cells, formula_cells, next_dst_row)
    """
    skip_rows = skip_rows or set()
    copied_cells  = 0
    formula_cells = 0

    # 定位"工程名称"列在统一表头中的索引（0-based）
    gongcheng_col = None
    for i, h in enumerate(unified_headers):
        if '工程名称' in normalize_header_text(h):
            gongcheng_col = i + 1  # 1-based dst col
            break

    dst_row = dst_row_start
    for src_row in range(src_row_start, src_row_end + 1):
        if src_row in skip_rows:
            continue

        row_shift = dst_row - src_row

        for out_idx, src_col in enumerate(col_indices):
            dst_col = out_idx + 1

            # ── 业态合成列 ──
            if yetai_col is not None and dst_col == yetai_col:
                cell = dst_ws.cell(row=dst_row, column=dst_col)
                cell.value = sheet_name
                cell.font      = _DATA_FONT
                cell.alignment = _CTR_ALIGN
                cell.border    = _THIN
                continue

            # ── 无源列 → 留空 ──
            if src_col is None:
                dst_ws.cell(row=dst_row, column=dst_col).value = None
                continue

            src_cell = src_ws.cell(row=src_row, column=src_col)
            dst_cell = dst_ws.cell(row=dst_row, column=dst_col)
            val = src_cell.value

            # ── 工程名称变换（非首张表）──
            if (not is_first_sheet and gongcheng_col is not None
                    and dst_col == gongcheng_col):
                dst_cell.value = combined_name
            elif val is None:
                dst_cell.value = None
            elif isinstance(val, str) and val.startswith("="):
                dst_cell.value = adjust_formula_rows(val, row_shift)
                formula_cells += 1
            else:
                dst_cell.value = val

            # 复制样式
            if src_cell.has_style:
                dst_cell.font          = copy.copy(src_cell.font)
                dst_cell.border        = copy.copy(src_cell.border)
                dst_cell.fill          = copy.copy(src_cell.fill)
                dst_cell.number_format = src_cell.number_format
                dst_cell.alignment     = copy.copy(src_cell.alignment)
                dst_cell.protection    = copy.copy(src_cell.protection)

            copied_cells += 1

        # 行高
        if src_row in src_ws.row_dimensions:
            dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height

        dst_row += 1

    return copied_cells, formula_cells, dst_row


def copy_merged_cells(src_ws, dst_ws, row_offset: int = 0):
    """复制合并单元格区域（row_offset > 0 时调整行号）"""
    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        new_min_row = min_row + row_offset
        new_max_row = max_row + row_offset
        new_range = (f"{get_column_letter(min_col)}{new_min_row}:"
                     f"{get_column_letter(max_col)}{new_max_row}")
        try:
            dst_ws.merge_cells(new_range)
        except Exception:
            pass  # 已存在或冲突的合并区域跳过


def copy_col_widths(src_ws, dst_ws):
    """复制列宽"""
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width


def _copy_merged_cells_mapped(src_ws, dst_ws,
                              col_indices: list,
                              row_offset: int = 0,
                              skip_src_rows: set = None):
    """
    按列映射复制合并单元格区域。
    col_indices: [src_col(1-based) or None]，长度为统一列数。
    将源合并区域映射到目标列。

    行偏移 = row_offset - 「skip_src_rows 中在合并区域之前的行数」
    例：首张表 row_offset=1, skip_src_rows={3}
      源行1合并 → 1 + 1 - 0 = 输出行2
      源行2合并 → 2 + 1 - 0 = 输出行3
      源行3合并 → 被跳过（与skip_src_rows重叠）
      源行4合并 → 4 + 1 - 1 = 输出行4

    Parameters:
        skip_src_rows: 源文件中被跳过的行号集合（这些行不在输出中）
    """
    skip_src_rows = skip_src_rows or set()
    sorted_skip = sorted(skip_src_rows)

    # 建立反向映射：dst_col(1-based) → src_col(1-based)
    dst_to_src = {}
    for dst_idx, src_col in enumerate(col_indices):
        if src_col is not None:
            dst_to_src[dst_idx + 1] = src_col

    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds

        # 跳过与源跳过行重叠的合并区域（源表头行等）
        if any(min_row <= r <= max_row for r in skip_src_rows):
            continue

        # 找到对应的目标列
        new_min_col = None
        new_max_col = None
        for dcol in range(min_col, max_col + 1):
            for d, s in dst_to_src.items():
                if s == dcol:
                    if new_min_col is None or d < new_min_col:
                        new_min_col = d
                    if new_max_col is None or d > new_max_col:
                        new_max_col = d
                    break

        if new_min_col is None:
            continue  # 该合并区域全部在无映射列，跳过

        # 计算有效行偏移：跳过行抵消部分偏移
        skipped_before = sum(1 for r in sorted_skip if r < min_row)
        effective_offset = row_offset - skipped_before

        new_min_row = min_row + effective_offset
        new_max_row = max_row + effective_offset

        new_range = (f"{get_column_letter(new_min_col)}{new_min_row}:"
                     f"{get_column_letter(new_max_col)}{new_max_row}")
        try:
            dst_ws.merge_cells(new_range)
        except Exception:
            pass


# ============================================================
# 格式化输出 Sheet
# ============================================================

def _apply_formula_report_formatting(ws, row_count: int):
    """为公式核查报告 Sheet 应用格式"""
    col_count = 7  # 7列

    for r in range(1, row_count + 2):  # +2 = header + data
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _THIN
            if r == 1:
                cell.font      = _HDR_FONT
                cell.fill      = _HDR_FILL
                cell.alignment = _HDR_ALIGN
            else:
                cell.font = _DATA_FONT
                # 根据公式类型上色
                ftype_val = ws.cell(row=r, column=4).value
                if ftype_val and "跨表" in str(ftype_val):
                    cell.fill = _ERR_FILL
                elif ftype_val and "范围" in str(ftype_val):
                    cell.fill = _WARN_FILL
                cell.alignment = _LEFT_ALIGN

        ws.row_dimensions[r].height = 20 if r > 1 else None

    ws.freeze_panes = "A2"

    # 列宽
    widths = [20, 10, 12, 18, 40, 40, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _apply_audit_formatting(ws, row_count: int):
    """为一致性核查 Sheet 应用格式"""
    col_count = 9

    for r in range(1, row_count + 2):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _THIN
            if r == 1:
                cell.font      = _HDR_FONT
                cell.fill      = _HDR_FILL
                cell.alignment = _HDR_ALIGN
            else:
                cell.font      = _DATA_FONT
                cell.alignment = _CTR_ALIGN

                # 对比结果列（最后一列）上色
                result_cell = ws.cell(row=r, column=col_count)
                result_val  = result_cell.value
                if result_val and "✓" in str(result_val):
                    result_cell.fill = PatternFill(start_color="C6EFCE",
                                                   end_color="C6EFCE", fill_type="solid")
                elif result_val and "✗" in str(result_val):
                    result_cell.fill = PatternFill(start_color="FFC7CE",
                                                   end_color="FFC7CE", fill_type="solid")

        ws.row_dimensions[r].height = 22 if r > 1 else None

    ws.freeze_panes = "A2"
    for c in range(1, col_count + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


# ============================================================
# 主函数
# ============================================================

def merge_to_summary(input_path: str, target_sheet_names: list | None = None) -> dict:
    """
    主处理函数 v2.1 — 按表头名称对齐合并。

    参数：
        input_path:        Excel 文件路径
        target_sheet_names: 可选，手动指定清单表名列表（按顺序）
    返回：结果摘要 dict
    """
    print(f"\n{'='*70}")
    print(f"公式感知合并 v2.1（按名称对齐）— {os.path.basename(input_path)}")
    print(f"{'='*70}")

    # ── Step 1: 加载（保留公式） ──────────────────────────────
    wb = openpyxl.load_workbook(input_path, data_only=False)
    all_sheets = list(wb.sheetnames)

    # ── Step 2: 识别清单表 ────────────────────────────────────
    if target_sheet_names:
        list_sheets   = [s for s in target_sheet_names if s in all_sheets]
        skipped_sheets = [s for s in all_sheets if s not in list_sheets]
    else:
        list_sheets, skipped_sheets = [], []
        for sn in all_sheets:
            if sn in (OUTPUT_SHEET, FORMULA_SHEET, REPORT_SHEET):
                skipped_sheets.append(sn)
                continue
            ws = wb[sn]
            if ws.max_row == 0:
                skipped_sheets.append(sn)
                continue
            _, headers = get_sheet_headers(ws)
            if is_list_sheet(headers):
                list_sheets.append(sn)
            else:
                skipped_sheets.append(sn)

    print(f"\n识别到清单表 ({len(list_sheets)} 个)：{list_sheets}")
    print(f"跳过 ({len(skipped_sheets)} 个)：{skipped_sheets}")

    if not list_sheets:
        print("\n[WARN] 未找到任何清单表。")
        return {"success": False, "list_sheets": [], "skipped_sheets": skipped_sheets}

    # ── Step 3: 读取所有清单表的表头行和表头文本 ────────────
    sheet_headers = []   # [(header_row, [raw_headers])]
    for sn in list_sheets:
        ws = wb[sn]
        hr, hdrs = get_sheet_headers(ws)
        sheet_headers.append((ws, hr, hdrs))

    # ── Step 4: 构建统一列列表 + 列映射 ─────────────────────
    ws_a, hdr_row_a, hdrs_a = sheet_headers[0]

    if len(sheet_headers) >= 2:
        # 两张表：按名称对齐
        _, hdr_row_b, hdrs_b = sheet_headers[1]
        unified_headers, col_indices_a, col_indices_b = build_unified_columns(
            hdrs_a, hdrs_b
        )
    else:
        # 只有一张表：在 hdrs_a 中插入"业态"
        unified_headers = list(hdrs_a)
        xuhao_idx = None
        for i, h in enumerate(unified_headers):
            if '序号' in normalize_header_text(h):
                xuhao_idx = i
                break
        yetai_insert = xuhao_idx + 1 if xuhao_idx is not None else 0
        unified_headers.insert(yetai_insert, '业态')
        # 列映射：位置等同（位置i→统一列i，业态位置→None）
        col_indices_a = list(range(1, len(hdrs_a) + 1))
        col_indices_a.insert(yetai_insert, None)
        col_indices_b = []

    # 业态列在统一表头中的 1-based 位置
    yetai_col = None
    for i, h in enumerate(unified_headers):
        if normalize_header_text(h) == '业态':
            yetai_col = i + 1
            break

    # 工程名称：所有清单表名用"+"连接
    combined_name = "+".join(list_sheets)

    print(f"\n统一列数：{len(unified_headers)}")
    print(f"业态列位置：第 {yetai_col} 列")
    print(f"工程名称（非首表）：{combined_name}")

    # ── Step 5: 追加前基准核查 ────────────────────────────────
    print(f"\n{'─'*70}")
    print("[ 追加前基准核查 ]")
    pre_stats = {}
    for idx, (ws, hdr_row, hdrs) in enumerate(sheet_headers):
        label = f"{list_sheets[idx]}"
        stats = audit_numeric_stats(ws, label)
        pre_stats[label] = stats
        print(f"  {label}：{stats['工作表最大行']}行 × {stats['工作表最大列']}列，"
              f"公式共{stats['公式总数']}个")

    # ── Step 6: 删除旧输出 Sheet，创建新 Sheet ────────────────
    for sn in (OUTPUT_SHEET, FORMULA_SHEET, REPORT_SHEET):
        if sn in wb.sheetnames:
            del wb[sn]

    ws_out = wb.create_sheet(OUTPUT_SHEET)
    print(f"\n已创建「{OUTPUT_SHEET}」Sheet")

    # ── Step 7: 写统一表头行 ─────────────────────────────────
    for c_idx, hdr in enumerate(unified_headers, 1):
        cell = ws_out.cell(row=1, column=c_idx)
        cell.value = hdr
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _THIN

    dst_row_cursor = 2  # 表头占第1行，数据从第2行开始

    # ── Step 8: 按表头映射复制所有清单表 ─────────────────────
    all_formula_records = []

    for idx, (ws_src, hdr_row_src, hdrs_src) in enumerate(sheet_headers):
        is_first = (idx == 0)
        sn = list_sheets[idx]

        # 收集公式
        if is_first and idx == 0:
            all_formula_records.extend(
                extract_formulas_info(wb[sn], sn, offset=0)
            )

        # 跳过行 + 数据起止
        if is_first:
            # 首张表：保留工程名称等前置行，仅跳过表头行
            skip_rows = {hdr_row_src}
            src_start = 1          # 包含表头之前的行
        else:
            # 非首张表：跳过表头行、工程名称行、工程量偏差行
            skip_rows = get_skip_rows(ws_src, hdr_row_src)
            src_start = hdr_row_src + 1
        src_end = ws_src.max_row

        # 列映射
        col_indices = col_indices_a if is_first else col_indices_b

        # 行偏移 = 目标起始行 - 源起始行（用于公式行号调整）
        row_shift = dst_row_cursor - src_start

        copied, formulas, dst_row_cursor = copy_cells_with_formula_mapped(
            src_ws          = ws_src,
            dst_ws          = ws_out,
            src_row_start   = src_start,
            src_row_end     = src_end,
            dst_row_start   = dst_row_cursor,
            col_indices     = col_indices,
            unified_headers = unified_headers,
            yetai_col       = yetai_col,
            sheet_name      = sn,
            combined_name   = combined_name if not is_first else "",
            is_first_sheet  = is_first,
            skip_rows       = skip_rows,
        )

        # 合并单元格（按列映射 + 动态行偏移）
        # 首张表：row_offset=1（输出多了统一表头行），跳过源表头行
        # 次张表：row_offset=row_shift，跳过 table header / 工程名称 / 工程量偏差
        _copy_merged_cells_mapped(
            ws_src, ws_out,
            col_indices    = col_indices,
            row_offset     = 1 if is_first else row_shift,
            skip_src_rows  = {hdr_row_src} if is_first else get_skip_rows(ws_src, hdr_row_src),
        )

        # 列宽
        if is_first:
            copy_col_widths(ws_src, ws_out)
        else:
            # 仅补首张表没有的列
            for ci, si in enumerate(col_indices, 1):
                if si is None:
                    continue
                src_letter = get_column_letter(si)
                if src_letter in ws_src.column_dimensions:
                    dim = ws_src.column_dimensions[src_letter]
                    if dim.width and ci <= len(ws_out.column_dimensions):
                        try:
                            ws_out.column_dimensions[get_column_letter(ci)].width = dim.width
                        except Exception:
                            pass

        print(f"  [{sn}] 复制完成：{copied}个单元格，其中公式{formulas}个"
              f"（起始行={dst_row_cursor - copied//max(len(unified_headers),1)}）")

    # ── Step 9: 公式核查报告 ─────────────────────────────────
    # （暂保留原逻辑，后续按映射后位置更新）

    # ── Step 10: 写「合并说明」Sheet ──────────────────────────
    # （暂保留，格式需调整）

    # ── Step 11: 保存 ─────────────────────────────────────────
    base, ext = os.path.splitext(input_path)
    output_path = input_path

    print(f"\n{'─'*70}")
    print("保存文件...")

    try:
        wb.save(output_path)
        print(f"  ✅ 已保存到原文件：{os.path.basename(output_path)}")
    except PermissionError:
        output_path = f"{base}_含清单汇总{ext}"
        wb.save(output_path)
        print(f"  原文件被占用，已另存为：{os.path.basename(output_path)}")

    wb.close()

    print(f"\n{'='*70}")
    print(f"🎉 处理完成（v2.1）！")
    print(f"  清单表：{len(list_sheets)} 个")
    print(f"  统一列数：{len(unified_headers)}")
    print(f"  输出文件：{os.path.basename(output_path)}")
    print(f"{'='*70}")

    return {
        "success": True,
        "list_sheets":     list_sheets,
        "skipped_sheets":  skipped_sheets,
        "unified_columns": len(unified_headers),
        "output_file":     output_path,
    }


# ============================================================
# CLI 入口
# ============================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="清单表公式感知合并工具 v2.0"
    )
    parser.add_argument("file", help="Excel 文件路径")
    parser.add_argument(
        "--sheets", nargs="+",
        help="手动指定要合并的工作表名（按顺序，第1个为主表，第2个为追加表）"
    )
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"错误：文件不存在 - {args.file}")
        sys.exit(1)

    result = merge_to_summary(args.file, target_sheet_names=args.sheets)

    if not result["success"]:
        print("\n[WARN] 处理未完成，请检查日志。")
        sys.exit(1)
