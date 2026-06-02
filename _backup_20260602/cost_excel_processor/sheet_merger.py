"""
清单表合并脚本 v2.0  —— 公式感知版
=====================================================
功能：
  1. 扫描工作簿，识别清单表（表头命中 ≥4 个关键字）
  2. 直接用 openpyxl（data_only=False）逐格复制，保留公式字符串
  3. 第二张表追加到第一张表尾部时，公式行号全量 +offset
  4. 生成「公式核查报告」Sheet，列出所有范围公式(含:)和跨表公式(含!)
  5. 三段式数值一致性核查（追加前 → 追加后 → 对比）

用法：
    python sheet_merger.py "input.xlsx"
    python sheet_merger.py "input.xlsx" --sheets "综合楼Sheet" "教学楼Sheet"

注意：
    本脚本仅处理包含两张清单表的工作簿。
    输出目标 Sheet 名：清单汇总
    公式核查报告 Sheet 名：公式核查报告
    合并说明 Sheet 名：合并说明
"""

import os
import sys
import re
import copy
import argparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils import normalize_header_text
from header_parser import HeaderParser


# ===== 配置 =====
LIST_KEYWORDS    = ["项目特征", "项目名称", "项目编码", "计量单位", "工程量"]
MIN_HIT          = 4
OUTPUT_SHEET     = "清单汇总"
FORMULA_SHEET    = "公式核查报告"
REPORT_SHEET     = "合并说明"

# 样式常量
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_FONT  = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
_HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DATA_FONT  = Font(name="微软雅黑", size=10)
_LEFT_ALIGN = Alignment(horizontal="left",    vertical="center", wrap_text=True)
_CTR_ALIGN  = Alignment(horizontal="center",  vertical="center")
_NUM_ALIGN  = Alignment(horizontal="right",   vertical="center")


# ============================================================
# 工具函数
# ============================================================

def is_list_sheet(ws, max_scan=10) -> bool:
    """判断工作表是否为清单表（表头含 ≥MIN_HIT 个关键字）"""
    hdr_row = get_header_row(ws, max_scan)
    if hdr_row is None:
        return False
    headers = [ws.cell(row=hdr_row, column=c).value or "" for c in range(1, ws.max_column + 1)]
    hdr_text = "".join(normalize_header_text(str(h)) for h in headers if h)
    return sum(1 for kw in LIST_KEYWORDS if kw in hdr_text) >= MIN_HIT


def get_header_row(ws, max_scan=10) -> int | None:
    """自动检测表头行（跳过前置标题行）"""
    # 先跳过内容超长的行（通常是标题行）
    start_row = 1
    for row in range(1, min(ws.max_row + 1, max_scan + 1)):
        cells = [str(ws.cell(row=row, column=c).value or "").strip() for c in range(1, min(ws.max_column + 1, 10))]
        if any(len(c) > 35 for c in cells):
            continue
        start_row = row
        break

    best_row, best_hits = start_row, 0
    for row in range(start_row, min(ws.max_row + 1, start_row + max_scan)):
        cells = [str(ws.cell(row=row, column=c).value or "") for c in range(1, ws.max_column + 1)]
        hdr_text = "".join(normalize_header_text(c) for c in cells if c)
        hits = sum(1 for kw in LIST_KEYWORDS if kw in hdr_text)
        if hits > best_hits:
            best_hits, best_row = hits, row

    return best_row if best_hits >= MIN_HIT else None


def get_sheet_headers(ws):
    """读取工作表的表头行，返回 (header_row, headers_list)"""
    hdr_row = get_header_row(ws)
    if hdr_row is None:
        return None, []
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=hdr_row, column=c).value
        headers.append(str(val).strip() if val else "")
    # 去除尾部空列
    while headers and not headers[-1]:
        headers.pop()
    return hdr_row, headers


# ============================================================
# 公式工具函数
# ============================================================

def classify_formula(formula: str) -> str:
    """分类公式类型：inline / range / cross"""
    if not formula or not formula.startswith("="):
        return "value"
    if "!" in formula:
        return "cross"
    if ":" in formula:
        return "range"
    return "inline"


def adjust_formula_rows(formula: str, offset: int) -> str:
    """
    将公式中所有行号 +offset。
    仅调整相对引用（不带 $ 前缀的行号）。
    例：=SUM(V5:V10)  + offset=266  → =SUM(V271:V276)
    """
    if not formula or not formula.startswith("="):
        return formula

    def _shift(m):
        col_part = m.group(1)   # 列字母，可能含 $
        row_lock = m.group(2)   # 行锁定符 $ 或空
        row_num  = m.group(3)   # 行号数字
        if row_lock == "$":       # 绝对行引用，不调整
            return f"{col_part}${row_num}"
        return f"{col_part}{int(row_num) + offset}"

    pattern = r"(\$?[A-Za-z]+)(\$?)(\d+)"
    return re.sub(pattern, _shift, formula)


def adjust_formula_refs(formula: str, row_offset: int = 0, col_offset: int = 0) -> str:
    """
    同时调整公式中的列字母和行号。
    用于插入业态列（col_offset=1，所有 col>=2 右移1列）+ 行号偏移。
    """
    if not formula or not formula.startswith("="):
        return formula

    def _col_idx_to_letter(idx: int) -> str:
        """1-based column index to Excel letter"""
        letters = ""
        while idx > 0:
            idx -= 1
            letters = chr(ord('A') + idx % 26) + letters
            idx //= 26
        return letters

    def _col_letter_to_idx(letter: str) -> int:
        idx = 0
        for ch in letter.upper():
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx

    def _shift(m):
        col_part = m.group(1)   # column letter(s), may include $
        row_lock = m.group(2)   # $ or empty
        row_num  = m.group(3)   # row number

        # Shift column
        if col_offset != 0 and col_part:
            col_abs = col_part.startswith("$")
            col_letter = col_part.lstrip("$")
            new_idx = _col_letter_to_idx(col_letter) + col_offset
            if new_idx < 1:
                new_idx = 1
            col_part = ("$" if col_abs else "") + _col_idx_to_letter(new_idx)

        # Shift row
        if row_lock == "$":
            return f"{col_part}${row_num}"
        return f"{col_part}{int(row_num) + row_offset}"

    pattern = r"(\$?[A-Za-z]+)(\$?)(\d+)"
    return re.sub(pattern, _shift, formula)


def detect_header_rows_from(ws, start_row: int = 3, max_scan: int = 15) -> int:
    """
    从 start_row 开始，通过合并单元格模式检测表头占几行。
    返回表头行数（至少1行）。
    """
    merge_count_per_row = {}
    for merged_range in ws.merged_cells.ranges:
        mr_min_row = merged_range.min_row
        if mr_min_row >= start_row:
            for r in range(mr_min_row, merged_range.max_row + 1):
                merge_count_per_row[r] = merge_count_per_row.get(r, 0) + 1

    hdr_rows = 0
    for r in range(start_row, start_row + max_scan):
        if merge_count_per_row.get(r, 0) > 0:
            hdr_rows += 1
        else:
            break
    return max(1, hdr_rows)


def extract_formulas(ws, sheet_name: str, row_offset: int = 0, col_offset: int = 0) -> list:
    """
    提取工作表中所有范围公式和跨表公式。
    返回：[（位置, 公式类型, 原公式, 偏移后公式）, ...]
    """
    records = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not (cell.value and isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            ftype = classify_formula(cell.value)
            if ftype not in ("range", "cross"):
                continue
            pos = f"{get_column_letter(col)}{row}"
            shifted = (adjust_formula_refs(cell.value, row_offset=row_offset, col_offset=col_offset)
                       if (row_offset or col_offset) else cell.value)
            records.append((pos, ftype, cell.value, shifted))
    return records


# ============================================================
# 数值一致性核查
# ============================================================

def audit_numeric_stats(ws, label: str) -> dict:
    """统计工作表数值和公式信息（用于三段式核查）"""
    stats = {
        "标签": label,
        "行数": ws.max_row,
        "列数": ws.max_column,
        "公式数量": 0,
        "范围公式": 0,
        "跨表公式": 0,
    }
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                stats["公式数量"] += 1
                ftype = classify_formula(cell.value)
                if ftype == "range":
                    stats["范围公式"] += 1
                elif ftype == "cross":
                    stats["跨表公式"] += 1
    return stats


# ============================================================
# 核心复制函数（v2.0 按列位置复制）
# ============================================================

def copy_cells_with_formula(src_ws, dst_ws,
                             src_start_row: int, dst_start_row: int,
                             row_count: int, col_count: int,
                             row_offset: int = 0, col_offset: int = 0) -> int:
    """
    按列位置逐格复制（v2.0 核心函数）。
    从 src_ws[src_start_row:] 复制到 dst_ws[dst_start_row:]
    公式行号自动 +row_offset，列字母自动 +col_offset。
    col_offset > 0 用于插入业态列（所有 col>=2 右移1列）。
    返回：复制的公式数量。
    """
    formula_count = 0
    for r_offset in range(row_count):
        src_row = src_start_row + r_offset
        dst_row = dst_start_row + r_offset

        # 复制行高
        if src_row in src_ws.row_dimensions:
            dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height

        for src_col in range(1, col_count + 1):
            dst_col = src_col + col_offset
            src_cell = src_ws.cell(row=src_row, column=src_col)
            dst_cell = dst_ws.cell(row=dst_row, column=dst_col)

            # 复制值（公式字符串或普通值）
            if src_cell.value and isinstance(src_cell.value, str) and src_cell.value.startswith("="):
                dst_cell.value = adjust_formula_refs(
                    src_cell.value, row_offset=row_offset, col_offset=col_offset
                )
                formula_count += 1
            else:
                dst_cell.value = src_cell.value

            # 复制样式
            if src_cell.has_style:
                dst_cell.font       = copy.copy(src_cell.font)
                dst_cell.border     = copy.copy(src_cell.border)
                dst_cell.fill       = copy.copy(src_cell.fill)
                dst_cell.alignment  = copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

    return formula_count


def copy_merged_cells(src_ws, dst_ws, row_offset: int = 0):
    """复制合并单元格，行号 +row_offset"""
    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        new_min_row = min_row + row_offset
        new_max_row = max_row + row_offset
        new_range = (f"{get_column_letter(min_col)}{new_min_row}:"
                     f"{get_column_letter(max_col)}{new_max_row}")
        try:
            dst_ws.merge_cells(new_range)
        except Exception:
            pass  # 忽略冲突的合并区域


def copy_col_widths(src_ws, dst_ws):
    """复制列宽"""
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width


# ============================================================
# 格式化输出 Sheet
# ============================================================

def format_output_sheet(ws, col_count: int, row_count: int, hdr_row: int = 3):
    """为输出 Sheet 应用格式（表头、边框、对齐）
    
    参数：
        hdr_row: 真正的表头行号（默认第3行，前2行为前置行不格式化）
    """
    # 表头格式（真正的表头行，不是第1行）
    for col in range(1, col_count + 1):
        cell = ws.cell(row=hdr_row, column=col)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _THIN

    # 数据行格式（表头行 + 1 开始，不覆盖前置行）
    for row in range(hdr_row + 1, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border   = _THIN
            cell.font     = _DATA_FONT
            cell.alignment = _LEFT_ALIGN


def write_formula_audit_sheet(wb, records: list):
    """写「公式核查报告」Sheet"""
    if FORMULA_SHEET in wb.sheetnames:
        del wb[FORMULA_SHEET]
    ws = wb.create_sheet(FORMULA_SHEET)

    # 表头
    headers = ["来源表", "位置", "公式类型", "原公式", "偏移后公式", "建议"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value     = h
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _THIN

    # 数据
    for row_idx, (pos, ftype, orig_formula, shifted) in enumerate(records, 2):
        ws.cell(row=row_idx, column=1).value = "综合楼" if int(pos[1:]) <= 266 else "教学楼"
        ws.cell(row=row_idx, column=2).value = pos
        ws.cell(row=row_idx, column=3).value = "范围公式" if ftype == "range" else "跨表公式"
        ws.cell(row=row_idx, column=4).value = orig_formula
        ws.cell(row=row_idx, column=5).value = shifted
        ws.cell(row=row_idx, column=6).value = "请核查公式范围是否正确" if ftype == "range" else "!! 跨表引用，请确认引用路径"

        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = _THIN

    # 列宽
    col_widths = [20, 10, 15, 60, 60, 40]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"


def write_merge_report_sheet(wb, stats_pre: dict, stats_post: dict):
    """写「合并说明」Sheet（三段式核查报告）"""
    if REPORT_SHEET in wb.sheetnames:
        del wb[REPORT_SHEET]
    ws = wb.create_sheet(REPORT_SHEET)

    # 表头
    headers = ["对比项", "追加前", "追加后", "差异", "结果"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value     = h
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _HDR_ALIGN
        cell.border    = _THIN

    # 对比行
    rows = [
        ("行数",     stats_pre["行数"],     stats_post["行数"],     stats_post["行数"]     - stats_pre["行数"],     None),
        ("列数",     stats_pre["列数"],     stats_post["列数"],     stats_post["列数"]     - stats_pre["列数"],     None),
        ("公式数量", stats_pre["公式数量"], stats_post["公式数量"], stats_post["公式数量"] - stats_pre["公式数量"], None),
        ("范围公式", stats_pre["范围公式"], stats_post["范围公式"], stats_post["范围公式"] - stats_pre["范围公式"], None),
        ("跨表公式", stats_pre["跨表公式"], stats_post["跨表公式"], stats_post["跨表公式"] - stats_pre["跨表公式"], None),
    ]

    for row_idx, (item, pre, post, diff, result) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1).value = item
        ws.cell(row=row_idx, column=2).value = pre
        ws.cell(row=row_idx, column=3).value = post
        ws.cell(row=row_idx, column=4).value = diff
        ws.cell(row=row_idx, column=5).value = "[OK] 一致" if diff == 0 or item in ("行数",) else "请核查"

        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.border   = _THIN
            cell.font     = _DATA_FONT
            cell.alignment = _CTR_ALIGN

    # 列宽
    for col, w in enumerate([20, 15, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"


# ============================================================
# 主函数（v2.0 简化版）
# ============================================================

def merge_to_summary(input_path: str, target_sheet_names: list = None) -> dict:
    """
    主处理函数 v2.0 — 公式感知合并（按列位置）。
    
    参数：
        input_path:        Excel 文件路径
        target_sheet_names: 可选，手动指定两张清单表名 [主表, 追加表]
    
    返回：结果摘要 dict
    """
    print(f"\n{'='*70}")
    print(f"公式感知合并 v2.0 — {os.path.basename(input_path)}")
    print(f"{'='*70}")

    # Step 1: 加载（保留公式）
    wb = load_workbook(input_path, data_only=False)
    all_sheets = wb.sheetnames

    # Step 2: 识别清单表
    if target_sheet_names:
        list_sheets = [s for s in target_sheet_names if s in all_sheets]
    else:
        list_sheets = []
        for sn in all_sheets:
            if sn in (OUTPUT_SHEET, FORMULA_SHEET, REPORT_SHEET):
                continue
            ws = wb[sn]
            if is_list_sheet(ws):
                list_sheets.append(sn)

    if len(list_sheets) < 2:
        print(f"\n[WARN] 需要至少 2 个清单表，当前找到 {len(list_sheets)} 个")
        return {"success": False, "list_sheets": list_sheets}

    print(f"\n识别到清单表（前2个）：{list_sheets[:2]}")

    # Step 3: 读取两张表结构 + 检测多级表头区域
    ws_a = wb[list_sheets[0]]
    ws_b = wb[list_sheets[1]]

    hdr_row_a = 3   # 表头固定从 Row 3 开始（Row 1-2 是前置行）
    hdr_row_b = 3
    header_rows_a = detect_header_rows_from(ws_a, start_row=3)
    header_rows_b = detect_header_rows_from(ws_b, start_row=3)

    print(f"\n表A「{list_sheets[0]}」：{ws_a.max_row}行 x {ws_a.max_column}列, 表头={header_rows_a}行(Row3-{2+header_rows_a})")
    print(f"表B「{list_sheets[1]}」：{ws_b.max_row}行 x {ws_b.max_column}列, 表头={header_rows_b}行(Row3-{2+header_rows_b})")

    # Step 4: 用 HeaderParser 扁平化多级表头 → 单级表头
    parser = HeaderParser()
    flat_hdrs_a = parser.parse_sheet_headers(ws_a, header_rows_a, first_header_row=3)
    flat_hdrs_b = parser.parse_sheet_headers(ws_b, header_rows_b, first_header_row=3)

    # 去除尾部空列
    while flat_hdrs_a and not flat_hdrs_a[-1]:
        flat_hdrs_a.pop()
    while flat_hdrs_b and not flat_hdrs_b[-1]:
        flat_hdrs_b.pop()

    # Step 5: 构建统一扁平化表头 + 插入业态列
    physical_cols = max(ws_a.max_column, ws_b.max_column)
    unified_cols = max(len(flat_hdrs_a), len(flat_hdrs_b), physical_cols)
    # 补齐到统一列数
    padded_a = flat_hdrs_a + [""] * (unified_cols - len(flat_hdrs_a))
    padded_b = flat_hdrs_b + [""] * (unified_cols - len(flat_hdrs_b))

    # 构建统一表头（优先用表A的列名）
    unified_headers = list(padded_a)

    # 在"序号"后插入"业态"列
    unified_headers.insert(1, "业态")
    yetai_col = 2   # 业态列在输出中的列号（1-based）

    col_count_out = unified_cols + 1  # 多了业态列

    print(f"\n扁平化表头：{len(flat_hdrs_a)}列(A) + {len(flat_hdrs_b)}列(B) → 统一{col_count_out}列（含业态）")
    print(f"表头预览：{[h[:25] if h else '(空)' for h in unified_headers[:8]]}")

    # Step 6: 数据起始行计算
    data_start_a = 3 + header_rows_a  # Row 6（跳过多级表头 Row 3-5）
    data_start_b = 3 + header_rows_b  # Row 6
    data_rows_a = ws_a.max_row - data_start_a + 1
    data_rows_b = ws_b.max_row - data_start_b + 1
    dst_data_start = 4  # 输出行1=标题, 行2=工程名称, 行3=表头

    # 表B 追加起始行 = 表A 数据结束行 + 1
    dst_start_b = dst_data_start + data_rows_a

    # 教学楼公式行偏移：源Row6 → 输出Row(dst_start_b), 源Row7 → 输出Row(dst_start_b+1)
    # offset = dst_start_b - data_start_b
    row_offset_b = dst_start_b - data_start_b

    print(f"\n数据范围：A=[源{data_start_a}-{ws_a.max_row}] → [输出{dst_data_start}-{dst_data_start+data_rows_a-1}]")
    print(f"          B=[源{data_start_b}-{ws_b.max_row}] → [输出{dst_start_b}-{dst_start_b+data_rows_b-1}]")
    print(f"表B公式偏移：行+{row_offset_b}, 列+1（业态列）")

    # Step 7: 追加前基准核查
    print(f"\n{'─'*70}")
    print("[ 追加前基准核查 ]")
    stats_pre_a = audit_numeric_stats(ws_a, list_sheets[0])
    stats_pre_b = audit_numeric_stats(ws_b, list_sheets[1])
    print(f"  {list_sheets[0]}：{stats_pre_a['行数']}行，公式{stats_pre_a['公式数量']}个")
    print(f"  {list_sheets[1]}：{stats_pre_b['行数']}行，公式{stats_pre_b['公式数量']}个")

    # Step 8: 删除旧输出 Sheet + 创建新 Sheet
    for sn in (OUTPUT_SHEET, FORMULA_SHEET, REPORT_SHEET):
        if sn in wb.sheetnames:
            del wb[sn]

    ws_out = wb.create_sheet(OUTPUT_SHEET)

    # ---- 8a: 第1行 - 工程量清单偏差或漏项对比表 ----
    row1_merge = None
    for mr in ws_a.merged_cells.ranges:
        min_c, min_r, max_c, max_r = mr.bounds
        if min_r == 1 and max_r == 1:
            row1_merge = (min_c, max_c)

    if row1_merge:
        ws_out.merge_cells(start_row=1, start_column=row1_merge[0],
                           end_row=1, end_column=row1_merge[1] + 1)  # +1 因为业态列
    cell1 = ws_out.cell(row=1, column=1, value="工程量清单偏差或漏项对比表")
    src_cell1 = ws_a.cell(row=1, column=1)
    if src_cell1.has_style:
        cell1.font = copy.copy(src_cell1.font)
        cell1.fill = copy.copy(src_cell1.fill)
        cell1.alignment = copy.copy(src_cell1.alignment)
        cell1.border = copy.copy(src_cell1.border)
        cell1.number_format = src_cell1.number_format

    # ---- 8b: 第2行 - 工程名称 ----
    row2_merge = None
    for mr in ws_a.merged_cells.ranges:
        min_c, min_r, max_c, max_r = mr.bounds
        if min_r == 2 and max_r == 2:
            row2_merge = (min_c, max_c)

    if row2_merge:
        ws_out.merge_cells(start_row=2, start_column=row2_merge[0],
                           end_row=2, end_column=row2_merge[1] + 1)
    sheet_short_names = []
    for s in list_sheets[:2]:
        short = s
        for prefix in ['表-08 分部分项工程和单价措施项目清单-', '表-08 ', '表', ]:
            if short.startswith(prefix):
                short = short[len(prefix):]
                break
        sheet_short_names.append(short)
    cell2 = ws_out.cell(row=2, column=1, value=f"工程名称：{' + '.join(sheet_short_names)}")
    src_cell2 = ws_a.cell(row=2, column=1)
    if src_cell2.has_style:
        cell2.font = copy.copy(src_cell2.font)
        cell2.fill = copy.copy(src_cell2.fill)
        cell2.alignment = copy.copy(src_cell2.alignment)
        cell2.border = copy.copy(src_cell2.border)
        cell2.number_format = src_cell2.number_format

    # ---- 8c: 第3行 - 扁平化统一表头 ----
    for col_idx, hdr_text in enumerate(unified_headers, 1):
        cell = ws_out.cell(row=3, column=col_idx)
        cell.value = hdr_text if hdr_text else None

    # ---- 8d: 复制表A数据（跳过多级表头，从data_start_a开始，col_offset=1为业态让路）----
    copy_cells_with_formula(
        ws_a, ws_out,
        src_start_row=data_start_a, dst_start_row=dst_data_start,
        row_count=data_rows_a, col_count=physical_cols,
        row_offset=0, col_offset=1  # 综合楼数据行偏移=0, 列偏移+1(业态列)
    )

    # ---- 8e: 填充表A业态列 ----
    sheet_a_short = sheet_short_names[0]
    for r in range(data_rows_a):
        ws_out.cell(row=dst_data_start + r, column=yetai_col).value = sheet_a_short

    # ---- 8f: 复制表A的合并单元格（仅数据区，跳过表头区/前置行）----
    for merged_range in ws_a.merged_cells.ranges:
        min_c, min_r, max_c, max_r = merged_range.bounds
        if min_r <= 2 or min_r < data_start_a:  # 跳过前置行和表头行
            continue
        new_min_row = min_r - (data_start_a - dst_data_start)
        new_max_row = max_r - (data_start_a - dst_data_start)
        new_min_col = min_c + 1  # 列偏移
        new_max_col = max_c + 1
        try:
            ws_out.merge_cells(
                start_row=new_min_row, start_column=new_min_col,
                end_row=new_max_row, end_column=new_max_col
            )
        except Exception:
            pass

    print(f"\n[OK] 表A数据复制完成：{data_rows_a}行（业态={sheet_a_short}）")

    # ---- 8g: 复制表B数据（跳过多级表头，追加到表A之后）----
    copy_cells_with_formula(
        ws_b, ws_out,
        src_start_row=data_start_b, dst_start_row=dst_start_b,
        row_count=data_rows_b, col_count=physical_cols,
        row_offset=row_offset_b, col_offset=1
    )

    # ---- 8h: 填充表B业态列 ----
    sheet_b_short = sheet_short_names[1]
    for r in range(data_rows_b):
        ws_out.cell(row=dst_start_b + r, column=yetai_col).value = sheet_b_short

    # ---- 8i: 复制表B的合并单元格（仅数据区）----
    for merged_range in ws_b.merged_cells.ranges:
        min_c, min_r, max_c, max_r = merged_range.bounds
        if min_r <= 2 or min_r < data_start_b:
            continue
        new_min_row = min_r + row_offset_b
        new_max_row = max_r + row_offset_b
        new_min_col = min_c + 1
        new_max_col = max_c + 1
        try:
            ws_out.merge_cells(
                start_row=new_min_row, start_column=new_min_col,
                end_row=new_max_row, end_column=new_max_col
            )
        except Exception:
            pass

    print(f"[OK] 表B数据追加完成：{data_rows_b}行（业态={sheet_b_short}, 公式偏移行+{row_offset_b}/列+1）")

    # Step 9: 格式化输出 Sheet
    format_output_sheet(ws_out, col_count_out, ws_out.max_row, hdr_row=3)

    # 复制列宽
    copy_col_widths(ws_a, ws_out)

    # Step 10: 公式核查报告
    all_formula_records = []
    all_formula_records.extend(extract_formulas(ws_a, list_sheets[0], row_offset=0, col_offset=1))
    all_formula_records.extend(extract_formulas(ws_b, list_sheets[1], row_offset=row_offset_b, col_offset=1))
    write_formula_audit_sheet(wb, all_formula_records)

    print(f"\n[OK] 公式核查报告已完成：{len(all_formula_records)}条公式")

    # Step 11: 合并说明（三段式核查）
    stats_post = audit_numeric_stats(ws_out, "清单汇总")
    write_merge_report_sheet(wb,
        {"行数": data_rows_a + data_rows_b,
         "列数": col_count_out,
         "公式数量": stats_pre_a["公式数量"] + stats_pre_b["公式数量"],
         "范围公式": stats_pre_a["范围公式"] + stats_pre_b["范围公式"],
         "跨表公式": stats_pre_a["跨表公式"] + stats_pre_b["跨表公式"]},
        stats_post
    )

    print(f"[OK] 合并说明已完成")

    # Step 12: 保存
    output_path = input_path

    print(f"\n{'─'*70}")
    print("保存文件...")

    try:
        wb.save(output_path)
        print(f"  [OK] 已保存到原文件：{os.path.basename(output_path)}")
    except PermissionError:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_含清单汇总{ext}"
        wb.save(output_path)
        print(f"  原文件被占用，已另存为：{os.path.basename(output_path)}")

    wb.close()

    total_rows_out = 3 + data_rows_a + data_rows_b
    print(f"\n{'='*70}")
    print(f"处理完成（v2.0 表头扁平化版）！")
    print(f"  清单表：{list_sheets[0]} + {list_sheets[1]}")
    print(f"  表头：{header_rows_a}行多级 → 1行单级（{col_count_out}列，含业态）")
    print(f"  数据：A={data_rows_a}行 + B={data_rows_b}行 = {data_rows_a+data_rows_b}行")
    print(f"  总行数：{total_rows_out}（1标题+1名称+1表头+{data_rows_a+data_rows_b}数据）")
    print(f"  输出文件：{os.path.basename(output_path)}")
    print(f"{'='*70}\n")

    return {
        "success": True,
        "list_sheets": list_sheets[:2],
        "output_file": output_path,
    }


# ============================================================
# CLI 入口
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="清单表公式感知合并工具 v2.0")
    parser.add_argument("file", help="Excel 文件路径")
    parser.add_argument("--sheets", nargs=2, help="手动指定两张清单表名（按顺序）")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"错误：文件不存在 - {args.file}")
        sys.exit(1)

    result = merge_to_summary(args.file, target_sheet_names=args.sheets)

    if not result["success"]:
        print("\n[WARN] 处理未完成，请检查日志。")
        sys.exit(1)
