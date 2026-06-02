import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from collections import defaultdict

excel_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测.xlsx"
output_path = r"c:\Users\aa\Desktop\造价清单文件标准化\1.1综合楼_终测_成本分析.xlsx"

print("开始处理Excel文件...")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path)

print(f"原始工作表: {wb.sheetnames}")

df_summary = pd.read_excel(excel_path, sheet_name='清单汇总', header=0)
print(f"\n清单汇总数据形状: {df_summary.shape}")
print(f"列名: {list(df_summary.columns)}")

print("\n数据预览:")
print(df_summary.head(10))

df_clean = df_summary.copy()

df_clean = df_clean[df_clean['项目编码'].notna() & (df_clean['项目编码'] != '')]

print(f"\n清洗后数据形状: {df_clean.shape}")

def safe_float(value):
    try:
        return float(value) if pd.notna(value) else 0.0
    except:
        return 0.0

df_clean['合同合价'] = df_clean['合价（元）_合同合价（元）_J'].apply(safe_float)
df_clean['结算合价'] = df_clean['合价（元）_结算合价（元）_K'].apply(safe_float)
df_clean['审核合价'] = df_clean['合价（元）_审核合价（元）_L'].apply(safe_float)

print("\n成本数据统计:")
print(f"合同合价总和: {df_clean['合同合价'].sum():,.2f}")
print(f"结算合价总和: {df_clean['结算合价'].sum():,.2f}")
print(f"审核合价总和: {df_clean['审核合价'].sum():,.2f}")

print("\n按业态统计:")
summary_by_category = df_clean.groupby('业态').agg({
    '项目编码': 'count',
    '合同合价': 'sum',
    '结算合价': 'sum',
    '审核合价': 'sum'
}).rename(columns={'项目编码': '项目数量'})

summary_by_category['偏差（结算-合同）'] = summary_by_category['结算合价'] - summary_by_category['合同合价']
summary_by_category['偏差率（%）'] = (summary_by_category['偏差（结算-合同）'] / summary_by_category['合同合价'] * 100).round(2)

print(summary_by_category)

df_clean.to_pickle(r"c:\Users\aa\Desktop\造价清单文件标准化\cleaned_data.pkl")
summary_by_category.to_pickle(r"c:\Users\aa\Desktop\造价清单文件标准化\summary_by_category.pkl")

print("\n数据清洗完成！")
print("=" * 80)
